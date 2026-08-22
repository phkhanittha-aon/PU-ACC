# -*- coding: utf-8 -*-
"""สร้าง presentation/MGS-Traceability-Schema.xlsx — โครงสร้างฐานข้อมูลของระบบทวนสอบสินค้าและเรียกคืน

รันด้วย:  python3 tools/build-traceability-schema-xlsx.py   (ต้องมี openpyxl)

ไฟล์นี้อ่านโครงตารางจาก apps-script/traceability/00_Config.gs ตัวจริงผ่าน
tools/dump-traceability-schema.js จึงไม่มีวันเพี้ยนจากโค้ด
"""
import json
import os
import subprocess

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "presentation", "MGS-Traceability-Schema.xlsx")

S = json.loads(subprocess.check_output(
    ["node", os.path.join(ROOT, "tools", "dump-traceability-schema.js")], text=True))

SCHEMA, TAB, SAP_TABS, EN = S["schema"], S["tab"], S["sap_tabs"], S["enums"]

F = "Arial"
C_INK, C_MUTE, C_LINE = "17211D", "48544C", "C9D2CC"
C_SAP_BG, C_SAP_FG = "1A6A5C", "FFFFFF"      # เขียว = กระจกเงา SAP
C_MGS_BG, C_MGS_FG = "96731C", "FFFFFF"      # ทอง  = MGS เป็นเจ้าของ
C_SYS_BG, C_SYS_FG = "48544C", "FFFFFF"      # เทา  = ระบบ
C_DOC_BG = "17211D"
C_NOTE = "FAF3DF"
C_WARN_BG, C_WARN_FG = "FDEAEA", "A02B30"
C_OK_BG, C_OK_FG = "E6F1EA", "2C6E49"

_t = Side(style="thin", color=C_LINE)
BD = Border(left=_t, right=_t, top=_t, bottom=_t)

SYS_TABS = [TAB["USERS"], TAB["AUDIT"], TAB["SYNC_LOG"], TAB["SETTINGS"]]
MGS_TABS = [t for t in SCHEMA if t not in SAP_TABS and t not in SYS_TABS]

# ─────────────────────────────────────────────────────────────────────────────
# คำอธิบายคอลัมน์
#   COMMON = ชื่อที่ใช้ซ้ำหลายตาราง เขียนครั้งเดียว
#   PER    = คอลัมน์ที่ความหมายขึ้นกับตาราง เขียนทับ COMMON
# รูปแบบ: (คำอธิบาย, ที่มา / ฟิลด์ใน SAP B1)
# ─────────────────────────────────────────────────────────────────────────────
COMMON = {
    "doc_entry":   ("รหัสภายในของเอกสารใน B1 (ไม่ใช่เลขที่ที่คนเห็น)", "DocEntry"),
    "line_num":    ("ลำดับบรรทัดในเอกสาร เริ่มที่ 0", "LineNum"),
    "doc_num":     ("เลขที่เอกสารที่คนใช้อ้างอิงกัน", "DocNum"),
    "doc_date":    ("วันที่เอกสาร รูปแบบ ปี-เดือน-วัน", "DocDate"),
    "card_code":   ("รหัสคู่ค้า (ผู้ขายหรือลูกค้า)", "CardCode"),
    "card_name":   ("ชื่อคู่ค้า — เก็บซ้ำไว้เพื่อให้ประวัติไม่เปลี่ยนเมื่อแก้ทะเบียน", "CardName"),
    "item_code":   ("รหัสสินค้า", "ItemCode"),
    "dscription":  ("คำอธิบายสินค้าบนบรรทัดเอกสาร", "Dscription"),
    "quantity":    ("จำนวนบนบรรทัดเอกสาร", "Quantity"),
    "whs_code":    ("รหัสคลัง", "WhsCode"),
    "project":     ("ชื่อโครงการ — ใช้แยกว่าของไปลงหน้างานไหน", "OPRJ.PrjName"),
    "base_entry":  ("DocEntry ของเอกสารต้นทาง ใช้ต่อสายกลับไปหา PO/DO", "BaseEntry"),
    "base_line":   ("บรรทัดของเอกสารต้นทาง", "BaseLine"),
    "base_doc_num":("เลขที่เอกสารต้นทาง (เติมให้แล้วตอนดึงข้อมูล)", "join จากเอกสารต้นทาง"),
    "num_at_card": ("เลขที่เอกสารของคู่ค้า เช่น เลขใบกำกับของผู้ขาย", "NumAtCard"),
    "dist_number": ("เลขล็อตหรือหมายเลขซีเรียล", "OBTN/OSRN.DistNumber"),
    "lot_key":     ("คีย์ยืนของล็อต = item_code + | + dist_number", "ระบบสร้างเอง"),
    "kind":        ("ชนิดการติดตาม: SERIAL หรือ BATCH", "จาก OITM.ManSerNum/ManBtchNum"),
    "is_active":   ("TRUE = ยังใช้งานอยู่", "validFor / Inactive"),
    "synced_at":   ("เวลาที่ข้อมูลแถวนี้ถูกดึงมาล่าสุด", "ระบบเติมเอง"),
    "case_no":     ("เลขที่เคสเรียกคืน เช่น RC-2026-001", "ระบบออกให้"),
    "created_at":  ("เวลาที่สร้างแถวนี้", "ระบบเติมเอง"),
    "updated_by":  ("อีเมลคนที่แก้ล่าสุด", "ระบบเติมเอง"),
    "updated_at":  ("เวลาที่แก้ล่าสุด", "ระบบเติมเอง"),
    "row_version": ("เลขเวอร์ชันของแถว — กันสองคนแก้ทับกัน ห้ามแก้มือ", "ระบบเติมเอง"),
    "status":      ("สถานะ — ดูค่าที่ใช้ได้ในแท็บ _ค่าที่ใช้ได้", "ระบบควบคุม"),
    "remark":      ("หมายเหตุเพิ่มเติม", "คนกรอก"),
    "note":        ("หมายเหตุเพิ่มเติม", "คนกรอก"),
    "evidence_file_ids": ("รหัสไฟล์หลักฐานใน Drive คั่นด้วยจุลภาค", "ระบบเติมเมื่ออัปโหลด"),
    "qty_affected":("จำนวนที่ได้รับผลกระทบ", "ระบบคำนวณจาก SAP"),
    "model":       ("รุ่น / Part No.", "OITM.ItemName"),
}

PER = {
  TAB["ITEMS"]: {
    "item_name": ("ชื่อสินค้า / รุ่น", "ItemName"),
    "item_group": ("กลุ่มสินค้า", "OITB.ItmsGrpNam"),
    "brand": ("ยี่ห้อ — ต้องสร้างเป็น User Field ใน B1", "OITM.U_Brand"),
    "product_type": ("ประเภทสินค้า เช่น Inverter / Rail", "OITM.U_ProductType"),
    "mng_method": ("วิธีติดตาม: SERIAL / BATCH / NONE (NONE จะไม่ขึ้นในระบบ)", "ManSerNum + ManBtchNum"),
    "uom": ("หน่วยนับ", "InvntryUom"),
  },
  TAB["BP"]: {
    "card_type": ("S = ผู้ขาย · C = ลูกค้า", "CardType"),
    "phone": ("เบอร์โทร — ใช้ตอนต้องเรียกคืน", "Phone1"),
    "email": ("อีเมล", "E_Mail"),
    "contact_person": ("ผู้ติดต่อ", "CntctPrsn"),
  },
  TAB["WHS"]: {
    "whs_name": ("ชื่อคลัง", "WhsName"),
    "location_type": ("INTERNAL = คลังเรา · CUSTOMER = คลังหน้างานลูกค้า (ต้องสร้าง User Field)", "OWHS.U_LocationType"),
  },
  TAB["PO"]: {
    "open_qty": ("จำนวนที่ยังไม่ได้รับ", "OpenQty"),
    "doc_status": ("O = เปิดอยู่ · C = ปิดแล้ว", "DocStatus"),
  },
  TAB["DELIVERY"]: {
    "ship_to_code": ("รหัสที่อยู่จัดส่ง", "ShipToCode"),
    "address": ("ที่อยู่จัดส่ง", "Address2"),
  },
  TAB["RETURN"]: {
    "return_type": ("CUSTOMER_RETURN = ลูกค้าคืนเรา · SUPPLIER_RETURN = เราคืนผู้ขาย", "แยกจากตารางต้นทาง"),
    "reason": ("เหตุผลการคืน", "Comments"),
  },
  TAB["LOTS"]: {
    "sys_number": ("รหัสภายในของล็อต ใช้ join กับตารางเดินของ", "AbsEntry"),
    "mnf_serial": ("ซีเรียลของผู้ผลิต (ต่างจากซีเรียลที่ MGS ใช้)", "MnfSerial"),
    "supplier_lot": ("เลขล็อตของผู้ขาย — ใช้ตอนเคลมกับผู้ขาย", "LotNumber"),
    "in_date": ("วันที่รับเข้า", "InDate"),
    "exp_date": ("วันหมดอายุ (ถ้ามี)", "ExpDate"),
    "notes": ("หมายเหตุที่ติดมากับล็อตใน B1", "Notes"),
  },
  TAB["MOVES"]: {
    "move_id": ("คีย์ไม่ซ้ำของการเคลื่อนไหว 1 ครั้ง", "ระบบสร้างจาก obj_type~doc_entry~line~lot~direction"),
    "obj_type": ("ชนิดเอกสารตามรหัสของ B1 — ดูตารางในแท็บ _ค่าที่ใช้ได้", "OITL.ApplyType"),
    "direction": ("IN = ของเข้า · OUT = ของออก (quantity เป็นบวกเสมอ)", "OITL.Direction 0/1"),
    "quantity": ("จำนวน — เป็นค่าบวกเสมอ ทิศทางดูที่ direction เท่านั้น", "ABS(ITL1.Quantity)"),
    "doc_entry": ("DocEntry ของเอกสารที่ทำให้ของเคลื่อน", "OITL.ApplyEntry"),
    "line_num": ("บรรทัดของเอกสารนั้น", "OITL.ApplyLine"),
  },
  TAB["STOCK"]: {
    "quantity": ("จำนวนคงเหลือของล็อตนี้ในคลังนี้", "OIBT.Quantity / OSRI"),
    "status": ("AVAILABLE = พร้อมจ่าย", "OSRI.Status"),
  },
  TAB["TRACE_NOTES"]: {
    "note_id": ("รหัสไม่ซ้ำของหมายเหตุ", "ระบบสร้าง"),
    "mgs_receiving_lot": ("เลขล็อตรับเข้าที่ MGS ตั้งเอง เช่น MGS-INV-260807-01", "QC กรอก"),
    "qc_status": ("ผลตรวจของ QC ต่อล็อตนี้", "QC กรอก"),
    "note": ("สิ่งที่คนอ่านทีหลังต้องรู้", "QC กรอก"),
    "created_by": ("อีเมลคนที่สร้าง", "ระบบเติมเอง"),
  },
  TAB["CASES"]: {
    "case_id": ("รหัสภายในของเคส", "ระบบสร้าง"),
    "status": ("DRAFT/OPEN/CONTAINED/TRACKING/VERIFYING/CLOSED/CANCELLED", "ระบบควบคุม"),
    "opened_at": ("เวลาที่เปิดเคส", "ระบบเติมเอง"),
    "opened_by": ("อีเมลคนเปิดเคส", "ระบบเติมเอง"),
    "case_owner": ("ผู้รับผิดชอบเคส", "คนกรอก"),
    "source": ("ที่มา: ผู้ขายแจ้ง / ลูกค้าร้องเรียน / QC พบเอง / พบหน้างาน / หน่วยงานกำกับ", "คนเลือก"),
    "source_ref": ("เลขที่อ้างอิง เช่น เลขหนังสือแจ้งของผู้ขาย", "คนกรอก"),
    "problem": ("ปัญหาที่พบ — ข้อความนี้ถูกส่งเข้ากลุ่ม Lark", "คนกรอก (ขั้นต่ำ 10 ตัวอักษร)"),
    "risk_class": ("CRITICAL / MAJOR / MINOR — กำหนด SLA และการแจ้งผู้บริหาร", "คนเลือก"),
    "immediate_action": ("สิ่งที่ต้องทำทันที", "คนกรอก"),
    "escalated": ("TRUE = ต้องแจ้งผู้บริหาร (ตั้งอัตโนมัติจากระดับความเสี่ยง)", "ระบบตั้ง"),
    "escalated_at": ("เวลาที่แจ้งผู้บริหาร", "ระบบเติมเอง"),
    "recall_required": ("สรุปการตัดสินใจว่าต้องเรียกคืนหรือไม่", "สะท้อนจาก Recall_Actions"),
    "field_action": ("สรุปการแก้ไข/เปลี่ยนหน้างาน", "สะท้อนจาก Recall_Actions"),
    "customer_notify_required": ("สรุปว่าต้องแจ้งลูกค้าหรือไม่", "สะท้อนจาก Recall_Actions"),
    "supplier_claim_required": ("สรุปว่าต้องเคลมผู้ขายหรือไม่", "สะท้อนจาก Recall_Actions"),
    "qty_affected": ("จำนวนที่กระทบทั้งหมด", "★ ระบบคำนวณจาก SAP"),
    "qty_in_stock": ("จำนวนที่ยังอยู่ในคลัง", "★ ระบบคำนวณจาก SAP"),
    "qty_delivered": ("จำนวนที่ส่งลูกค้าไปแล้ว", "★ ระบบคำนวณจาก SAP"),
    "qty_unaccounted": ("★ จำนวนที่ระบุที่อยู่ไม่ได้ — ตัวเลขสำคัญที่สุดในระบบ ต้องเป็น 0", "★ ระบบคำนวณ"),
    "qty_returned": ("รวมจำนวนที่รับคืนแล้ว", "ระบบรวมจาก Recall_Tracking"),
    "qty_replaced": ("รวมจำนวนที่เปลี่ยนของใหม่แล้ว", "ระบบรวมจาก Recall_Tracking"),
    "qty_corrected": ("รวมจำนวนที่ตรวจ/แก้ไขแล้ว", "ระบบรวมจาก Recall_Tracking"),
    "effectiveness_pct": ("% ประสิทธิผล = (คืน+เปลี่ยน+แก้ไข) ÷ กระทบ", "ระบบคำนวณ"),
    "closed_at": ("เวลาที่ปิดเคส", "ระบบเติมเอง"),
    "closed_by": ("อีเมลคนปิดเคส", "ระบบเติมเอง"),
    "closure_note": ("สรุปผลการปิดเคส", "คนกรอก (ขั้นต่ำ 10 ตัวอักษร)"),
    "drive_folder_id": ("โฟลเดอร์เก็บหลักฐานของเคสนี้", "ระบบสร้าง"),
  },
  TAB["SCOPE"]: {
    "scope_id": ("รหัสไม่ซ้ำของขอบเขต 1 รายการ", "ระบบสร้าง"),
    "sn_from": ("ซีเรียลเริ่มต้นที่ผู้ใช้ระบุ", "คนกรอก"),
    "sn_to": ("ซีเรียลสิ้นสุดที่ผู้ใช้ระบุ", "คนกรอก"),
    "qty_affected": ("จำนวนที่กระทบของล็อตนี้ (ซีเรียล = 1 เสมอ)", "ระบบคำนวณ"),
    "added_by": ("อีเมลคนที่เพิ่มขอบเขตนี้", "ระบบเติมเอง"),
    "added_at": ("เวลาที่เพิ่ม", "ระบบเติมเอง"),
    "note": ("หมายเหตุของขอบเขตนี้", "คนกรอก"),
  },
  TAB["TRACKING"]: {
    "track_id": ("รหัสไม่ซ้ำของแถวติดตาม", "ระบบสร้าง"),
    "sn_display": ("ช่วงซีเรียลแบบย่อสำหรับอ่าน เช่น A006–A012 (7 หน่วย)", "ระบบสร้าง"),
    "location_type": ("WAREHOUSE = คลังเรา · CUSTOMER = ลูกค้า · CUSTOMER_SITE = คลังหน้างานลูกค้า", "ระบบจัดกลุ่ม"),
    "party_code": ("รหัสลูกค้า หรือรหัสคลัง", "ระบบเติมจาก SAP"),
    "party_name": ("ชื่อลูกค้า หรือชื่อคลัง", "ระบบเติมจาก SAP"),
    "doc_type": ("ชนิดเอกสารที่ส่งของออกไป หรือ STOCK ถ้ายังอยู่ในคลัง", "ระบบเติม"),
    "contacted_at": ("เวลาที่ติดต่อปลายทางครั้งแรก", "ระบบเติมเมื่อเปลี่ยนสถานะ"),
    "contacted_by": ("อีเมลคนที่ติดต่อ", "ระบบเติมเอง"),
    "required_action": ("HOLD/RETURN/REPLACE/INSPECT_CORRECT/SCRAP/NO_ACTION", "คนเลือก"),
    "qty_returned": ("จำนวนที่รับคืนแล้วจากปลายทางนี้", "คนกรอก"),
    "qty_replaced": ("จำนวนที่เปลี่ยนของใหม่แล้ว", "คนกรอก"),
    "qty_corrected": ("จำนวนที่ตรวจ/แก้ไขหน้างานแล้ว", "คนกรอก"),
    "qty_pending": ("จำนวนที่ยังค้าง = กระทบ − (คืน+เปลี่ยน+แก้ไข)", "ระบบคำนวณ"),
    "status": ("PENDING/CONTACTED/IN_PROGRESS/COMPLETED/NOT_REACHED", "คนเลือก + ระบบตรวจ"),
    "remark": ("หลักฐานหรือบันทึกการติดต่อ", "คนกรอก"),
  },
  TAB["ACTIONS"]: {
    "action_id": ("รหัสไม่ซ้ำของหัวข้อ", "ระบบสร้าง"),
    "section_type": ("INVESTIGATION = สอบสวน/CAPA · DECISION = การตัดสินใจ", "ระบบสร้างตอนเปิดเคส"),
    "section": ("รหัสหัวข้อ เช่น ROOT_CAUSE — ดูรายการในแท็บ _ค่าที่ใช้ได้", "ระบบสร้างตอนเปิดเคส"),
    "seq": ("ลำดับการแสดงผล", "ระบบตั้ง"),
    "details": ("เนื้อหาของหัวข้อ — ต้องกรอกก่อนตั้งสถานะเป็น DONE", "คนกรอก"),
    "responsible": ("ผู้รับผิดชอบ", "คนกรอก"),
    "target_date": ("วันที่กำหนดเสร็จ (ปี-เดือน-วัน)", "คนกรอก"),
    "status": ("OPEN / IN_PROGRESS / DONE / NA", "คนเลือก"),
    "approved_by": ("อีเมลผู้อนุมัติ — ต้องคนละคนกับผู้กรอก", "ระบบเติมเอง"),
    "approved_at": ("เวลาที่อนุมัติ", "ระบบเติมเอง"),
  },
  TAB["HOLDS"]: {
    "hold_id": ("รหัสไม่ซ้ำของการกัก", "ระบบสร้าง"),
    "qty_hold": ("จำนวนที่ถูกกัก", "ระบบคำนวณจากคงคลัง"),
    "status": ("HOLD = กักอยู่ · RELEASED = ปลดแล้ว", "ระบบควบคุม"),
    "placed_by": ("อีเมลคนที่สั่งกัก", "ระบบเติมเอง"),
    "placed_at": ("เวลาที่กัก", "ระบบเติมเอง"),
    "released_by": ("อีเมลคนที่ปลด — เฉพาะผู้จัดการคุณภาพ", "ระบบเติมเอง"),
    "released_at": ("เวลาที่ปลด", "ระบบเติมเอง"),
    "release_reason": ("เหตุผลที่ปลด — บังคับกรอกอย่างน้อย 5 ตัวอักษร", "คนกรอก"),
  },
  TAB["MOCK"]: {
    "test_id": ("รหัสภายในของการทดสอบ", "ระบบสร้าง"),
    "test_no": ("เลขที่การทดสอบ เช่น MR-2026-001", "ระบบออกให้"),
    "test_date": ("วันที่ทดสอบ", "ระบบเติมเอง"),
    "conducted_by": ("อีเมลผู้ทดสอบ", "ระบบเติมเอง"),
    "reviewed_by": ("อีเมลผู้ทบทวน — ต้องคนละคนกับผู้ทดสอบ", "ระบบเติมเอง"),
    "test_type": ("BACKWARD / FORWARD / BOTH", "คนเลือก"),
    "started_at": ("เวลาเริ่ม — ระบบจับเวลาให้", "ระบบเติมเอง"),
    "ended_at": ("เวลาปิดผล", "ระบบเติมเอง"),
    "duration_min": ("เวลาที่ใช้จริง (นาที)", "ระบบคำนวณ"),
    "target_min": ("เวลาเป้าหมาย (นาที) ค่าตั้งต้น 120", "ตั้งได้ในแท็บ Settings"),
    "qty_affected": ("จำนวนรวมที่ควรหาเจอ", "ระบบคำนวณจาก SAP"),
    "qty_located": ("จำนวนรวมที่หาเจอจริง", "รวมจาก Mock_Recall_Lines"),
    "completion_pct": ("% ที่หาเจอ — ต้อง 100% จึงผ่าน", "ระบบคำนวณ"),
    "reconcile_pct": ("% การกระทบยอด — ต้อง 100% จึงผ่าน", "ระบบคำนวณ"),
    "result": ("IN_PROGRESS / PASS / FAIL", "ระบบตัดสิน"),
    "gap_found": ("สาเหตุที่ไม่ผ่าน — ระบบเขียนให้เอง", "ระบบเติม"),
    "root_cause": ("สาเหตุที่แท้จริง — บังคับกรอกเมื่อไม่ผ่าน", "ผู้ทบทวนกรอก"),
    "corrective_action": ("การแก้ไข — บังคับกรอกเมื่อไม่ผ่าน", "ผู้ทบทวนกรอก"),
    "capa_owner": ("ผู้รับผิดชอบการแก้ไข", "ผู้ทบทวนกรอก"),
    "capa_due": ("วันที่กำหนดแล้วเสร็จ", "ผู้ทบทวนกรอก"),
    "capa_status": ("OPEN = ยังแก้ไม่เสร็จ · NA = ไม่ต้องแก้", "ระบบตั้ง"),
    "verified_by": ("อีเมลผู้ยืนยันผล", "ระบบเติมเอง"),
    "verified_at": ("เวลาที่ยืนยัน", "ระบบเติมเอง"),
  },
  TAB["MOCK_LINES"]: {
    "line_id": ("รหัสไม่ซ้ำของบรรทัด", "ระบบสร้าง"),
    "test_no": ("เลขที่การทดสอบที่บรรทัดนี้สังกัด", "ระบบเติม"),
    "leg": ("BACKWARD = ตามย้อนกลับ · FORWARD = ตามไปข้างหน้า", "ระบบสร้าง"),
    "seq": ("ลำดับ", "ระบบตั้ง"),
    "party_or_location": ("ปลายทางหรือต้นทางที่ต้องไปยืนยัน", "ระบบเติมจาก SAP"),
    "qty_expected": ("จำนวนที่ควรหาเจอ", "★ ระบบคำนวณจาก SAP"),
    "qty_located": ("จำนวนที่หาเจอจริง", "★ ผู้ทดสอบกรอก"),
    "result": ("PENDING / PASS / PARTIAL / FAIL", "ระบบตัดสินจากตัวเลข"),
    "evidence": ("หลักฐานที่ใช้ยืนยัน — บังคับกรอกถ้าหาเจอ", "ผู้ทดสอบกรอก"),
  },
  TAB["USERS"]: {
    "email": ("อีเมลบริษัท — ต้องตรงกับที่ใช้ล็อกอิน Google", "แอดมินกรอก"),
    "full_name": ("ชื่อที่แสดงในระบบ", "แอดมินกรอก"),
    "dept": ("แผนก", "แอดมินกรอก"),
    "role": ("QC / QCM / WH / SR / LS / GM / ADMIN / VIEWER", "แอดมินกรอก"),
    "lark_user_id": ("รหัสผู้ใช้ Lark (ถ้าจะแจ้งเตือนรายบุคคล)", "แอดมินกรอก"),
    "is_active": ("FALSE = ปิดการใช้งาน (ใช้ตอนพนักงานลาออก)", "แอดมินกรอก"),
    "note": ("หมายเหตุ", "แอดมินกรอก"),
  },
  TAB["AUDIT"]: {
    "at_ts": ("เวลาที่เกิดการกระทำ", "ระบบเติมเอง"),
    "actor_email": ("อีเมลคนที่ทำ", "ระบบเติมเอง"),
    "action": ("ชื่อการกระทำ เช่น OPEN_CASE / RELEASE_HOLD", "ระบบเติมเอง"),
    "entity": ("ตารางที่ถูกกระทำ", "ระบบเติมเอง"),
    "entity_id": ("รหัสของรายการที่ถูกกระทำ", "ระบบเติมเอง"),
    "before_json": ("ค่าก่อนแก้ (JSON) — ใช้ตอนมีข้อโต้แย้ง", "ระบบเติมเอง"),
    "after_json": ("ค่าหลังแก้ (JSON)", "ระบบเติมเอง"),
    "app_version": ("เวอร์ชันแอปตอนนั้น", "ระบบเติมเอง"),
    "request_id": ("clientKey ของคำขอ ใช้ไล่ปัญหาการกดซ้ำ", "ระบบเติมเอง"),
  },
  TAB["SYNC_LOG"]: {
    "at_ts": ("เวลาที่ sync", "ระบบเติมเอง"),
    "source": ("ที่มาของรอบนี้ เช่น PUSH", "ระบบเติมเอง"),
    "tab": ("แท็บที่ถูกเขียน", "ระบบเติมเอง"),
    "rows_in": ("จำนวนแถวที่ส่งมา", "ระบบเติมเอง"),
    "rows_written": ("จำนวนแถวที่เขียนสำเร็จ", "ระบบเติมเอง"),
    "status": ("OK / ERROR / REJECTED", "ระบบเติมเอง"),
    "duration_ms": ("เวลาที่ใช้ (มิลลิวินาที)", "ระบบเติมเอง"),
    "message": ("รายละเอียดหรือสาเหตุที่ล้มเหลว", "ระบบเติมเอง"),
    "batch_id": ("รหัสรอบการส่ง ใช้จับคู่หลายแท็บในรอบเดียวกัน", "ตัวส่งข้อมูลสร้าง"),
  },
  TAB["SETTINGS"]: {
    "key": ("ชื่อค่าตั้งค่า", "แอดมินกรอก"),
    "value": ("ค่า", "แอดมินกรอก"),
    "note": ("คำอธิบายว่าค่านี้มีผลกับอะไร", "แอดมินกรอก"),
  },
}


def desc(tab, col):
    if tab in PER and col in PER[tab]:
        return PER[tab][col]
    if col in COMMON:
        return COMMON[col]
    return ("—", "—")


# ─────────────────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)


def style_header(ws, row, ncols, bg, fg, height=26):
    ws.row_dimensions[row].height = height
    for c in range(1, ncols + 1):
        x = ws.cell(row=row, column=c)
        x.font = Font(name=F, bold=True, size=10, color=fg)
        x.fill = PatternFill("solid", fgColor=bg)
        x.alignment = Alignment(vertical="center", wrap_text=True)
        x.border = BD


def put(ws, r, c, v, bg=None, fg=C_INK, bold=False, size=10, wrap=True, al="left", italic=False):
    x = ws.cell(row=r, column=c, value=v)
    x.font = Font(name=F, size=size, bold=bold, italic=italic, color=fg)
    if bg:
        x.fill = PatternFill("solid", fgColor=bg)
    x.alignment = Alignment(vertical="top", wrap_text=wrap, horizontal=al)
    x.border = BD
    return x


def widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── แท็บที่ 1: อ่านก่อน ──────────────────────────────────────────────────────
ws = wb.create_sheet("_อ่านก่อน")
widths(ws, [4, 30, 96])
ws.merge_cells("B2:C2")
put(ws, 2, 2, "MGS — โครงสร้างฐานข้อมูลระบบทวนสอบสินค้าและเรียกคืน", bold=True, size=16, wrap=False)
ws.row_dimensions[2].height = 28
ws.merge_cells("B3:C3")
put(ws, 3, 2, "เวอร์ชันโครงสร้าง %s · %d ตาราง · %d คอลัมน์ · สร้างจาก apps-script/traceability/00_Config.gs โดยตรง"
    % (S["app_version"], len(SCHEMA), sum(len(v) for v in SCHEMA.values())),
    fg=C_MUTE, italic=True, wrap=False)

rows = [
    ("ไฟล์นี้คืออะไร",
     "โครงสร้างตารางทั้งหมดของฐานข้อมูล เอาไว้ดู ตรวจ และคุยกันก่อนสร้างชีตจริง\n"
     "แท็บที่ขึ้นต้นด้วย _ คือเอกสารประกอบ ไม่ใช่ตารางข้อมูล"),
    ("วิธีใช้ แบบที่ 1 — แนะนำ",
     "ให้ Apps Script สร้างชีตให้เอง: เปิดโปรเจกต์ Apps Script แล้วรัน setupCreateWorkbook()\n"
     "จะได้ครบทั้ง 23 แท็บ พร้อม header ที่ตรงกับโค้ด 100% และตั้งค่า frozen row ให้แล้ว\n"
     "วิธีนี้ไม่มีทางพิมพ์ชื่อคอลัมน์ผิด — ใช้วิธีนี้ถ้าไม่มีเหตุผลอื่น"),
    ("วิธีใช้ แบบที่ 2",
     "อยากเห็นและแก้โครงสร้างก่อน: เปิด Google Sheets ใหม่ > ไฟล์ > นำเข้า > อัปโหลดไฟล์นี้\n"
     "เลือก 'แทนที่สเปรดชีต' แล้วลบแท็บที่ขึ้นต้นด้วย _ ออกก่อนใช้งานจริง\n"
     "จากนั้นเอา ID ของชีต (ส่วนกลางของ URL) ไปใส่ใน Script Properties ชื่อ SS_ID"),
    ("ตรวจว่าโครงสร้างตรงกับโค้ด",
     "หลังสร้างชีตแล้ว รัน runSelfTest() ใน Apps Script\n"
     "มันจะเทียบ header ของทุกแท็บกับโค้ดทีละคอลัมน์ และบอกทันทีว่าคอลัมน์ไหนไม่ตรง"),
    ("⚠️ ห้ามแชร์ชีตนี้ให้ผู้ใช้ทั่วไป",
     "ถ้าผู้ใช้เปิดชีตตรงได้ กฎตรวจสอบทั้ง 13 ข้อและระบบสิทธิ์ทั้งหมดเป็นโมฆะทันที\n"
     "ทุกคนต้องเข้าผ่านเว็บแอปเท่านั้น — ให้สิทธิ์ชีตเฉพาะเจ้าของโปรเจกต์กับแอดมิน"),
    ("⚠️ ห้ามแทรกหรือสลับคอลัมน์",
     "โค้ดอ้างชื่อคอลัมน์จากแถวแรกเสมอ ไม่ได้อ้างตำแหน่ง แต่ถ้าแก้ 'ชื่อ' คอลัมน์เมื่อไร ระบบพัง\n"
     "จะเพิ่มคอลัมน์ ให้เพิ่มใน 00_Config.gs ก่อน แล้วรัน setupCreateWorkbook() ซ้ำ"),
    ("แท็บสีเขียว = กระจกเงา SAP",
     "11 แท็บนี้ถูกเขียนทับทั้งชุดทุก 30 นาทีจาก SAP Business One\n"
     "อย่าเก็บอะไรที่คนกรอกไว้ในนี้ รอบ sync ถัดไปจะทับหายทันที"),
    ("แท็บสีทอง = MGS เป็นเจ้าของ",
     "8 แท็บนี้คือข้อมูลที่คนกรอกและระบบคำนวณ รอบ sync ไม่แตะเลย"),
    ("แท็บสีเทา = ระบบ",
     "Users คือทะเบียนสิทธิ์ · Audit_Log เขียนเพิ่มอย่างเดียว ห้ามแก้/ลบ ·\n"
     "Sync_Log ดูสุขภาพการเชื่อม SAP · Settings แก้ค่าได้โดยไม่ต้องแก้โค้ด"),
    ("★ คอลัมน์ที่สำคัญที่สุด",
     "Recall_Cases.qty_unaccounted — จำนวนที่ระบบระบุที่อยู่ไม่ได้\n"
     "ถ้าไม่เป็น 0 แปลว่าถ้าเกิดเรียกคืนจริง จะตามของไม่เจอเท่านั้นชิ้น\n"
     "และเป็นเครื่องวัดวินัยการคีย์ล็อตของคลังไปในตัว"),
]
r = 5
style_header(ws, r, 3, C_DOC_BG, "FFFFFF")
put(ws, r, 2, "หัวข้อ", bg=C_DOC_BG, fg="FFFFFF", bold=True)
put(ws, r, 3, "รายละเอียด", bg=C_DOC_BG, fg="FFFFFF", bold=True)
r += 1
for k, v in rows:
    warn = k.startswith("⚠️")
    star = k.startswith("★")
    bg = C_WARN_BG if warn else (C_NOTE if star else None)
    fg = C_WARN_FG if warn else C_INK
    put(ws, r, 2, k, bg=bg, fg=fg, bold=True)
    put(ws, r, 3, v, bg=bg, fg=fg)
    ws.row_dimensions[r].height = 15 * (v.count("\n") + 1) + 6
    r += 1
ws.sheet_view.showGridLines = False

# ── แท็บที่ 2: พจนานุกรมข้อมูล ───────────────────────────────────────────────
ws = wb.create_sheet("_พจนานุกรมข้อมูล")
widths(ws, [22, 9, 26, 62, 34])
ws.merge_cells("A1:E1")
put(ws, 1, 1, "พจนานุกรมข้อมูล — ทุกคอลัมน์ของทุกตาราง", bold=True, size=14, wrap=False)
ws.row_dimensions[1].height = 24
ws.merge_cells("A2:E2")
put(ws, 2, 1, "คอลัมน์ที่มี ★ คือคอลัมน์ที่ระบบคำนวณให้ ไม่ใช่ให้คนกรอก", fg=C_MUTE, italic=True, wrap=False)

head = ["ตาราง", "ลำดับ", "ชื่อคอลัมน์", "ความหมาย", "ที่มา / ฟิลด์ใน SAP B1"]
r = 4
for i, h in enumerate(head, start=1):
    put(ws, r, i, h, bg=C_DOC_BG, fg="FFFFFF", bold=True)
style_header(ws, r, 5, C_DOC_BG, "FFFFFF")
r += 1

order = SAP_TABS + MGS_TABS + SYS_TABS
for tab in order:
    grp_bg = C_SAP_BG if tab in SAP_TABS else (C_MGS_BG if tab in MGS_TABS else C_SYS_BG)
    for i, col in enumerate(SCHEMA[tab], start=1):
        d, src = desc(tab, col)
        put(ws, r, 1, tab if i == 1 else "", bg=(grp_bg if i == 1 else None),
            fg=("FFFFFF" if i == 1 else C_INK), bold=(i == 1))
        put(ws, r, 2, i, al="center")
        put(ws, r, 3, col, bold=True)
        put(ws, r, 4, d, bg=(C_NOTE if d.startswith("★") else None))
        put(ws, r, 5, src, fg=C_MUTE)
        r += 1
ws.freeze_panes = "A5"
ws.auto_filter.ref = "A4:E%d" % (r - 1)

# ── แท็บที่ 3: ค่าที่ใช้ได้ ──────────────────────────────────────────────────
ws = wb.create_sheet("_ค่าที่ใช้ได้")
widths(ws, [30, 26, 60, 30])
ws.merge_cells("A1:D1")
put(ws, 1, 1, "ค่าที่ระบบยอมรับ — กรอกนอกเหนือจากนี้ระบบจะปฏิเสธ", bold=True, size=14, wrap=False)
ws.row_dimensions[1].height = 24
r = 3


def enum_block(title, pairs, note=""):
    global r
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    put(ws, r, 1, title + ("   —   " + note if note else ""), bg=C_DOC_BG, fg="FFFFFF", bold=True, wrap=False)
    ws.row_dimensions[r].height = 22
    r += 1
    put(ws, r, 1, "ใช้ที่", bg="EAECE9", bold=True)
    put(ws, r, 2, "ค่า", bg="EAECE9", bold=True)
    put(ws, r, 3, "ความหมาย", bg="EAECE9", bold=True)
    put(ws, r, 4, "หมายเหตุ", bg="EAECE9", bold=True)
    r += 1
    for where, val, mean, extra in pairs:
        put(ws, r, 1, where, fg=C_MUTE)
        put(ws, r, 2, val, bold=True)
        put(ws, r, 3, mean)
        put(ws, r, 4, extra, fg=C_MUTE)
        r += 1
    r += 1


enum_block("สถานะเคสเรียกคืน", [
    ("Recall_Cases.status", k, v["th"],
     ("ไปต่อได้: " + ", ".join(v["next"])) if v["next"] else "สถานะสุดท้าย")
    for k, v in EN["CASE_STATUS"].items()
], "เดินหน้าอย่างเดียว ปิดแล้วเปิดใหม่ไม่ได้")

enum_block("สถานะการติดตาม", [
    ("Recall_Tracking.status", k, v, "") for k, v in EN["TRACK_STATUS"].items()
], "ปิดเป็น COMPLETED ได้ต่อเมื่อจำนวนครบเท่านั้น")

enum_block("สิ่งที่ต้องทำกับของที่กระทบ", [
    ("Recall_Tracking.required_action", k, v, "") for k, v in EN["REQUIRED_ACTION"].items()
])

enum_block("ระดับความเสี่ยง", [
    ("Recall_Cases.risk_class", k, v["th"],
     "SLA %s ชม. · %s" % (v["sla_hours"], "แจ้งผู้บริหาร" if v["escalate"] else "ไม่ต้องแจ้งผู้บริหาร"))
    for k, v in EN["RISK_CLASS"].items()
])

enum_block("ที่มาของเคส", [
    ("Recall_Cases.source", k, v, "") for k, v in EN["CASE_SOURCE"].items()
])

enum_block("หัวข้อสอบสวนและการตัดสินใจ", [
    ("Recall_Actions.section", s["key"], s["th"],
     "สอบสวน/CAPA" if s["type"] == "INVESTIGATION" else "การตัดสินใจ")
    for s in EN["ACTION_SECTIONS"]
], "ระบบสร้างครบทุกหัวข้อให้อัตโนมัติตอนเปิดเคส")

enum_block("บทบาทผู้ใช้", [
    ("Users.role", k, v, "") for k, v in EN["ROLES"].items()
], "สิทธิ์ตรวจที่เซิร์ฟเวอร์ทุกครั้ง การซ่อนปุ่มเป็นแค่ UX")

enum_block("ชนิดเอกสารของ SAP B1 (obj_type)", [
    ("SAP_Lot_Moves.obj_type", k, v["name"],
     ("ของเข้า" if v["dir"] == "IN" else "ของออก" if v["dir"] == "OUT"
      else "ทั้งเข้าและออก" if v["dir"] == "BOTH" else "ไม่กระทบสต๊อก")
     + (" · นับเป็นส่งถึงลูกค้า" if k in EN["OUTBOUND_TO_CUSTOMER"] else ""))
    for k, v in EN["OBJ"].items()
], "มาจาก OITL.ApplyType")

enum_block("วิธีติดตามสินค้า", [
    ("SAP_Items.mng_method", k, {"SERIAL": "ติดตามรายเครื่องด้วยหมายเลขซีเรียล",
                                 "BATCH": "ติดตามเป็นล็อต",
                                 "NONE": "ไม่ติดตาม — สินค้าแบบนี้จะไม่ขึ้นในระบบ"}[k], "")
    for k in EN["MNG"]
])
ws.freeze_panes = "A3"

# ── แท็บข้อมูลจริง 23 แท็บ ───────────────────────────────────────────────────
GROUP_NOTE = {
    "SAP": "แท็บนี้ถูกเขียนทับทั้งชุดทุกรอบ sync จาก SAP B1 — ห้ามเก็บข้อมูลที่คนกรอกไว้ที่นี่",
    "MGS": "แท็บนี้เป็นข้อมูลที่ MGS เป็นเจ้าของ รอบ sync ไม่แตะ — แก้ผ่านเว็บแอปเท่านั้น",
    "SYS": "แท็บระบบ",
}

for tab in order:
    if tab in SAP_TABS:
        bg, fg, grp = C_SAP_BG, C_SAP_FG, "SAP"
    elif tab in MGS_TABS:
        bg, fg, grp = C_MGS_BG, C_MGS_FG, "MGS"
    else:
        bg, fg, grp = C_SYS_BG, C_SYS_FG, "SYS"

    ws = wb.create_sheet(tab)
    cols = SCHEMA[tab]

    # แถวที่ 1 ต้องเป็น header จริง เพราะโค้ดอ่านแถวแรกเป็นสัญญา
    for i, c in enumerate(cols, start=1):
        put(ws, 1, i, c, bg=bg, fg=fg, bold=True, wrap=False)
        ws.column_dimensions[get_column_letter(i)].width = max(14, min(30, len(c) + 6))
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # คำอธิบายไปไว้เป็น cell comment เพื่อไม่ให้ปนกับข้อมูล
    for i, c in enumerate(cols, start=1):
        d, src = desc(tab, c)
        cell = ws.cell(row=1, column=i)
        cell.comment = openpyxl.comments.Comment(
            "%s\n\nที่มา: %s" % (d, src), "MGS Schema", width=320, height=110)

    note_row = 3
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row,
                   end_column=max(3, min(len(cols), 8)))
    put(ws, note_row, 1,
        "หมายเหตุ: %s   ·   ลบแถวที่ %d นี้ทิ้งก่อนใช้งานจริง (ชี้เมาส์ที่ชื่อคอลัมน์เพื่อดูคำอธิบาย)"
        % (GROUP_NOTE[grp], note_row),
        bg=C_NOTE, fg=C_MUTE, italic=True, wrap=False)

wb.save(OUT)
print("สร้าง %s แล้ว" % os.path.relpath(OUT, ROOT))
print("  แท็บเอกสาร 3 · แท็บข้อมูล %d (SAP %d · MGS %d · ระบบ %d)"
      % (len(SCHEMA), len(SAP_TABS), len(MGS_TABS), len(SYS_TABS)))
print("  คอลัมน์รวม %d" % sum(len(v) for v in SCHEMA.values()))
