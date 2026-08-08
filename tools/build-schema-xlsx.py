# -*- coding: utf-8 -*-
"""สร้าง MGS-DocHub-Schema.xlsx — โครงสร้างหลังบ้านทั้งหมด แก้ไขได้
อ้างอิง docs/02, 04, 06, 07, 08, 09, 11 ของโปรเจกต์ PU-ACC"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

import os
# ไฟล์ปลายทางอยู่ใน repo เดียวกันเสมอ ไม่ผูกกับเครื่องใคร
OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                   "presentation", "MGS-DocHub-Schema.xlsx")
F = "Arial"

C_HD_BG, C_HD_FG = "1F3D33", "FFFFFF"
C_EDIT = "FFF7D6"   # เหลือง = ทีมแก้ได้
C_LOCK = "F2F3F1"   # เทา    = อ้างอิง อย่าแก้
C_NEW  = "E8F3FF"   # ฟ้า    = เพิ่มใหม่ล่าสุด
C_WARN_BG, C_WARN_FG = "FDEAEA", "A02B30"
C_GOOD_BG, C_GOOD_FG = "E6F1EA", "2C6E49"
C_TITLE, C_MUTE, C_LINE = "17211D", "48544C", "C9D2CC"

_t = Side(style="thin", color=C_LINE)
BD = Border(left=_t, right=_t, top=_t, bottom=_t)

wb = openpyxl.Workbook()
wb.remove(wb.active)

def hdr(ws, row, ncols, bg=C_HD_BG, fg=C_HD_FG, h=30):
    ws.row_dimensions[row].height = h
    for c in range(1, ncols + 1):
        x = ws.cell(row=row, column=c)
        x.font = Font(name=F, bold=True, color=fg, size=10)
        x.fill = PatternFill("solid", fgColor=bg)
        x.alignment = Alignment(vertical="center", wrap_text=True)
        x.border = BD

def widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def cell(ws, r, c, v, bg=None, fg=C_TITLE, bold=False, wrap=True, size=10, al="left", it=False):
    x = ws.cell(row=r, column=c, value=v)
    x.font = Font(name=F, size=size, bold=bold, italic=it, color=fg)
    if bg:
        x.fill = PatternFill("solid", fgColor=bg)
    x.alignment = Alignment(vertical="top", wrap_text=wrap, horizontal=al)
    x.border = BD
    return x

def title(ws, t, sub, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    x = ws.cell(row=1, column=1, value=t)
    x.font = Font(name=F, bold=True, size=15, color=C_TITLE)
    ws.row_dimensions[1].height = 26
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    y = ws.cell(row=2, column=1, value=sub)
    y.font = Font(name=F, italic=True, size=10, color=C_MUTE)
    ws.row_dimensions[2].height = 30
    ws.cell(row=2, column=1).alignment = Alignment(vertical="center", wrap_text=True)

# ---------- LISTS (hidden) ----------
LISTS = {
    "DTYPE": ["UUID", "TEXT", "NUMBER", "DATE", "DATETIME", "BOOLEAN", "ENUM", "FILE_ID", "JSON"],
    "YN": ["บังคับ", "ไม่บังคับ"],
    "SRC": ["ดึงจาก B1", "คนกรอก", "ระบบสร้าง/คำนวณ", "คัดลอกจากเอกสารต้นทาง"],
    "DEPT": ["SR", "QC", "LS", "WH", "AC", "GM", "IT/แอดมิน", "ระบบ"],
    "PHASE": ["BUY", "RECV", "AP", "PAY"],
    "SEV": ["บล็อก", "เตือน", "บันทึกเฉย ๆ"],
    "RULEST": ["ใช้งาน", "รอทีมยืนยันเกณฑ์", "ปิดใช้งาน"],
    "TBLST": ["ยืนยันแล้ว", "รอทีมตรวจ", "เพิ่มใหม่ล่าสุด"],
    "ISSUEST": ["รอตัดสินใจ", "ตัดสินใจแล้ว", "ไม่เกี่ยวข้องแล้ว"],
    "PRIO": ["สูง", "กลาง", "ต่ำ"],
    "TIER": ["1 ในเว็บ", "2 Lark", "3 อีเมล", "2 Lark + 3 อีเมล", "1 + 2", "1 + 2 + 3"],
}
lst = wb.create_sheet("รายการเลือก (ระบบ)")
RANGE = {}
for i, (k, vals) in enumerate(LISTS.items(), start=1):
    lst.cell(row=1, column=i, value=k).font = Font(name=F, bold=True)
    for j, v in enumerate(vals, start=2):
        lst.cell(row=j, column=i, value=v)
    L = get_column_letter(i)
    RANGE[k] = "'รายการเลือก (ระบบ)'!$%s$2:$%s$%d" % (L, L, len(vals) + 1)

def dv(ws, key, rng):
    d = DataValidation(type="list", formula1=RANGE[key], allow_blank=True, showDropDown=False)
    ws.add_data_validation(d)
    d.add(rng)

# ============================================================
# 1. เริ่มที่นี่
# ============================================================
ws = wb.create_sheet("เริ่มที่นี่")
title(ws, "MGS Document Hub — โครงสร้างระบบหลังบ้าน",
      "ไฟล์นี้คือโครงสร้างหลังบ้านทั้งหมดในที่เดียว สำหรับให้ทีมตรวจและแก้ก่อนเริ่มสร้างจริง · "
      "แก้ในไฟล์นี้ได้เลย แล้วส่งกลับมา ระบบจะสร้างตามไฟล์นี้", 6)
widths(ws, [34, 62, 26, 18, 18, 18])

r = 4
cell(ws, r, 1, "สีในไฟล์นี้หมายถึงอะไร", bold=True, size=12); r += 1
for bg, lab, desc in [
    (C_EDIT, "เหลือง = แก้ได้", "ช่องที่ตั้งใจให้ทีมแก้ — เกณฑ์ SLA อัตราภาษี วงเงินอนุมัติ ชื่อฟิลด์ ฯลฯ"),
    (C_LOCK, "เทา = อ้างอิง", "ค่าที่ระบบใช้เป็นคีย์หรือคำนวณเอง แก้ได้แต่ต้องบอกก่อน เพราะกระทบโค้ด"),
    (C_NEW,  "ฟ้า = เพิ่มใหม่ล่าสุด", "ตาราง Price_Requests / Quotations / Payment_Requests / Doc_Bundles และคอลัมน์เรื่องซื้อเงินสด-ไดรฟ์กลาง-รวมไฟล์ ที่เพิ่งเพิ่ม ยังไม่ผ่านตาทีม"),
    (C_WARN_BG, "แดง = ต้องตัดสินใจ", "ตัวเลขที่ผมตั้งค่าเริ่มต้นไว้เอง ต้องให้ทีมยืนยันก่อนขึ้นระบบ"),
]:
    cell(ws, r, 1, lab, bg=bg, bold=True)
    cell(ws, r, 2, desc)
    ws.row_dimensions[r].height = 28
    r += 1

r += 1
cell(ws, r, 1, "แต่ละแท็บใช้ทำอะไร", bold=True, size=12); r += 1
hdr(ws, r, 3); cell(ws, r, 1, "แท็บ", bg=C_HD_BG, fg=C_HD_FG, bold=True)
cell(ws, r, 2, "ใช้ทำอะไร", bg=C_HD_BG, fg=C_HD_FG, bold=True)
cell(ws, r, 3, "ใครควรตรวจ", bg=C_HD_BG, fg=C_HD_FG, bold=True)
r += 1
TABS = [
    ("สารบัญตาราง", "รายชื่อตารางทั้งหมด หน้าที่ และจำนวนคอลัมน์ (นับอัตโนมัติ)", "IT + หัวหน้าทุกแผนก"),
    ("พจนานุกรมข้อมูล", "ทุกคอลัมน์ของทุกตาราง — ชนิด บังคับไหม มาจากไหน", "IT + แผนกเจ้าของข้อมูล"),
    ("ขั้นตอนงาน", "17 ขั้นตอน เจ้าของ SLA เฟส — แก้ SLA ได้ที่นี่", "หัวหน้าทุกแผนก"),
    ("เอกสารที่ต้องมี", "เอกสารบังคับต่อ PO ใครเป็นเจ้าของใบไหน", "SR / QC / WH / AC"),
    ("กฎตรวจสอบ", "กฎที่ระบบใช้บล็อก/เตือนก่อนบันทึก — แก้เกณฑ์ได้ (ดูจำนวนจริงในสรุปด้านล่าง)", "AC เป็นหลัก"),
    ("สิทธิ์แต่ละแผนก", "ใครสร้าง/ดู/แก้/อนุมัติอะไรได้", "หัวหน้าทุกแผนก + ผู้ตรวจสอบ"),
    ("การแจ้งเตือน", "เหตุการณ์ไหนแจ้งใคร ช่องทางไหน เมื่อไหร่", "หัวหน้าทุกแผนก"),
    ("สถานะ", "สถานะทั้งหมดของแต่ละเอกสาร และเปลี่ยนไปไหนได้บ้าง", "IT + AC"),
    ("ค่าตั้งต้น", "ตัวเลขทั้งหมดที่ปรับได้อยู่ในแท็บเดียว", "AC + หัวหน้าทุกแผนก"),
    ("เชื่อมต่อ B1", "ฟิลด์ไหนดึงจากตารางไหนของ Business One", "ผู้ดูแล B1"),
    ("ประเด็นค้าง", "สิ่งที่ยังต้องตัดสินใจก่อนขึ้นระบบ พร้อมผู้รับผิดชอบ", "ทุกคน — ดูแท็บนี้ก่อน"),
]
for t, u, who in TABS:
    cell(ws, r, 1, t, bold=True); cell(ws, r, 2, u); cell(ws, r, 3, who)
    ws.row_dimensions[r].height = 26
    r += 1

r += 1
cell(ws, r, 1, "สรุปตัวเลข ณ ตอนนี้", bold=True, size=12); r += 1
SUMMARY_ROW = r
stats = [
    ("จำนวนตารางทั้งหมด", '=COUNTA(สารบัญตาราง!$B$5:$B$22)'),
    ("จำนวนคอลัมน์ทั้งหมด", '=COUNTA(พจนานุกรมข้อมูล!$B$5:$B$347)'),
    ("ขั้นตอนงาน", '=COUNTA(ขั้นตอนงาน!$B$5:$B$21)'),
    ("เอกสารบังคับต่อ PO", '=COUNTIF(เอกสารที่ต้องมี!$E$5:$E$13,"บังคับ")'),
    ("กฎตรวจสอบ", '=COUNTA(กฎตรวจสอบ!$A$5:$A$34)'),
    ("ค่าตั้งต้นที่ปรับได้", '=COUNTA(ค่าตั้งต้น!$A$5:$A$39)'),
    ("ประเด็นที่ยังรอตัดสินใจ", '=COUNTIF(ประเด็นค้าง!$G$5:$G$39,"รอตัดสินใจ")'),
    ("ประเด็นที่ตัดสินใจแล้ว", '=COUNTIF(ประเด็นค้าง!$G$5:$G$39,"ตัดสินใจแล้ว")'),
]
SUMMARY_LABELS = [lab for lab, _ in stats]
for lab, formula in stats:
    cell(ws, r, 1, lab, bold=True)
    x = cell(ws, r, 2, formula, bg=C_LOCK, al="left")
    x.font = Font(name=F, size=11, bold=True, color=C_TITLE)
    r += 1
START_SHEET = ws

r += 1
cell(ws, r, 1, "วิธีใช้ไฟล์นี้", bold=True, size=12); r += 1
for step in [
    "1. เปิดแท็บ «ประเด็นค้าง» ก่อน — เป็นรายการที่ผมตั้งค่าเริ่มต้นไว้เองและต้องให้ทีมยืนยัน",
    "2. แต่ละแผนกดูแท็บที่เกี่ยวกับตัวเอง (ดูตารางด้านบนว่าใครควรตรวจแท็บไหน)",
    "3. แก้ได้เฉพาะช่องพื้นเหลือง · ถ้าจะแก้ช่องเทา ให้เขียนหมายเหตุไว้ในคอลัมน์สุดท้ายของแถวนั้น",
    "4. ช่องไหนไม่เห็นด้วย ให้พิมพ์ทับได้เลย ไม่ต้องกลัวผิด — ไฟล์นี้มีไว้ให้แก้",
    "5. ส่งไฟล์กลับมา ระบบจะถูกสร้างตามไฟล์นี้ ไม่ใช่ตามเอกสารเดิม",
]:
    c = cell(ws, r, 1, step)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.row_dimensions[r].height = 20
    r += 1

r += 1
w = cell(ws, r, 1, "สำคัญ: ไฟล์นี้เป็น «ต้นฉบับ» ของโครงสร้างหลังบ้าน — ที่ตกลงกันในไฟล์นี้คือสิ่งที่จะถูกสร้าง "
                    "ถ้าแก้ในไฟล์นี้แล้วไม่ตรงกับเอกสาร .md ให้ยึดไฟล์นี้",
         bg=C_WARN_BG, fg=C_WARN_FG, bold=True)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
ws.row_dimensions[r].height = 32
ws.sheet_view.showGridLines = False

# ============================================================
# 2. สารบัญตาราง
# ============================================================
ws = wb.create_sheet("สารบัญตาราง")
title(ws, "สารบัญตาราง", "ตารางทั้งหมดในฐานข้อมูล — คอลัมน์ «จำนวนคอลัมน์» นับจากแท็บพจนานุกรมข้อมูลอัตโนมัติ", 7)
widths(ws, [5, 26, 46, 15, 15, 16, 40])
H = ["#", "ตาราง", "หน้าที่", "จำนวนคอลัมน์", "โตเร็วแค่ไหน", "สถานะ", "หมายเหตุ / ที่ทีมอยากแก้"]
for i, h in enumerate(H, 1):
    cell(ws, 4, i, h)
hdr(ws, 4, 7)

DICT_LAST_ROW = 4 + 390  # จำนวนแถวในพจนานุกรมข้อมูล (ตรวจซ้ำอัตโนมัติท้ายสคริปต์)

TABLES = [
    ("Purchase_Orders", "1 แถว = 1 PO ตั้งแต่เปิดจนปิดบัญชี", "~600 แถว/ปี", "ยืนยันแล้ว", ""),
    ("PO_Payments", "งวดจ่ายของแต่ละ PO (1 PO จ่ายได้หลายงวด)", "~900 แถว/ปี", "ยืนยันแล้ว", ""),
    ("Doc_Checklist", "เอกสารที่ต้องมีต่อ PO และสถานะแต่ละใบ", "~4,800 แถว/ปี", "ยืนยันแล้ว", ""),
    ("Documents", "ไฟล์จริงทุกเวอร์ชันใน Drive", "~5,000 แถว/ปี", "ยืนยันแล้ว", ""),
    ("Approvals", "การอนุมัติและลงนามทุกจุด (PO / จ่าย / เคลม / รหัสสินค้า)", "~2,000 แถว/ปี", "ยืนยันแล้ว", ""),
    ("Stage_Log", "บันทึกทุกการเปลี่ยนสถานะ — เขียนเพิ่มเท่านั้น", "~10,200 แถว/ปี", "ยืนยันแล้ว",
     "ที่มาของสถานะ คอขวด และ KPI ทุกตัว · แยกแท็บต่อปี"),
    ("Notifications", "คิวแจ้งเตือน Lark / อีเมล", "~25,000 แถว/ปี", "ยืนยันแล้ว", "แยกแท็บต่อปี"),
    ("Issues", "ปัญหาที่ผู้ใช้กดแจ้งจากหน้าจอ", "~200 แถว/ปี", "ยืนยันแล้ว", ""),
    ("Claims", "ใบเคลม/ร้องเรียน ตามฟอร์ม QC-F006", "~60 แถว/ปี", "รอทีมตรวจ",
     "ยังรอฟิลด์จริงของ «ส่วนที่ 3» ในฟอร์ม (หน้า 2–3 เป็นภาพสแกน)"),
    ("ItemCode_Requests", "ใบขอเปิดรหัสสินค้าใหม่", "~80 แถว/ปี", "รอทีมตรวจ",
     "รอดูว่ามีฟอร์มควบคุมเดิมอยู่แล้วหรือไม่"),
    ("Price_Requests", "ใบขอราคา — จุดเริ่มก่อนมี PO", "~700 แถว/ปี", "เพิ่มใหม่ล่าสุด",
     "เพิ่มรอบล่าสุด แก้ปัญหาใบขอราคาดึงไปทำใบเสนอราคาไม่ได้"),
    ("Quotations", "ใบเสนอราคาจากผู้ขาย (1 ใบขอราคา → หลายใบ)", "~1,800 แถว/ปี", "เพิ่มใหม่ล่าสุด",
     "เพิ่มรอบล่าสุด"),
    ("Stage_Templates", "นิยามขั้นตอนต่อ module — เพิ่ม module = เพิ่มแถว", "~20 แถว/module", "ยืนยันแล้ว", ""),
    ("Doc_Templates", "นิยามเอกสารที่ต้องมีต่อ module", "~10 แถว/module", "ยืนยันแล้ว", ""),
    ("Users", "ผู้ใช้ แผนก สิทธิ์ และ lark_user_id", "~40 แถว", "รอทีมตรวจ",
     "ต้องมี lark_user_id ครบทุกคน เพราะใช้ Lark ช่องทางเดียว"),
    ("Suppliers", "ผู้ขาย — บางมาก เพราะ master จริงอยู่ใน B1", "~150 แถว", "ยืนยันแล้ว", ""),
    ("Config / Lookups", "ค่าตั้งต้นและรายการ dropdown", "~120 แถว", "ยืนยันแล้ว",
     "จริง ๆ เป็น 2 แท็บ (Config กับ Lookups) นับรวมเป็น 1 รายการ"),
    ("Sync_Log", "บันทึกทุกรอบที่ดึงข้อมูลจาก B1", "~17,500 แถว/ปี", "ยืนยันแล้ว", "แยกแท็บต่อปี"),
    ("Payment_Requests", "คำขอทำจ่ายที่จัดซื้อส่งให้หัวหน้า (1 งวดขอได้หลายรอบ)", "~1,100 แถว/ปี",
     "เพิ่มใหม่ล่าสุด", "เพิ่มรอบล่าสุด — ยอดเงินและ Effective date เป็นช่องบังคับ"),
    ("Doc_Bundles", "ชุดเอกสารรวมเล่มที่สร้างอัตโนมัติเมื่อปิดรายการ", "~660 แถว/ปี",
     "เพิ่มใหม่ล่าสุด", "เพิ่มรอบล่าสุด — 1 รายการ = 1 ชุด"),
]
r = 5
for i, (name, purpose, growth, status, note) in enumerate(TABLES, start=1):
    bg = C_NEW if status == "เพิ่มใหม่ล่าสุด" else None
    cell(ws, r, 1, i, al="center", bg=bg)
    cell(ws, r, 2, name, bold=True, bg=bg)
    cell(ws, r, 3, purpose, bg=bg)
    cell(ws, r, 4, '=COUNTIF(พจนานุกรมข้อมูล!$A$5:$A$%d,$B%d)' % (DICT_LAST_ROW, r), bg=C_LOCK, al="center")
    cell(ws, r, 5, growth, bg=bg, al="center")
    cell(ws, r, 6, status, bg=C_EDIT)
    cell(ws, r, 7, note, bg=C_EDIT)
    ws.row_dimensions[r].height = 30
    r += 1
dv(ws, "TBLST", "F5:F%d" % (r - 1))
ws.freeze_panes = "A5"
ws.auto_filter.ref = "A4:G%d" % (r - 1)

note_r = r + 1
c = cell(ws, note_r, 2, "หมายเหตุ: เอกสาร .md ระบุจำนวนตารางไม่ตรงกัน (07 บอก 15, 09 บอก 16, 11 บอก 17) "
                        "เพราะเขียนคนละรอบ — ตัวเลขจริงคือ 20 รายการตามตารางนี้ (เพิ่ม Payment_Requests และ Doc_Bundles รอบล่าสุด) ให้ยึดไฟล์นี้",
         bg=C_WARN_BG, fg=C_WARN_FG)
ws.merge_cells(start_row=note_r, start_column=2, end_row=note_r, end_column=7)
ws.row_dimensions[note_r].height = 30

# ============================================================
# 3. พจนานุกรมข้อมูล
# ============================================================
ws = wb.create_sheet("พจนานุกรมข้อมูล")
title(ws, "พจนานุกรมข้อมูล", "ทุกคอลัมน์ของทุกตาราง · «มาจากไหน» สำคัญที่สุด — บอกว่าคนต้องพิมพ์เองหรือระบบดึงให้", 9)
widths(ws, [22, 26, 12, 11, 8, 30, 44, 20, 30])
H = ["ตาราง", "คอลัมน์", "ชนิด", "บังคับ", "คีย์", "ค่าที่เป็นไปได้ / รูปแบบ", "คำอธิบาย", "มาจากไหน", "หมายเหตุของทีม"]
for i, h in enumerate(H, 1):
    cell(ws, 4, i, h)
hdr(ws, 4, 9)

B1, USR, SYS, CPY = "ดึงจาก B1", "คนกรอก", "ระบบสร้าง/คำนวณ", "คัดลอกจากเอกสารต้นทาง"
REQ, OPT = "บังคับ", "ไม่บังคับ"

D = []  # (table, col, type, req, key, values, desc, source)
def add(t, rows):
    for row in rows:
        D.append((t,) + row)

add("Purchase_Orders", [
    ("po_id", "UUID", REQ, "PK", "UUID", "คีย์ภายใน ไม่แสดงให้ผู้ใช้เห็น", SYS),
    ("po_no", "TEXT", REQ, "UQ", "PO-26-0042", "เลขที่ผู้ใช้เห็นและใช้อ้างอิงกันทุกแผนก", SYS),
    ("module", "ENUM", REQ, "", "DOMESTIC_FOOD / IMPORT_FOOD / SOLAR", "ตัวกำหนดว่าใช้ template ขั้นตอนชุดไหน", SYS),
    ("sap_doc_entry", "NUMBER", OPT, "FK", "OPOR.DocEntry", "คีย์จริงของ PO ใน B1 — ใช้จับคู่ตอนดึงข้อมูล", B1),
    ("sap_doc_num", "TEXT", OPT, "", "OPOR.DocNum เช่น 10245", "เลข PO ที่คนอ่าน (ต่างจาก DocEntry)", B1),
    ("supplier_code", "TEXT", OPT, "FK", "OCRD.CardCode", "รหัสผู้ขาย", B1),
    ("supplier_name", "TEXT", OPT, "", "OCRD.CardName", "เก็บชื่อซ้ำไว้ไม่ให้ประวัติเปลี่ยนเมื่อ master เปลี่ยน", B1),
    ("item_summary", "TEXT", REQ, "", "กุ้งแช่แข็ง 16/20", "คำอธิบายสั้นให้จำ PO ได้", CPY),
    ("qty_summary", "TEXT", OPT, "", "500 KG", "จำนวนรวมแบบอ่านง่าย", B1),
    ("currency", "TEXT", REQ, "", "THB", "สกุลเงินของ PO", B1),
    ("total_amount", "NUMBER", REQ, "", "130000.00", "มูลค่า PO", B1),
    ("payment_terms_code", "TEXT", OPT, "FK", "OCTG.GroupNum", "เงื่อนไขชำระจาก B1 — ใช้สร้างงวดจ่ายอัตโนมัติ", B1),
    ("payment_terms_name", "TEXT", OPT, "", "เครดิต 30 วัน", "ชื่อเงื่อนไขชำระแบบอ่านง่าย", B1),
    ("owner_email", "TEXT", REQ, "FK", "Users.email", "SR ผู้ดูแล PO นี้", USR),
    ("effective_date", "DATE", OPT, "", "", "วันที่มีผล ที่เดิม SR ต้องแจ้ง AC ทางเมล", USR),
    ("expected_delivery_date", "DATE", OPT, "", "", "วันนัดรับของ", USR),
    ("current_stage", "TEXT", REQ, "", "QC_INCOMING", "ขั้นตอนที่ค้างอยู่ตอนนี้", SYS),
    ("current_owner_email", "TEXT", REQ, "FK", "Users.email", "ใครถืองานอยู่ตอนนี้ — หัวใจของหน้าหลัก", SYS),
    ("stage_due_at", "DATETIME", OPT, "", "", "กำหนดเสร็จของขั้นปัจจุบัน คำนวณจาก SLA", SYS),
    ("status", "ENUM", REQ, "", "ACTIVE / COMPLETED / CANCELLED", "ใช้แยกแท็บกำลังดำเนินการ/เสร็จสิ้น/ยกเลิก", SYS),
    ("has_open_claim", "BOOLEAN", REQ, "", "TRUE / FALSE", "ถ้า TRUE = ปิด PO ไม่ได้ และงวดจ่ายถูกล็อก", SYS),
    ("doc_complete", "BOOLEAN", REQ, "", "TRUE / FALSE", "เอกสารบังคับครบหรือยัง", SYS),
    ("match_ready", "BOOLEAN", REQ, "", "TRUE / FALSE", "ตรวจ 3 ทางผ่านหรือยัง", SYS),
    ("match_diff", "NUMBER", OPT, "", "2000.00", "ส่วนต่างที่ทำให้ 3 ทางไม่ผ่าน", SYS),
    ("match_block", "TEXT", OPT, "", "ยอดใบแจ้งหนี้เกิน PO", "เหตุผลที่ยังปิดไม่ได้ แสดงให้ AC เห็น", SYS),
    ("from_quote_no", "TEXT", OPT, "FK", "Quotations.quote_no", "ถ้าเปิดจากใบเสนอราคา เก็บเลขไว้ตามรอยกลับ", SYS),
    ("drive_folder_id", "FILE_ID", REQ, "", "", "โฟลเดอร์ของรายการนี้ในไดรฟ์กลาง MGS-Documents", SYS),
    ("entry_type", "ENUM", REQ, "", "PO / CASH", "PO = เปิด PO ใน B1 · CASH = ซื้อเงินสดไม่มี PO", SYS),
    ("cash_reason", "TEXT", OPT, "", "ร้านไม่รับเครดิต ต้องจ่ายสดหน้างาน",
     "บังคับเมื่อ entry_type = CASH — เหตุผลที่ไม่เปิด PO (กฎ C1)", USR),
    ("cash_paid_from", "ENUM", OPT, "", "PETTY / OWN / TRANSFER",
     "จ่ายจากเงินสดย่อย / พนักงานสำรองจ่าย / โอนจากบัญชีบริษัท", USR),
    ("qc_bypassed", "BOOLEAN", REQ, "", "TRUE / FALSE", "ล็อตนี้ข้ามการตรวจรับของ QC หรือไม่", USR),
    ("qc_bypass_reason", "TEXT", OPT, "", "ผักสดตรวจหน้างานแล้ว",
     "บังคับเมื่อ qc_bypassed = TRUE (กฎ C3) — แก้ย้อนหลังไม่ได้", USR),
    ("qc_bypass_by", "TEXT", OPT, "FK", "Users.email", "ใครเป็นผู้อนุมัติให้ข้าม QC", SYS),
    ("qc_bypass_at", "DATETIME", OPT, "", "", "เวลาที่กดยืนยันการข้าม QC", SYS),
    ("bundle_doc_id", "UUID", OPT, "FK", "Documents.doc_id", "ไฟล์รวมเล่ม — ว่างจนกว่าจะปิดรายการ", SYS),
    ("bundle_at", "DATETIME", OPT, "", "", "เวลาที่รวมเล่มเสร็จ", SYS),
    ("created_by", "TEXT", REQ, "", "", "ผู้สร้าง", SYS),
    ("created_at", "DATETIME", REQ, "", "", "เวลาสร้าง", SYS),
    ("updated_at", "DATETIME", REQ, "", "", "เวลาแก้ล่าสุด", SYS),
    ("row_version", "NUMBER", REQ, "", "1, 2, 3...", "กันสองคนแก้ทับกันเงียบ ๆ", SYS),
])

add("PO_Payments", [
    ("payment_id", "UUID", REQ, "PK", "UUID", "คีย์ภายใน", SYS),
    ("po_no", "TEXT", REQ, "FK", "Purchase_Orders.po_no", "PO ที่งวดนี้สังกัด", SYS),
    ("seq", "NUMBER", REQ, "", "1, 2, 3", "งวดที่เท่าไหร่", SYS),
    ("payment_type", "ENUM", REQ, "", "DEPOSIT / BALANCE / FULL / PARTIAL", "ประเภทงวด", SYS),
    ("percent", "NUMBER", OPT, "", "30", "สัดส่วนของยอด PO", SYS),
    ("amount", "NUMBER", REQ, "", "39000.00", "ยอดตามงวด", SYS),
    ("currency", "TEXT", REQ, "", "THB", "สกุลเงิน", B1),
    ("due_date", "DATE", REQ, "", "", "วันครบกำหนด คำนวณจาก OCTG ของ B1", SYS),
    ("payment_date", "DATE", OPT, "", "", "วันที่ทำรายการจ่ายจริง", USR),
    ("value_date", "DATE", OPT, "", "", "วันที่ธนาคารตัดเงินจริง — ตัวกำหนดงวดบัญชี", USR),
    ("posting_period", "TEXT", OPT, "", "2026-08", "งวดบัญชี คำนวณจาก value_date", SYS),
    ("days_late", "NUMBER", OPT, "", "-2, 0, 5", "จ่ายช้ากี่วัน (ติดลบ = จ่ายก่อนกำหนด)", SYS),
    ("method", "ENUM", OPT, "", "TRANSFER / CHEQUE / SWIFT / CASH", "วิธีจ่าย", USR),
    ("bank_account_code", "TEXT", OPT, "FK", "Config", "บัญชีที่จ่ายออก เลือกจากรายการ ไม่ให้พิมพ์", USR),
    ("bank_ref", "TEXT", OPT, "UQ", "KB26083000456", "เลขอ้างอิงธนาคาร — ห้ามซ้ำ เพราะเป็นตัวกันจ่ายซ้ำ", USR),
    ("cheque_no", "TEXT", OPT, "", "", "เลขเช็ค ถ้าจ่ายด้วยเช็ค", USR),
    ("wht_type", "ENUM", OPT, "", "GOODS / FREIGHT / SERVICE / RENT / ADS", "ประเภทรายจ่ายสำหรับหัก ณ ที่จ่าย", USR),
    ("wht_rate", "NUMBER", OPT, "", "0, 1, 3, 5", "อัตราหัก ณ ที่จ่าย (%)", SYS),
    ("wht_amount", "NUMBER", OPT, "", "0.00", "จำนวนภาษีหัก ณ ที่จ่าย", USR),
    ("wht_cert_no", "TEXT", OPT, "", "", "เลขที่หนังสือรับรองหัก ณ ที่จ่าย", USR),
    ("fee_amount", "NUMBER", OPT, "", "35.00", "ค่าธรรมเนียมโอน", USR),
    ("fee_borne_by", "ENUM", OPT, "", "OUR / SUPPLIER", "ใครรับภาระค่าธรรมเนียม", USR),
    ("paid_amount", "NUMBER", OPT, "", "91035.00", "ยอดที่ตัดจากบัญชีจริง", USR),
    ("net_to_supplier", "NUMBER", OPT, "", "91000.00", "ยอดที่ผู้ขายได้รับ", SYS),
    ("variance_amount", "NUMBER", OPT, "", "0.00", "ส่วนต่างจากที่ระบบคำนวณ", SYS),
    ("variance_reason", "TEXT", OPT, "", "", "เหตุผลของส่วนต่าง — บังคับเมื่อต่างเกิน 1 บาท", USR),
    ("fx_rate", "NUMBER", OPT, "", "35.20", "อัตราแลกเปลี่ยน ถ้าจ่ายต่างประเทศ", USR),
    ("amount_thb", "NUMBER", OPT, "", "", "ยอดแปลงเป็นบาท", SYS),
    ("status", "ENUM", REQ, "", "PENDING / BLOCKED / REQUESTED / APPROVED / PAID / FAILED / VOID / CANCELLED",
     "สถานะงวด — ดูแท็บ «สถานะ»", SYS),
    ("blocked_by_claim_no", "TEXT", OPT, "FK", "Claims.claim_no", "ใบเคลมที่ทำให้งวดนี้ถูกล็อก", SYS),
    ("requested_by", "TEXT", OPT, "FK", "Users.email", "ผู้ตั้งเรื่องขอจ่าย", SYS),
    ("requested_at", "DATETIME", OPT, "", "", "เวลาตั้งเรื่อง", SYS),
    ("approved_by", "TEXT", OPT, "FK", "Users.email", "ผู้อนุมัติ — ต้องคนละคนกับผู้ตั้งเรื่อง", SYS),
    ("approved_at", "DATETIME", OPT, "", "", "เวลาอนุมัติ", SYS),
    ("reject_reason", "TEXT", OPT, "", "", "เหตุผลที่ตีกลับ", USR),
    ("paid_by", "TEXT", OPT, "FK", "Users.email", "ผู้บันทึกจ่าย — ต้องคนละคนกับผู้อนุมัติ", SYS),
    ("paid_at", "DATETIME", OPT, "", "", "เวลาที่บันทึกจ่ายในระบบ", SYS),
    ("slip_doc_id", "FILE_ID", OPT, "FK", "Documents.doc_id", "สลิป/Swift ของงวดนี้", USR),
    ("receipt_doc_id", "FILE_ID", OPT, "FK", "Documents.doc_id", "ใบเสร็จจากผู้ขาย", USR),
    ("void_of_payment_id", "UUID", OPT, "FK", "PO_Payments.payment_id", "ถ้าเป็นการกลับรายการ ชี้ไปรายการเดิม", SYS),
    ("void_reason", "TEXT", OPT, "", "", "เหตุผลการกลับรายการ", USR),
    ("voided_by", "TEXT", OPT, "FK", "Users.email", "ผู้กลับรายการ (หัวหน้าบัญชี)", SYS),
    ("voided_at", "DATETIME", OPT, "", "", "เวลากลับรายการ", SYS),
    ("note", "TEXT", OPT, "", "", "หมายเหตุ", USR),
    ("idempotency_key", "TEXT", REQ, "UQ", "po_no+seq+payment_date+bank_ref", "กันกดปุ่มบันทึกซ้ำตอนเน็ตช้า", SYS),
    ("row_version", "NUMBER", REQ, "", "", "กันสองคนแก้ทับกัน", SYS),
    ("created_at", "DATETIME", REQ, "", "", "", SYS),
    ("updated_at", "DATETIME", REQ, "", "", "", SYS),
])

add("Doc_Checklist", [
    ("checklist_id", "UUID", REQ, "PK", "", "คีย์ภายใน", SYS),
    ("po_no", "TEXT", REQ, "FK", "Purchase_Orders.po_no", "PO ที่เอกสารใบนี้สังกัด", SYS),
    ("doc_type", "TEXT", REQ, "FK", "Doc_Templates.doc_type", "ประเภทเอกสาร", SYS),
    ("doc_name_th", "TEXT", REQ, "", "ใบแจ้งหนี้", "ชื่อที่ผู้ใช้เห็น", SYS),
    ("owner_dept", "ENUM", REQ, "", "SR / QC / LS / WH / AC", "แผนกเจ้าของเอกสารใบนี้", SYS),
    ("is_required", "BOOLEAN", REQ, "", "TRUE / FALSE", "บังคับหรือไม่", SYS),
    ("stage_code", "TEXT", REQ, "", "INVOICE_RECEIVED", "เอกสารใบนี้เกิดที่ขั้นไหน", SYS),
    ("doc_status", "ENUM", REQ, "", "WAITING / UPLOADED / APPROVED / REJECTED / NA", "สถานะเอกสาร", SYS),
    ("current_doc_id", "FILE_ID", OPT, "FK", "Documents.doc_id", "ไฟล์เวอร์ชันปัจจุบัน", SYS),
    ("amount", "NUMBER", OPT, "", "", "ยอดบนเอกสาร ใช้เทียบ 3 ทาง", USR),
    ("qty", "NUMBER", OPT, "", "", "จำนวนบนเอกสาร", USR),
    ("currency", "TEXT", OPT, "", "THB", "สกุลเงิน", USR),
    ("tracking_no", "TEXT", OPT, "", "EX123456789TH", "เลขพัสดุ สำหรับเอกสารตัวจริง", USR),
    ("sent_at", "DATE", OPT, "", "", "วันที่ส่งเอกสารตัวจริง", USR),
    ("received_by", "TEXT", OPT, "", "", "ผู้รับเอกสารตัวจริง", USR),
    ("received_at", "DATE", OPT, "", "", "วันที่รับเอกสารตัวจริง", USR),
    ("reject_reason", "TEXT", OPT, "", "", "เหตุผลที่ตีกลับเอกสาร", USR),
    ("updated_by", "TEXT", OPT, "", "", "", SYS),
    ("updated_at", "DATETIME", OPT, "", "", "", SYS),
])

add("Documents", [
    ("doc_id", "UUID", REQ, "PK", "", "คีย์ภายใน", SYS),
    ("ref_type", "ENUM", REQ, "", "PO / CLAIM / ITEM_REQ / PAYMENT / QUOTE", "ไฟล์นี้ผูกกับเอกสารประเภทไหน", SYS),
    ("ref_no", "TEXT", REQ, "", "PO-26-0042 หรือ PO-26-0042#2", "เลขเอกสารแม่ — บังคับ ห้ามว่าง ไม่มีไฟล์ลอย", SYS),
    ("doc_type", "TEXT", REQ, "", "INVOICE", "ประเภทเอกสาร", SYS),
    ("drive_file_id", "FILE_ID", REQ, "", "", "ไอดีไฟล์ใน Google Drive", SYS),
    ("file_name", "TEXT", REQ, "", "SLIP_PO-26-0042_s2_20260830.pdf", "ระบบตั้งชื่อให้เอง ไม่ฝากไว้กับวินัยคน", SYS),
    ("file_url", "TEXT", REQ, "", "", "ลิงก์เปิดไฟล์", SYS),
    ("mime_type", "TEXT", REQ, "", "application/pdf", "ชนิดไฟล์", SYS),
    ("size_bytes", "NUMBER", REQ, "", "284000", "ขนาดหลังบีบอัด", SYS),
    ("version", "NUMBER", REQ, "", "1, 2, 3", "เวอร์ชันที่เท่าไหร่", SYS),
    ("is_current", "BOOLEAN", REQ, "", "TRUE / FALSE", "เป็นฉบับล่าสุดหรือไม่ — ของเก่าไม่ลบ", SYS),
    ("supersedes_doc_id", "UUID", OPT, "FK", "Documents.doc_id", "แทนที่ไฟล์ไหน", SYS),
    ("checksum_md5", "TEXT", REQ, "", "", "ใช้กันอัพไฟล์ซ้ำ และเป็นหลักฐาน e-signature", SYS),
    ("uploaded_by", "TEXT", REQ, "FK", "Users.email", "ผู้อัพโหลด", SYS),
    ("uploaded_at", "DATETIME", REQ, "", "", "เวลาอัพโหลด", SYS),
    ("source", "ENUM", REQ, "", "WEB / LARK / EMAIL", "อัพจากช่องทางไหน", SYS),
    ("note", "TEXT", OPT, "", "ของมาไม่ครบ 5 KG", "หมายเหตุที่ผู้ใช้พิมพ์ในป๊อปอัปของขั้นนั้น", USR),
    ("original_name", "TEXT", REQ, "", "scan_2026-08-08.pdf", "ชื่อไฟล์ที่ผู้ใช้เลือกมา — เก็บไว้สืบกลับ", SYS),
    ("system_name", "TEXT", REQ, "", "PO-26-0042_GR.jpg",
     "ชื่อที่ระบบตั้งให้ <เลขรายการ>_<รหัสเอกสาร>.<นามสกุลเดิม> — ผู้ใช้ไม่ต้องตั้งเอง", SYS),
    ("stage_code", "TEXT", OPT, "FK", "Stage_Templates.stage_code", "แนบในขั้นไหน", SYS),
    ("is_merged", "BOOLEAN", REQ, "", "TRUE / FALSE", "ถูกรวมเข้าไฟล์ชุดแล้วหรือยัง", SYS),
    ("purge_after", "DATE", OPT, "", "", "วันที่จะลบต้นฉบับ = bundle_at + MERGE_KEEP_DAYS", SYS),
    ("derived_from_doc_id", "UUID", OPT, "FK", "Documents.doc_id",
     "ใช้กับไฟล์ที่ประทับข้อความ — ชี้กลับไปต้นฉบับที่ไม่ถูกแก้", SYS),
    ("stamp_json", "TEXT", OPT, "", '{"text":"...","pos":"BR","size":13}',
     "ข้อความ ตำแหน่ง ขนาด สี ที่ใช้ประทับลงเอกสาร", SYS),
])

add("Approvals", [
    ("approval_id", "UUID", REQ, "PK", "", "คีย์ภายใน", SYS),
    ("ref_type", "ENUM", REQ, "", "PO / PAYMENT / CLAIM / ITEM_REQ", "อนุมัติอะไร", SYS),
    ("ref_no", "TEXT", REQ, "", "PO-26-0042", "เลขเอกสารที่ขออนุมัติ", SYS),
    ("doc_id", "UUID", OPT, "FK", "Documents.doc_id", "ไฟล์ที่เซ็น (ถ้ามี)", SYS),
    ("approval_round", "NUMBER", REQ, "", "1", "รอบที่เท่าไหร่ (ตีกลับแล้วส่งใหม่ = รอบใหม่)", SYS),
    ("seq", "NUMBER", REQ, "", "1, 2", "ลำดับผู้เซ็นในรอบนั้น", SYS),
    ("approver_email", "TEXT", REQ, "FK", "Users.email", "ผู้อนุมัติ", SYS),
    ("approver_role", "TEXT", REQ, "", "SR Head / GM / AC Head", "ตำแหน่งที่ใช้อนุมัติ", SYS),
    ("decision", "ENUM", REQ, "", "PENDING / APPROVED / REJECTED", "ผลการตัดสิน", USR),
    ("decided_at", "DATETIME", OPT, "", "", "เวลาที่ตัดสิน", SYS),
    ("reject_reason", "TEXT", OPT, "", "", "เหตุผลที่ไม่อนุมัติ — บังคับเมื่อ REJECTED", USR),
    ("on_behalf_of", "TEXT", OPT, "FK", "Users.email", "เซ็นแทนใคร (กรณีรับมอบอำนาจ)", SYS),
    ("signed_file_checksum", "TEXT", OPT, "", "", "หลักฐาน: checksum ของไฟล์ ณ เวลาที่เซ็น", SYS),
    ("signer_ip", "TEXT", OPT, "", "", "หลักฐาน: IP ผู้เซ็น", SYS),
    ("user_agent", "TEXT", OPT, "", "", "หลักฐาน: เบราว์เซอร์ผู้เซ็น", SYS),
    ("evidence_json", "JSON", OPT, "", "", "หลักฐานเพิ่มเติมตาม พ.ร.บ.ธุรกรรมทางอิเล็กทรอนิกส์", SYS),
    ("notified_at", "DATETIME", OPT, "", "", "เวลาที่แจ้งผู้อนุมัติ", SYS),
])

add("Stage_Log", [
    ("log_id", "UUID", REQ, "PK", "", "คีย์ภายใน", SYS),
    ("at_ts", "DATETIME", REQ, "", "", "เวลาที่เกิดเหตุการณ์", SYS),
    ("po_no", "TEXT", OPT, "FK", "Purchase_Orders.po_no", "PO ที่เกี่ยวข้อง", SYS),
    ("entity_type", "ENUM", REQ, "", "PO / PAYMENT / CLAIM / ITEM_REQ / PRICE_REQ / QUOTE", "เอกสารประเภทไหน", SYS),
    ("entity_no", "TEXT", REQ, "", "", "เลขเอกสาร", SYS),
    ("action", "TEXT", REQ, "", "STAGE_ADVANCE / UPLOAD / APPROVE / BLOCKED_BY_RULE", "ทำอะไร", SYS),
    ("from_status", "TEXT", OPT, "", "", "สถานะก่อนหน้า", SYS),
    ("to_status", "TEXT", OPT, "", "", "สถานะหลังเปลี่ยน", SYS),
    ("actor_email", "TEXT", REQ, "FK", "Users.email", "ใครทำ", SYS),
    ("actor_dept", "ENUM", REQ, "", "SR / QC / LS / WH / AC / GM", "แผนกของผู้ทำ", SYS),
    ("on_behalf_of", "TEXT", OPT, "", "", "ทำแทนใคร", SYS),
    ("duration_prev_hrs", "NUMBER", OPT, "", "26.5", "ขั้นก่อนหน้าใช้เวลากี่ชั่วโมง — ที่มาของรายงานคอขวด", SYS),
    ("sla_hours", "NUMBER", OPT, "", "24", "SLA ของขั้นนั้น", SYS),
    ("sla_breached", "BOOLEAN", OPT, "", "TRUE / FALSE", "เลยกำหนดหรือไม่", SYS),
    ("reason", "TEXT", OPT, "", "", "เหตุผล (กรณีตีกลับ ข้ามขั้น หรือถูกกฎบล็อก)", USR),
    ("changed_fields", "JSON", OPT, "", "", "ฟิลด์ที่เปลี่ยนและค่าก่อน/หลัง", SYS),
    ("source", "ENUM", OPT, "", "WEB / SYSTEM / SYNC", "เกิดจากคนหรือระบบ", SYS),
    ("request_id", "TEXT", OPT, "", "", "ใช้ไล่ปัญหาย้อนหลัง", SYS),
])

add("Notifications", [
    ("notif_id", "UUID", REQ, "PK", "", "คีย์ภายใน", SYS),
    ("created_at", "DATETIME", REQ, "", "", "เวลาที่เข้าคิว", SYS),
    ("channel", "ENUM", REQ, "", "LARK / EMAIL / WEB", "ช่องทาง — ไม่มี LINE เพราะภายในใช้ Lark อย่างเดียว", SYS),
    ("template_id", "TEXT", REQ, "", "", "แม่แบบข้อความ", SYS),
    ("to_recipients", "TEXT", REQ, "", "", "ผู้รับ", SYS),
    ("cc", "TEXT", OPT, "", "", "สำเนาถึง", SYS),
    ("lark_user_id", "TEXT", OPT, "FK", "Users.lark_user_id", "ไอดี Lark ของผู้รับ", SYS),
    ("subject", "TEXT", OPT, "", "", "หัวเรื่อง", SYS),
    ("ref_no", "TEXT", OPT, "", "", "เอกสารที่เกี่ยวข้อง", SYS),
    ("vars_json", "JSON", OPT, "", "", "ตัวแปรที่เติมลงแม่แบบ", SYS),
    ("status", "ENUM", REQ, "", "QUEUED / SENT / FAILED / SUPPRESSED / FALLBACK_EMAIL", "สถานะการส่ง", SYS),
    ("attempts", "NUMBER", REQ, "", "0-5", "ลองส่งไปกี่ครั้ง", SYS),
    ("next_retry_at", "DATETIME", OPT, "", "", "รอบถัดไป (1 → 5 → 25 นาที)", SYS),
    ("sent_at", "DATETIME", OPT, "", "", "เวลาที่ส่งสำเร็จ", SYS),
    ("message_id", "TEXT", OPT, "", "", "ไอดีข้อความจากปลายทาง", SYS),
    ("error", "TEXT", OPT, "", "", "ข้อผิดพลาดล่าสุด", SYS),
    ("dedupe_key", "TEXT", REQ, "UQ", "event+ref+user+date", "กันเรื่องเดิมเด้งซ้ำใน 24 ชม.", SYS),
])

add("Issues", [
    ("issue_id", "UUID", REQ, "PK", "", "คีย์ภายใน", SYS),
    ("issue_no", "TEXT", REQ, "UQ", "BUG-26-0031", "เลขที่ผู้ใช้เห็น", SYS),
    ("severity", "ENUM", REQ, "", "BLOCKER / WRONG / SUGGEST", "ความรุนแรงที่ผู้แจ้งเลือก", USR),
    ("title", "TEXT", REQ, "", "", "หัวข้อสั้น", USR),
    ("description", "TEXT", OPT, "", "", "รายละเอียดที่ผู้ใช้พิมพ์", USR),
    ("page", "TEXT", REQ, "", "หน้า PO → งวดจ่าย", "หน้าจอที่เกิดปัญหา — ระบบเก็บเอง", SYS),
    ("ref_type", "TEXT", OPT, "", "", "ประเภทเอกสารที่เปิดอยู่", SYS),
    ("ref_no", "TEXT", OPT, "", "", "เลขเอกสารที่เปิดอยู่", SYS),
    ("reported_by", "TEXT", REQ, "FK", "Users.email", "ผู้แจ้ง", SYS),
    ("reported_role", "ENUM", REQ, "", "SR / QC / LS / WH / AC / GM", "แผนกผู้แจ้ง", SYS),
    ("user_agent", "TEXT", OPT, "", "", "เบราว์เซอร์ — ระบบเก็บเอง", SYS),
    ("screen_size", "TEXT", OPT, "", "1440x900", "ขนาดจอ — ระบบเก็บเอง", SYS),
    ("last_actions_json", "JSON", OPT, "", "", "20 การกระทำล่าสุดก่อนเกิดปัญหา", SYS),
    ("console_error", "TEXT", OPT, "", "", "ข้อผิดพลาดล่าสุดในเบราว์เซอร์", SYS),
    ("stack_hash", "TEXT", OPT, "", "", "ใช้รวมเรื่องซ้ำเป็นใบเดียว", SYS),
    ("occurrences", "NUMBER", REQ, "", "1", "เจอกี่ครั้ง", SYS),
    ("screenshot_file_id", "FILE_ID", OPT, "", "", "ภาพหน้าจอ", SYS),
    ("status", "ENUM", REQ, "", "NEW / ACK / FIXING / FIXED / CLOSED / WONTFIX", "สถานะ — ไม่มีการลบ ปิดได้อย่างเดียว", USR),
    ("assigned_to", "TEXT", OPT, "", "", "ผู้รับผิดชอบฝั่ง IT", USR),
    ("acked_at", "DATETIME", OPT, "", "", "เวลาที่ IT รับเรื่อง (SLA 1 วันทำการ)", SYS),
    ("fixed_at", "DATETIME", OPT, "", "", "เวลาที่แก้เสร็จ", SYS),
    ("fix_note", "TEXT", OPT, "", "", "แก้อะไรไป", USR),
    ("confirmed_by_reporter", "BOOLEAN", OPT, "", "TRUE / FALSE", "ผู้แจ้งยืนยันว่าใช้ได้แล้ว — คนแจ้งเป็นคนปิด ไม่ใช่ IT", USR),
    ("created_at", "DATETIME", REQ, "", "", "", SYS),
])

add("Claims", [
    ("claim_id", "UUID", REQ, "PK", "", "คีย์ภายใน", SYS),
    ("claim_no", "TEXT", REQ, "UQ", "EX-2026001 / CL-26-0004", "เลขที่ใบเคลม", SYS),
    ("source", "ENUM", REQ, "", "CUSTOMER / INCOMING", "มาจากลูกค้าร้องเรียน หรือตรวจรับไม่ผ่าน — เปลี่ยนทั้งคนกรอกและเอกสารอ้างอิง", USR),
    ("claim_date", "DATE", REQ, "", "", "วันที่เปิดเรื่อง", USR),
    ("product_group", "TEXT", OPT, "", "", "กลุ่มสินค้า", USR),
    ("seller_name", "TEXT", OPT, "", "", "ส่วนที่ 1 — ชื่อผู้ขายฝั่งเรา", USR),
    ("seller_unit", "TEXT", OPT, "", "", "ส่วนที่ 1 — หน่วยงาน", USR),
    ("seller_position", "TEXT", OPT, "", "", "ส่วนที่ 1 — ตำแหน่ง", USR),
    ("customer_name", "TEXT", OPT, "", "", "ส่วนที่ 1 — ลูกค้าที่ร้องเรียน", USR),
    ("pi_no", "TEXT", OPT, "", "", "ส่วนที่ 1 — เลขที่ PI (ฝั่งขาย)", USR),
    ("dn_no", "TEXT", OPT, "", "", "ส่วนที่ 1 — เลขที่ DN (ฝั่งขาย)", USR),
    ("buy_date", "DATE", OPT, "", "", "ส่วนที่ 1 — วันที่ซื้อ", USR),
    ("item_name", "TEXT", REQ, "", "", "ส่วนที่ 1 — สินค้า", USR),
    ("brand", "TEXT", OPT, "", "", "ส่วนที่ 1 — แบรนด์", USR),
    ("lot_no", "TEXT", OPT, "", "", "ส่วนที่ 1 — เลข Lot", USR),
    ("warehouse", "TEXT", OPT, "", "", "ส่วนที่ 1 — คลัง", USR),
    ("packing", "TEXT", OPT, "", "", "ส่วนที่ 1 — ขนาดบรรจุ", USR),
    ("qty_defect", "NUMBER", OPT, "", "", "ส่วนที่ 1 — จำนวนที่มีปัญหา", USR),
    ("uom", "TEXT", OPT, "", "KG", "หน่วยนับ", USR),
    ("weight_kg", "NUMBER", OPT, "", "", "น้ำหนัก", USR),
    ("description", "TEXT", REQ, "", "", "ส่วนที่ 1 — ลักษณะปัญหา", USR),
    ("inspect_place", "TEXT", OPT, "", "", "สถานที่ตรวจ", USR),
    ("return_truck_type", "TEXT", OPT, "", "", "ส่วนที่ 2 — ประเภทรถที่ส่งคืน", USR),
    ("qc_result", "ENUM", OPT, "", "PASS / FAIL", "ส่วนที่ 2 — ผลตรวจ · FAIL จะเปิดส่วนที่ 3 ให้เอง", USR),
    ("qc_note", "TEXT", OPT, "", "", "ส่วนที่ 2 — หมายเหตุ QC", USR),
    ("customer_wants_doc", "BOOLEAN", OPT, "", "TRUE / FALSE", "ลูกค้าขอเอกสารหรือไม่", USR),
    ("photo_count", "NUMBER", OPT, "", "3", "จำนวนรูปหลักฐาน", SYS),
    ("inspection_report_file_id", "FILE_ID", OPT, "", "", "Inspection Report ที่ระบบสร้างจากส่วนที่ 2 — ไม่ต้องกรอกซ้ำ", SYS),
    ("supplier_code", "TEXT", OPT, "FK", "Suppliers.supplier_code", "ส่วนที่ 3 — ผู้ขายที่เคลม", USR),
    ("supplier_po_no", "TEXT", OPT, "", "", "ส่วนที่ 3 — PO ของผู้ขาย", USR),
    ("po_no", "TEXT", OPT, "FK", "Purchase_Orders.po_no", "ส่วนที่ 3 — PO ฝั่งซื้อที่เกี่ยวข้อง", USR),
    ("claim_status", "ENUM", REQ, "", "OPEN / SENT_TO_SUPPLIER / CN_RECEIVED / CLOSED", "สถานะเคลม", SYS),
    ("credit_note_no", "TEXT", OPT, "", "", "เลขใบลดหนี้จากผู้ขาย", USR),
    ("credit_note_amount", "NUMBER", OPT, "", "", "ยอดใบลดหนี้", USR),
    ("credit_note_received_at", "DATE", OPT, "", "", "วันที่ได้รับใบลดหนี้ — ตัวปลดล็อกงวดจ่าย", USR),
    ("blocks_payment", "BOOLEAN", REQ, "", "TRUE / FALSE", "ใบนี้ล็อกงวดจ่ายอยู่หรือไม่", SYS),
    ("created_by", "TEXT", REQ, "", "", "", SYS),
    ("created_at", "DATETIME", REQ, "", "", "", SYS),
    ("updated_at", "DATETIME", REQ, "", "", "", SYS),
    ("row_version", "NUMBER", REQ, "", "", "", SYS),
])

add("ItemCode_Requests", [
    ("request_id", "UUID", REQ, "PK", "", "คีย์ภายใน", SYS),
    ("request_no", "TEXT", REQ, "UQ", "IC-26-0007", "เลขที่ใบขอ", SYS),
    ("po_no", "TEXT", OPT, "FK", "Purchase_Orders.po_no", "PO ที่รออยู่ (ว่างได้)", SYS),
    ("item_name", "TEXT", REQ, "", "", "ชื่อสินค้า", USR),
    ("brand", "TEXT", OPT, "", "", "แบรนด์", USR),
    ("product_group", "TEXT", OPT, "", "", "กลุ่มสินค้า", USR),
    ("uom", "TEXT", REQ, "", "KG", "หน่วยนับ", USR),
    ("packing", "TEXT", OPT, "", "10 KG x 1 CTN", "ขนาดบรรจุ", USR),
    ("supplier_code", "TEXT", OPT, "FK", "Suppliers.supplier_code", "ผู้ขาย", USR),
    ("hs_code", "TEXT", OPT, "", "0307.43.00", "พิกัดศุลกากร", USR),
    ("storage_condition", "TEXT", OPT, "", "แช่แข็ง", "การเก็บรักษา", USR),
    ("shelf_life_days", "NUMBER", OPT, "", "730", "อายุสินค้า (วัน)", USR),
    ("reason", "TEXT", OPT, "", "", "เหตุผลที่ขอเปิดรหัส", USR),
    ("status", "ENUM", REQ, "", "DRAFT / WAIT_APPROVE / WAIT_LS / DONE / REJECTED", "สถานะใบขอ", SYS),
    ("approved_by", "TEXT", OPT, "FK", "Users.email", "หัวหน้าที่อนุมัติ", SYS),
    ("approved_at", "DATETIME", OPT, "", "", "เวลาอนุมัติ", SYS),
    ("sap_item_code", "TEXT", OPT, "FK", "OITM.ItemCode", "รหัสจริงจาก B1 — PO เดินต่อไม่ได้จนกว่าจะมีค่านี้", B1),
    ("sap_created_by", "TEXT", OPT, "", "", "LS ผู้เปิดรหัสใน B1", SYS),
    ("sap_created_at", "DATETIME", OPT, "", "", "เวลาที่เปิดรหัสใน B1", SYS),
    ("created_by", "TEXT", REQ, "", "", "", SYS),
    ("created_at", "DATETIME", REQ, "", "", "", SYS),
])

add("Price_Requests", [
    ("request_id", "UUID", REQ, "PK", "", "คีย์ภายใน", SYS),
    ("request_no", "TEXT", REQ, "UQ", "PR-26-0004", "เลขที่ใบขอราคา", SYS),
    ("item_name", "TEXT", REQ, "", "ปลาแซลมอนแล่ เกรด A", "ชื่อสินค้า — กรอกที่นี่ครั้งเดียว ใบเสนอราคาและ PO ดึงต่อ", USR),
    ("spec", "TEXT", OPT, "", "แล่ไม่มีหนัง ไม่มีก้าง แช่แข็ง -18°C", "สเปกที่ใช้ขอราคา", USR),
    ("qty", "NUMBER", REQ, "", "200", "จำนวนที่ต้องการ", USR),
    ("uom", "TEXT", REQ, "", "KG", "หน่วยนับ", USR),
    ("product_group", "TEXT", OPT, "", "Food", "กลุ่มสินค้า", USR),
    ("need_by_date", "DATE", OPT, "", "", "ต้องการภายในวันไหน", USR),
    ("status", "ENUM", REQ, "", "NEW / QUOTING / CONVERTED / CANCELLED", "สถานะใบขอราคา", SYS),
    ("requested_by", "TEXT", REQ, "FK", "Users.email", "ผู้ขอ", SYS),
    ("note", "TEXT", OPT, "", "", "หมายเหตุ", USR),
    ("created_at", "DATETIME", REQ, "", "", "", SYS),
    ("updated_at", "DATETIME", REQ, "", "", "", SYS),
])

add("Quotations", [
    ("quote_id", "UUID", REQ, "PK", "", "คีย์ภายใน", SYS),
    ("quote_no", "TEXT", REQ, "UQ", "QT-26-0009", "เลขที่ใบเสนอราคา", SYS),
    ("price_request_id", "UUID", REQ, "FK", "Price_Requests.request_id", "ใบขอราคาต้นทาง", SYS),
    ("supplier", "TEXT", REQ, "", "Ocean Best", "ผู้ขายที่เสนอราคา", USR),
    ("supplier_code", "TEXT", OPT, "FK", "OCRD.CardCode", "รหัสผู้ขายใน B1 ถ้ามี", B1),
    ("unit_price", "NUMBER", REQ, "", "295.00", "ราคาต่อหน่วยที่เสนอมา", USR),
    ("qty", "NUMBER", REQ, "", "200", "จำนวน — คัดลอกจากใบขอราคาตอนสร้าง ไม่ให้พิมพ์เอง", CPY),
    ("total_amount", "NUMBER", REQ, "", "59000.00", "ราคารวม = unit_price x qty", SYS),
    ("currency", "TEXT", REQ, "", "THB", "สกุลเงิน", USR),
    ("lead_time", "TEXT", OPT, "", "7 วัน", "ระยะเวลาส่งมอบ", USR),
    ("valid_until", "DATE", OPT, "", "", "ยืนราคาถึงวันไหน", USR),
    ("selected", "BOOLEAN", REQ, "", "TRUE / FALSE", "ใบที่เลือกซื้อ — มีได้ใบเดียวต่อใบขอราคา", USR),
    ("converted_po_no", "TEXT", OPT, "FK", "Purchase_Orders.po_no", "PO ที่เปิดจากใบนี้ — ตามรอยได้สองทาง", SYS),
    ("created_by", "TEXT", REQ, "", "", "", SYS),
    ("created_at", "DATETIME", REQ, "", "", "", SYS),
])

add("Stage_Templates", [
    ("template_id", "TEXT", REQ, "PK", "DOMESTIC_FOOD-01", "คีย์", SYS),
    ("module", "TEXT", REQ, "", "DOMESTIC_FOOD", "งานประเภทไหน", USR),
    ("seq", "NUMBER", REQ, "", "1-17", "ลำดับขั้น", USR),
    ("stage_code", "TEXT", REQ, "", "QC_INCOMING", "รหัสขั้น", USR),
    ("stage_name_th", "TEXT", REQ, "", "QC ตรวจรับ", "ชื่อที่ผู้ใช้เห็น", USR),
    ("phase_code", "ENUM", REQ, "", "BUY / RECV / AP / PAY", "เฟสที่ใช้แสดงบนหน้าจอ (ยุบ 17 ขั้นเหลือ 4)", USR),
    ("owner_dept", "ENUM", REQ, "", "SR / QC / LS / WH / AC / GM", "แผนกเจ้าของขั้น", USR),
    ("owner_resolve", "ENUM", REQ, "", "DEPT_QUEUE / PO_OWNER / ROLE / FIXED_EMAIL", "วิธีหาตัวคนที่รับผิดชอบ", USR),
    ("owner_value", "TEXT", OPT, "", "", "ค่าประกอบของ owner_resolve", USR),
    ("sla_hours", "NUMBER", REQ, "", "24", "ควรเสร็จในกี่ชั่วโมง — แก้ที่แท็บ «ขั้นตอนงาน»", USR),
    ("skip_condition", "TEXT", OPT, "", "has_sap_item_code = TRUE", "เงื่อนไขข้ามขั้น", USR),
    ("branch_on_fail", "TEXT", OPT, "", "CLAIM", "ถ้าไม่ผ่านให้แตกไปทางไหน", USR),
    ("requires_approval", "BOOLEAN", REQ, "", "TRUE / FALSE", "ขั้นนี้ต้องมีการอนุมัติไหม", USR),
    ("is_active", "BOOLEAN", REQ, "", "TRUE / FALSE", "เปิดใช้อยู่ไหม", USR),
])

add("Doc_Templates", [
    ("doc_template_id", "TEXT", REQ, "PK", "", "คีย์", SYS),
    ("module", "TEXT", REQ, "", "DOMESTIC_FOOD", "งานประเภทไหน", USR),
    ("doc_type", "TEXT", REQ, "", "INVOICE", "รหัสประเภทเอกสาร", USR),
    ("doc_name_th", "TEXT", REQ, "", "ใบแจ้งหนี้", "ชื่อที่ผู้ใช้เห็น", USR),
    ("owner_dept", "ENUM", REQ, "", "SR / QC / LS / WH / AC", "แผนกเจ้าของเอกสาร", USR),
    ("is_required", "BOOLEAN", REQ, "", "TRUE / FALSE", "บังคับหรือไม่", USR),
    ("stage_code", "TEXT", REQ, "", "INVOICE_RECEIVED", "เกิดที่ขั้นไหน", USR),
    ("needs_amount", "BOOLEAN", REQ, "", "TRUE / FALSE", "ต้องกรอกยอดไหม (ใช้เทียบ 3 ทาง)", USR),
    ("needs_approval", "BOOLEAN", REQ, "", "TRUE / FALSE", "ต้องอนุมัติไหม", USR),
    ("is_physical", "BOOLEAN", REQ, "", "TRUE / FALSE", "ต้องติดตามตัวจริงไหม", USR),
    ("allowed_mime", "TEXT", REQ, "", "application/pdf,image/jpeg,image/png", "ชนิดไฟล์ที่รับ", USR),
    ("seq", "NUMBER", REQ, "", "", "ลำดับแสดงผล", USR),
    ("is_active", "BOOLEAN", REQ, "", "TRUE / FALSE", "เปิดใช้อยู่ไหม", USR),
])

add("Users", [
    ("email", "TEXT", REQ, "PK", "somchai@mgs.co.th", "อีเมลคือ identity ของผู้ใช้", USR),
    ("full_name", "TEXT", REQ, "", "", "ชื่อที่แสดงในระบบ", USR),
    ("dept", "ENUM", REQ, "", "SR / QC / LS / WH / AC / GM / IT", "แผนก", USR),
    ("roles", "TEXT", REQ, "", "SR,SR_HEAD", "บทบาท (มีได้หลายบทบาท)", USR),
    ("lark_user_id", "TEXT", REQ, "", "ou_xxxxx", "จำเป็นทุกคน เพราะแจ้งเตือนภายในใช้ Lark ช่องทางเดียว", USR),
    ("manager_email", "TEXT", OPT, "FK", "Users.email", "หัวหน้า — ใช้ตอน escalate เมื่อเลย SLA", USR),
    ("delegate_email", "TEXT", OPT, "FK", "Users.email", "ผู้รับมอบอำนาจตอนลา", USR),
    ("delegate_until", "DATE", OPT, "", "", "มอบอำนาจถึงวันไหน", USR),
    ("approval_limit", "NUMBER", OPT, "", "50000", "วงเงินอนุมัติ", USR),
    ("notify_channel", "ENUM", OPT, "", "LARK / EMAIL / BOTH", "ช่องทางที่ผู้ใช้เลือก", USR),
    ("is_active", "BOOLEAN", REQ, "", "TRUE / FALSE", "ยังทำงานอยู่ไหม", USR),
])

add("Suppliers", [
    ("supplier_code", "TEXT", REQ, "PK", "OCRD.CardCode", "รหัสผู้ขาย — ตรงกับ B1", B1),
    ("supplier_name", "TEXT", REQ, "", "OCRD.CardName", "ชื่อผู้ขาย", B1),
    ("contact_email", "TEXT", OPT, "", "OCRD.E_Mail", "อีเมล — ใช้ส่งสลิปให้อัตโนมัติ", B1),
    ("contact_line", "TEXT", OPT, "", "", "LINE ของผู้ขาย (ภายนอกเท่านั้น ไม่ใช่ช่องแจ้งเตือนภายใน)", USR),
    ("payment_term", "TEXT", OPT, "", "OCTG.PymntGroup", "เงื่อนไขชำระตั้งต้นของผู้ขาย", B1),
    ("is_active", "BOOLEAN", REQ, "", "TRUE / FALSE", "ยังค้าขายอยู่ไหม", B1),
])

add("Config / Lookups", [
    ("config_key", "TEXT", REQ, "PK", "SLA_DEFAULT_HOURS", "ชื่อค่าตั้งต้น", USR),
    ("config_value", "TEXT", REQ, "", "24", "ค่า — แก้ที่แท็บ «ค่าตั้งต้น»", USR),
    ("description", "TEXT", OPT, "", "", "ค่านี้ใช้ทำอะไร", USR),
    ("lookup_group", "TEXT", REQ, "", "PAYMENT_METHOD", "กลุ่ม dropdown", USR),
    ("code", "TEXT", REQ, "", "TRANSFER", "รหัสค่า", USR),
    ("label_th", "TEXT", REQ, "", "โอนเงิน", "ข้อความที่ผู้ใช้เห็น", USR),
    ("sort_order", "NUMBER", OPT, "", "1", "ลำดับใน dropdown", USR),
    ("is_active", "BOOLEAN", REQ, "", "TRUE / FALSE", "เปิดใช้อยู่ไหม", USR),
])

add("Payment_Requests", [
    ("request_id", "TEXT", REQ, "PK", "PRQ-26-0031", "เลขคำขอทำจ่าย", SYS),
    ("po_no", "TEXT", REQ, "FK", "Purchase_Orders.po_no", "รายการที่ขอจ่าย", SYS),
    ("payment_seq", "NUMBER", REQ, "FK", "PO_Payments.seq", "งวดที่ขอ", SYS),
    ("amount", "NUMBER", REQ, "", "122500.00",
     "ยอดเงินที่ต้องจ่าย — ช่องบังคับตามที่ทีมขอ (กฎ S1, S2)", USR),
    ("effective_date", "DATE", REQ, "", "26/08/2026",
     "Effective date — ช่องบังคับตามที่ทีมขอ (กฎ S3, S4) · ลงเป็นวันครบกำหนดของงวดนั้น", USR),
    ("stamp_on_doc", "BOOLEAN", REQ, "", "TRUE / FALSE", "ติ๊กให้พิมพ์ข้อความนี้ลงบนเอกสารที่แนบ", USR),
    ("stamp_target_doc_id", "UUID", OPT, "FK", "Documents.doc_id", "เอกสารที่จะประทับ", USR),
    ("stamp_text", "TEXT", OPT, "", "ขออนุมัติทำจ่าย PO-26-0045 ...",
     "ข้อความที่ผู้ใช้แก้ในโหมดตัวอย่าง — บังคับเมื่อ stamp_on_doc (กฎ S7)", USR),
    ("stamp_position", "ENUM", OPT, "", "TL / TR / BR / CT", "ตำแหน่งบนหน้ากระดาษ", USR),
    ("stamp_size_pt", "NUMBER", OPT, "", "11 / 13 / 16", "ขนาดตัวอักษร", USR),
    ("stamp_color", "TEXT", OPT, "", "ดำ / แดง / เทาจาง", "สีข้อความ", USR),
    ("requested_by", "TEXT", REQ, "FK", "Users.email", "ผู้ขอ (จัดซื้อ)", SYS),
    ("requested_at", "DATETIME", REQ, "", "", "เวลาที่ส่งคำขอ", SYS),
    ("approver_role", "ENUM", REQ, "", "AC_HEAD / GM", "เลือกอัตโนมัติจากวงเงิน APPROVAL_LIMIT", SYS),
    ("decision", "ENUM", REQ, "", "PENDING / APPROVED / REJECTED", "ผลการอนุมัติ", SYS),
    ("decided_by", "TEXT", OPT, "FK", "Users.email",
     "ผู้ตัดสิน — ระบบบล็อกถ้าเป็นคนเดียวกับ requested_by", SYS),
    ("decided_at", "DATETIME", OPT, "", "", "เวลาที่ตัดสิน", SYS),
    ("reject_reason", "TEXT", OPT, "", "", "บังคับเมื่อ decision = REJECTED", USR),
    ("attached_doc_ids", "TEXT", OPT, "", "doc1,doc2,doc3", "เอกสารครบชุดที่แนบไปกับคำขออัตโนมัติ", SYS),
    ("idempotency_key", "TEXT", REQ, "UQ", "", "กันกดส่งซ้ำแล้วเกิดคำขอซ้ำ", SYS),
    ("row_version", "NUMBER", REQ, "", "1, 2, 3...", "กันสองคนแก้ทับกันเงียบ ๆ", SYS),
])

add("Doc_Bundles", [
    ("bundle_id", "UUID", REQ, "PK", "", "คีย์ภายใน", SYS),
    ("po_no", "TEXT", REQ, "FK", "Purchase_Orders.po_no", "1 รายการ = 1 ชุด", SYS),
    ("doc_id", "UUID", REQ, "FK", "Documents.doc_id", "ไฟล์รวมที่สร้างขึ้น", SYS),
    ("parts_json", "TEXT", REQ, "", '["PI","PO_SIGNED","QC_RC",...]',
     "รายชื่อไฟล์ที่รวม เรียงตามลำดับจริง", SYS),
    ("page_count", "NUMBER", REQ, "", "14", "จำนวนหน้าของไฟล์รวม", SYS),
    ("size_kb", "NUMBER", REQ, "", "2240", "ใช้ติดตามพื้นที่ที่ใช้เพิ่ม", SYS),
    ("built_at", "DATETIME", REQ, "", "", "เวลาที่รวมเสร็จ", SYS),
    ("build_status", "ENUM", REQ, "", "OK / PARTIAL / FAILED",
     "PARTIAL = มีไฟล์แปลงเป็นหน้า PDF ไม่ได้ ต้องแจ้ง IT", SYS),
    ("failed_parts_json", "TEXT", OPT, "", "", "ใบที่รวมไม่สำเร็จ", SYS),
    ("purge_scheduled_at", "DATE", OPT, "", "", "วันที่จะลบต้นฉบับที่รวมแล้ว", SYS),
])

add("Sync_Log", [
    ("sync_id", "UUID", REQ, "PK", "", "คีย์ภายใน", SYS),
    ("object", "ENUM", REQ, "", "PO / BP / ITEM / GRPO / APINV / PAYMENT / TERMS", "ดึงข้อมูลอะไร", SYS),
    ("started_at", "DATETIME", REQ, "", "", "เริ่มรอบเมื่อไหร่", SYS),
    ("finished_at", "DATETIME", OPT, "", "", "จบรอบเมื่อไหร่", SYS),
    ("source_file", "TEXT", OPT, "", "PO_20260803_0830.json", "ไฟล์ที่ตัวส่งวางไว้", SYS),
    ("rows_read", "NUMBER", REQ, "", "", "อ่านมากี่แถว", SYS),
    ("rows_upserted", "NUMBER", REQ, "", "", "เขียนลงระบบกี่แถว", SYS),
    ("rows_rejected", "NUMBER", REQ, "", "", "ตกไปกี่แถว", SYS),
    ("status", "ENUM", REQ, "", "OK / PARTIAL / FAILED / SKIPPED", "ผลของรอบนั้น", SYS),
    ("error_msg", "TEXT", OPT, "", "", "ข้อผิดพลาด", SYS),
    ("watermark", "DATETIME", REQ, "", "", "ดึงถึงเวลาไหนแล้ว — รอบถัดไปเริ่มจากตรงนี้", SYS),
])

r = 5
prev_table = None
for (t, colname, dtype, req, key, vals, desc, src) in D:
    is_new = t in ("Price_Requests", "Quotations", "Payment_Requests", "Doc_Bundles")
    bg_row = C_NEW if is_new else None
    cell(ws, r, 1, t, bold=(t != prev_table), bg=bg_row)
    cell(ws, r, 2, colname, bg=bg_row)
    cell(ws, r, 3, dtype, bg=bg_row, al="center")
    cell(ws, r, 4, req, bg=bg_row, al="center")
    cell(ws, r, 5, key, bg=C_LOCK, al="center")
    cell(ws, r, 6, vals, bg=bg_row)
    cell(ws, r, 7, desc, bg=bg_row)
    cell(ws, r, 8, src, bg=C_EDIT if src == USR else bg_row, al="center")
    cell(ws, r, 9, "", bg=C_EDIT)
    prev_table = t
    r += 1

DICT_LAST = r - 1
dv(ws, "DTYPE", "C5:C%d" % DICT_LAST)
dv(ws, "YN", "D5:D%d" % DICT_LAST)
dv(ws, "SRC", "H5:H%d" % DICT_LAST)
ws.freeze_panes = "C5"
ws.auto_filter.ref = "A4:I%d" % DICT_LAST

# ============================================================
# 4. ขั้นตอนงาน
# ============================================================
ws = wb.create_sheet("ขั้นตอนงาน")
title(ws, "ขั้นตอนงาน (Stage_Templates) — module DOMESTIC_FOOD",
      "SLA แก้ได้ที่คอลัมน์เหลือง · เพิ่มงานนำเข้า/Solar ในอนาคต = เพิ่มแถว ไม่ต้องแก้โปรแกรม", 9)
widths(ws, [5, 24, 30, 12, 10, 14, 30, 20, 34])
H = ["ลำดับ", "รหัสขั้น", "ชื่อขั้น (ที่ผู้ใช้เห็น)", "แผนกเจ้าของ", "เฟส", "SLA (ชม.)",
     "เงื่อนไขข้ามขั้น", "ถ้าไม่ผ่านไปไหน", "หมายเหตุของทีม"]
for i, h in enumerate(H, 1):
    cell(ws, 4, i, h)
hdr(ws, 4, 9)

STAGES = [
    ("PRICE_REQUEST", "เช็คราคา", "SR", "BUY", 24, "เปิดจากใบเสนอราคา = ข้ามขั้นนี้", "", "ผูกกับใบขอราคา/ใบเสนอราคาแล้ว"),
    ("SAMPLE_CHECK", "ตรวจตัวอย่าง", "QC", "BUY", 48, "ดูตัวอย่างก่อนซื้อไม่ได้", "REJECTED", ""),
    ("GM_DECISION", "ผู้บริหารตัดสินใจ", "GM", "BUY", 24, "ตรวจตัวอย่างผ่านแล้ว", "CANCELLED", ""),
    ("ORDER_CONFIRM", "คอนเฟิร์มออเดอร์", "SR", "BUY", 24, "", "", ""),
    ("PI_RECEIVED", "รับ PI จากผู้ขาย", "SR", "BUY", 48, "", "", ""),
    ("ITEM_CODE", "เปิดรหัสสินค้า", "LS", "BUY", 24, "มีรหัสใน B1 แล้ว", "", "ข้ามไม่ได้จนกว่าจะมีรหัสจริงจาก B1"),
    ("PO_CREATED", "ผูกกับ PO ใน B1", "SR", "BUY", 24, "", "", "กดปุ่มดึงจาก B1 ไม่ให้พิมพ์เลขเอง"),
    ("INTERNAL_APPROVAL", "อนุมัติภายใน", "GM", "BUY", 24, "", "ตีกลับหา SR", ""),
    ("SUPPLIER_SIGN", "ผู้ขายเซ็น PO กลับ", "SR", "BUY", 72, "", "", "เก็บเฉพาะ PO ที่เซ็นครบแล้ว"),
    ("DELIVERY_BOOKING", "จองรถรับของ", "LS", "RECV", 24, "", "", "จองล่วงหน้า 1 วัน"),
    ("QC_INCOMING", "QC ตรวจรับ", "QC", "RECV", 24, "", "CLAIM", "ไม่ผ่าน = เปิดเคลม + ล็อกงวดจ่าย"),
    ("GR", "รับเข้าคลัง + GRPO", "WH", "RECV", 24, "", "", "ดึงจาก OPDN/PDN1 ของ B1"),
    ("INVOICE_RECEIVED", "รับใบแจ้งหนี้", "AC", "AP", 72, "", "", "ยังต้องยืนยันว่า SR หรือ AC เป็นคนอัพ"),
    ("DOC_HANDOVER", "ส่งมอบเอกสาร", "SR", "AP", 24, "", "", ""),
    ("PAYMENT", "ทำจ่าย", "AC", "PAY", 48, "", "", "ตั้งเรื่อง → อนุมัติ → บันทึกจ่าย"),
    ("RECEIPT_ORIGINAL", "ใบเสร็จตัวจริง", "AC", "PAY", 336, "", "", "336 ชม. = 14 วัน · รอยืนยันว่าส่งตรงถึง AC ได้ไหม"),
    ("AC_CLOSE", "ตรวจ 3 ทาง + ปิด", "AC", "PAY", 48, "", "", ""),
]
r = 5
for i, (code, name, dept, phase, sla, skip, branch, note) in enumerate(STAGES, start=1):
    cell(ws, r, 1, i, al="center")
    cell(ws, r, 2, code, bg=C_LOCK)
    cell(ws, r, 3, name, bg=C_EDIT)
    cell(ws, r, 4, dept, bg=C_EDIT, al="center")
    cell(ws, r, 5, phase, bg=C_EDIT, al="center")
    cell(ws, r, 6, sla, bg=C_EDIT, al="center")
    cell(ws, r, 7, skip, bg=C_EDIT)
    cell(ws, r, 8, branch, bg=C_EDIT)
    cell(ws, r, 9, note, bg=C_EDIT)
    ws.row_dimensions[r].height = 26
    r += 1
STAGE_LAST = r - 1
dv(ws, "DEPT", "D5:D%d" % STAGE_LAST)
dv(ws, "PHASE", "E5:E%d" % STAGE_LAST)
ws.freeze_panes = "C5"

r += 1
cell(ws, r, 2, "รวม SLA ทั้งกระบวนการ (ชม.)", bold=True)
cell(ws, r, 6, "=SUM(F5:F%d)" % STAGE_LAST, bg=C_LOCK, bold=True, al="center")
r += 1
cell(ws, r, 2, "คิดเป็นวันทำการ (8 ชม./วัน)", bold=True)
cell(ws, r, 6, "=ROUND(F%d/8,1)" % (r - 1), bg=C_LOCK, bold=True, al="center")
r += 2
c = cell(ws, r, 2, "SLA ที่ใส่ไว้เป็นค่าที่ผมตั้งเอง — ต้องให้หัวหน้าแต่ละแผนกยืนยันว่าทำได้จริง "
                    "ก่อนเปิดใช้ระบบเตือน มิฉะนั้นระบบจะเตือนตลอดเวลาจนคนปิดการแจ้งเตือน",
         bg=C_WARN_BG, fg=C_WARN_FG)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
ws.row_dimensions[r].height = 30

# ============================================================
# 5. เอกสารที่ต้องมี
# ============================================================
ws = wb.create_sheet("เอกสารที่ต้องมี")
title(ws, "เอกสารที่ต้องมี (Doc_Templates)",
      "เอกสารบังคับต่อ PO หนึ่งใบ · สลิปไม่อยู่ในรายการนี้เพราะผูกกับ «งวดจ่าย» ไม่ใช่ผูกกับ PO", 8)
widths(ws, [22, 28, 12, 22, 12, 14, 26, 32])
H = ["รหัสเอกสาร", "ชื่อเอกสาร", "เจ้าของ", "เกิดที่ขั้น", "บังคับ", "ต้องกรอกยอด", "เก็บที่โฟลเดอร์", "หมายเหตุของทีม"]
for i, h in enumerate(H, 1):
    cell(ws, 4, i, h)
hdr(ws, 4, 8)

DOCS = [
    ("PI", "PI จากผู้ขาย", "SR", "PI_RECEIVED", "บังคับ", "TRUE", "01_PO/.../01_PI/", ""),
    ("PO_SIGNED", "PO ที่เซ็นครบแล้ว", "SR", "SUPPLIER_SIGN", "บังคับ", "TRUE", "01_PO/.../02_PO/",
     "เก็บเฉพาะใบที่เซ็นครบ — PO ที่ยังไม่เซ็นอยู่ใน B1 แล้ว"),
    ("PATTERN_LABEL", "Pattern label", "SR", "SUPPLIER_SIGN", "ไม่บังคับ", "FALSE", "01_PO/.../02_PO/",
     "ยังต้องยืนยันว่าใช้กับสินค้ากลุ่มไหน"),
    ("QC_RC", "ผลตรวจรับ / RC + รูป", "QC", "QC_INCOMING", "บังคับ", "TRUE", "01_PO/.../03_QC/",
     "จำนวนที่ QC รับ คือตัวเลขที่ใช้ตรวจ 3 ทาง ไม่ใช่จำนวนที่ส่งมา"),
    ("GR", "ใบรับสินค้า (GRPO)", "WH", "GR", "บังคับ", "TRUE", "01_PO/.../04_GR/", "ดึงจาก B1 (OPDN/PDN1)"),
    ("INVOICE", "ใบแจ้งหนี้", "AC", "INVOICE_RECEIVED", "บังคับ", "TRUE", "01_PO/.../05_Invoice/",
     "เลขที่ใบแจ้งหนี้ดึงจาก OPCH.NumAtCard ของ B1"),
    ("CREDIT_NOTE", "ใบลดหนี้", "AC", "(เมื่อมีเคลม)", "ตามเงื่อนไข", "TRUE", "01_PO/.../05_Invoice/",
     "ตัวปลดล็อกงวดจ่ายที่ถูกล็อกจากเคลม"),
    ("SLIP", "สลิป / Swift", "AC", "PAYMENT", "บังคับทุกงวด", "TRUE", "01_PO/.../06_Payment/",
     "ผูกกับงวดจ่าย ไม่ใช่กับ PO — 2 งวดก็มี 2 สลิป"),
    ("RECEIPT", "ใบเสร็จตัวจริง", "AC", "RECEIPT_ORIGINAL", "บังคับ", "TRUE", "01_PO/.../06_Payment/", ""),
]
r = 5
for code, name, dept, stage, req, amt, folder, note in DOCS:
    cell(ws, r, 1, code, bg=C_LOCK)
    cell(ws, r, 2, name, bg=C_EDIT)
    cell(ws, r, 3, dept, bg=C_EDIT, al="center")
    cell(ws, r, 4, stage, bg=C_EDIT)
    cell(ws, r, 5, req, bg=C_EDIT, al="center")
    cell(ws, r, 6, amt, bg=C_EDIT, al="center")
    cell(ws, r, 7, folder, bg=C_LOCK)
    cell(ws, r, 8, note, bg=C_EDIT)
    ws.row_dimensions[r].height = 30
    r += 1
dv(ws, "DEPT", "C5:C%d" % (r - 1))
ws.freeze_panes = "B5"

r += 1
cell(ws, r, 1, "จำนวนเอกสารบังคับต่อ PO", bold=True)
cell(ws, r, 2, '=COUNTIF(E5:E%d,"บังคับ")' % (r - 2), bg=C_LOCK, bold=True, al="center")

# ============================================================
# 6. กฎตรวจสอบ
# ============================================================
ws = wb.create_sheet("กฎตรวจสอบ")
title(ws, "กฎตรวจสอบก่อนบันทึก",
      "กฎที่ระบบใช้บล็อกหรือเตือน · «เกณฑ์» และ «ระดับ» แก้ได้ — ตัวเลขที่เป็นสีแดงคือค่าที่ผมตั้งเองและต้องให้ทีมยืนยัน", 8)
widths(ws, [7, 14, 44, 14, 16, 40, 14, 30])
H = ["รหัส", "จังหวะ", "กฎ", "ระดับ", "เกณฑ์ (แก้ได้)", "ข้อความที่ผู้ใช้เห็น", "สถานะ", "หมายเหตุของทีม"]
for i, h in enumerate(H, 1):
    cell(ws, 4, i, h)
hdr(ws, 4, 8)

RULES = [
    ("V1", "ขณะพิมพ์", "จำนวน / มูลค่า ต้องมากกว่า 0", "บล็อก", "> 0", "จำนวนต้องมากกว่า 0", "ใช้งาน", "", False),
    ("V2", "ขณะพิมพ์", "วันครบกำหนดจ่าย ต้องไม่ก่อนวันเปิด PO", "บล็อก", "≥ วันเปิด PO", "วันครบกำหนดอยู่ก่อนวันเปิด PO", "ใช้งาน", "", False),
    ("V3", "ขณะพิมพ์", "อีเมล / เบอร์ผู้ขาย รูปแบบถูกต้อง", "เตือน", "รูปแบบมาตรฐาน", "ขีดเส้นใต้แดงใต้ช่อง", "ใช้งาน", "", False),
    ("V4", "ตอนบันทึก", "ไฟล์แนบครบตามที่ขั้นนั้นบังคับ", "บล็อก", "ครบทุกใบที่ is_required", "ยังขาด: ใบแจ้งหนี้", "ใช้งาน", "", False),
    ("V5", "ตอนบันทึก", "ชนิดและขนาดไฟล์", "บล็อก", "pdf/jpg/png · ≤ 20 MB", "รองรับ PDF/JPG/PNG ไม่เกิน 20 MB", "ใช้งาน", "", True),
    ("V6", "ตอนบันทึก", "ไฟล์ซ้ำ (checksum ตรงกับที่มีแล้ว)", "เตือน", "MD5 ตรงกัน", "ไฟล์นี้อัพไปแล้วเมื่อ 2 ส.ค. — ต้องการทับหรือไม่", "ใช้งาน", "", False),
    ("V7", "ตอนบันทึก", "ผลรวมงวดจ่าย = มูลค่า PO", "บล็อก", "ต่างได้ไม่เกิน 1 บาท", "งวดจ่ายรวมไม่เท่ากับมูลค่า PO", "ใช้งาน", "", True),
    ("V8", "ตอนบันทึก", "จำนวนใน GR ต้องไม่เกินจำนวนใน PO", "บล็อก", "GR ≤ PO", "GR 520 KG เกิน PO 500 KG", "ใช้งาน", "", False),
    ("V9", "ตอนบันทึก", "ยอดใบแจ้งหนี้ต่างจาก PO (ระดับเตือน)", "เตือน", "> 2%", "ยอดต่างจาก PO 1.5% — ระบุเหตุผล", "รอทีมยืนยันเกณฑ์",
     "AC ต้องยืนยัน 2% ว่าเหมาะสม", True),
    ("V10", "ตอนบันทึก", "ยอดใบแจ้งหนี้ต่างจาก PO (ระดับบล็อก)", "บล็อก", "> 5%", "ต่างเกิน 5% ต้องให้จัดซื้อแก้ที่ต้นทางก่อน", "รอทีมยืนยันเกณฑ์",
     "AC ต้องยืนยัน 5% ว่าเหมาะสม", True),
    ("V11", "ตอนบันทึก", "วันที่ในเอกสารต้องไม่อยู่ในอนาคต", "เตือน", "≤ วันนี้", "วันที่ในใบแจ้งหนี้เป็นวันข้างหน้า", "ใช้งาน", "", False),
    ("V12", "ตอนบันทึก", "เลขที่ใบแจ้งหนี้ไม่ซ้ำต่อผู้ขายเดียวกัน", "บล็อก", "ไม่ซ้ำ", "เลขใบแจ้งหนี้นี้ใช้แล้วใน PO-26-0031", "ใช้งาน",
     "เทียบกับ OPCH.NumAtCard ของ B1 ด้วย", False),
    ("V13", "เปลี่ยนขั้น", "เอกสารบังคับของขั้นก่อนหน้าครบทุกใบ", "บล็อก", "ครบ 100%", "ยังขาดเอกสารจากขั้นก่อน", "ใช้งาน", "", False),
    ("V14", "เปลี่ยนขั้น", "ไม่มีใบเคลมที่ยังไม่ได้ใบลดหนี้", "บล็อก", "ไม่มีเคลมค้าง", "มีใบเคลมที่ยังไม่ได้รับใบลดหนี้", "ใช้งาน", "", False),
    ("V15", "เปลี่ยนขั้น", "ทุกงวดจ่ายเป็น PAID และมีสลิปครบ", "บล็อก", "ทุกงวด", "ยังจ่ายไม่ครบ", "ใช้งาน", "", False),
    ("V16", "เปลี่ยนขั้น", "ตรวจ 3 ทางผ่าน — เทียบกับจำนวนที่ QC รับ", "บล็อก", "PO x QC รับ x ใบแจ้งหนี้", "ตรวจ 3 ทางยังไม่ผ่าน", "ใช้งาน",
     "สำคัญ: เทียบกับจำนวนที่ QC รับ ไม่ใช่จำนวนที่ส่งมา", False),
    ("V17", "เปลี่ยนขั้น", "ผู้อนุมัติต้องไม่ใช่ผู้สร้าง/ผู้อัพโหลด", "บล็อก", "คนละคน", "อนุมัติเอกสารของตัวเองไม่ได้", "ใช้งาน", "", False),
    ("V18", "เปลี่ยนขั้น", "ต้องมีรหัสสินค้าจริงจาก B1 แล้ว", "บล็อก", "sap_item_code ไม่ว่าง", "ยังไม่มีรหัสสินค้าจาก B1", "ใช้งาน", "", False),
    ("P1", "ขณะพิมพ์", "วันที่จ่ายต้องไม่เป็นวันในอนาคต", "บล็อก", "≤ วันนี้", "วันที่จ่ายเป็นวันข้างหน้า — บันทึกได้เมื่อจ่ายจริงแล้ว", "ใช้งาน", "", False),
    ("P2", "ขณะพิมพ์", "วันที่ธนาคารตัดเงิน ≥ วันที่จ่าย", "บล็อก", "≥ payment_date", "วันที่ธนาคารตัดเงินต้องไม่ก่อนวันที่จ่าย", "ใช้งาน", "", False),
    ("P3", "ขณะพิมพ์", "จำนวนเงินที่จ่ายต้องมากกว่า 0", "บล็อก", "> 0", "จำนวนเงินต้องมากกว่า 0", "ใช้งาน", "", False),
    ("P4", "ขณะพิมพ์", "อัตราภาษีหัก ณ ที่จ่ายอยู่ในช่วงปกติ", "บล็อก", "0–15%", "อัตราหัก ณ ที่จ่ายผิดปกติ", "รอทีมยืนยันเกณฑ์",
     "รออัตราจริงที่บริษัทใช้", True),
    ("P5", "ตอนบันทึก", "ต้องแนบสลิป / Swift", "บล็อก", "บังคับ", "ต้องแนบสลิปหรือ Swift ก่อนบันทึกการจ่าย", "ใช้งาน", "", False),
    ("P6", "ตอนบันทึก", "ต้องมีเลขอ้างอิงธนาคาร", "บล็อก", "บังคับ", "ต้องใส่เลขอ้างอิงเพื่อกระทบยอดกับ statement", "ใช้งาน", "", False),
    ("P7", "ตอนบันทึก", "ยอดที่จ่ายต่างจากที่คำนวณ", "เตือน", "> 1 บาท", "ยอดต่างจากที่คำนวณ — ระบุเหตุผล", "ใช้งาน",
     "เหตุผลขึ้นในรายงานของหัวหน้าบัญชี", True),
    ("P8", "ตอนบันทึก", "ยอดสะสมที่จ่ายต้องไม่เกินมูลค่า PO", "บล็อก", "≤ มูลค่า PO", "จ่ายรวมเกินมูลค่า PO", "ใช้งาน", "", False),
    ("P9", "ตอนบันทึก", "จ่ายย้อนหลังเกินกำหนด", "เตือน", "> 7 วัน", "จ่ายย้อนหลัง X วัน — ต้องขออนุมัติเพิ่ม", "รอทีมยืนยันเกณฑ์",
     "AC ต้องยืนยันว่า 7 วันเหมาะสม", True),
    ("P10", "ตอนบันทึก", "งวดบัญชีต้องยังไม่ปิด", "บล็อก", "posting_period ยังเปิด", "งวดบัญชีปิดแล้ว — บันทึกในงวดปัจจุบันแทน", "ใช้งาน", "", False),
    ("P11", "ตอนบันทึก", "เลขอ้างอิงธนาคารห้ามซ้ำ", "บล็อก", "ไม่ซ้ำทั้งระบบ", "เลขอ้างอิงนี้ใช้แล้วใน PO-26-0031 งวด 1", "ใช้งาน",
     "กฎที่ป้องกันการจ่ายซ้ำ — ความผิดพลาดที่แพงที่สุดในงาน AP", False),
    ("P12", "ตอนบันทึก", "งวดต้องอยู่ในสถานะ APPROVED", "บล็อก", "= APPROVED", "งวดนี้ยังไม่ผ่านการอนุมัติจ่าย", "ใช้งาน", "", False),

    ("C1", "ตอนบันทึก", "ซื้อเงินสดต้องระบุเหตุผลที่ไม่เปิด PO", "บล็อก", "ต้องมีข้อความ",
     "ต้องบอกเหตุผลที่ไม่เปิด PO — ผู้ตรวจสอบบัญชีถามทุกครั้ง", "ใช้งาน",
     "ช่องนี้คือสิ่งเดียวที่ทำให้การซื้อเงินสดตรวจสอบได้", False),
    ("C2", "ตอนบันทึก", "วันที่จ่ายเงินสดต้องไม่เป็นวันข้างหน้า", "บล็อก", "≤ วันนี้",
     "วันที่จ่ายเป็นวันข้างหน้า — บันทึกได้เมื่อจ่ายจริงแล้ว", "ใช้งาน", "", False),
    ("C3", "ตอนบันทึก", "ข้าม QC ต้องระบุเหตุผลและผู้อนุมัติ", "บล็อก", "ต้องมีทั้ง 2 ช่อง",
     "ข้าม QC ต้องระบุเหตุผลและชื่อผู้อนุมัติ — เก็บเป็นหลักฐานถาวร", "ใช้งาน",
     "หัวหน้า QC เห็นทุกใบที่ข้ามในทะเบียนรายการเงินสด", False),
    ("C4", "ตอนบันทึก", "ยอดซื้อเงินสดต่อรายการต้องไม่เกินเพดาน", "บล็อก", "≤ 50,000 บาท",
     "เงินสดเกิน 50,000 บาท ต้องเปิด PO", "รอทีมยืนยันเกณฑ์",
     "ผู้บริหาร + บัญชี ต้องยืนยันเพดานนี้ · ตั้งไว้เท่าวงเงินอนุมัติจ่าย", True),
    ("C5", "ขณะพิมพ์", "ร้าน / สินค้า / จำนวน / ยอดเงิน ต้องไม่ว่าง", "บล็อก", "ครบทั้ง 4 ช่อง",
     "ยังกรอกไม่ครบ", "ใช้งาน", "", False),

    ("U1", "ตอนบันทึก", "ขั้นที่มีเอกสารบังคับ ต้องแนบไฟล์ก่อนส่งต่อ", "บล็อก", "ครบตาม Doc_Templates",
     "ขั้นนี้ต้องแนบใบรับสินค้า (GRPO) ก่อนถึงจะส่งต่อได้", "ใช้งาน",
     "แทนพฤติกรรมเดิมที่กดปุ่มเสร็จแล้วส่งต่อทันทีโดยไม่มีหลักฐาน", False),
    ("U2", "ตอนบันทึก", "ขั้นที่ไม่แนบไฟล์ ต้องเขียนหมายเหตุแทน", "บล็อก", "ติ๊กไม่มีเอกสาร + มีหมายเหตุ",
     "ไม่แนบไฟล์ต้องเขียนหมายเหตุว่าทำอะไรไปในขั้นนี้", "ใช้งาน", "", False),

    ("S1", "ขณะพิมพ์", "ยอดเงินที่ขอจ่ายต้องมากกว่า 0", "บล็อก", "> 0",
     "ต้องระบุยอดเงินที่ต้องจ่าย", "ใช้งาน", "ช่องบังคับตามที่ทีมขอ", False),
    ("S2", "ตอนบันทึก", "ยอดที่ขอต้องไม่เกินยอดคงเหลือของ PO", "บล็อก", "≤ ยอดคงเหลือ",
     "ยอดที่ขอเกินยอดคงเหลือของ PO", "ใช้งาน", "กันขอจ่ายเกินมูลค่าที่สั่งซื้อ", False),
    ("S3", "ขณะพิมพ์", "ต้องระบุ Effective date", "บล็อก", "ต้องมีค่า",
     "ต้องระบุ Effective date", "ใช้งาน", "ช่องบังคับตามที่ทีมขอ", False),
    ("S4", "ตอนบันทึก", "Effective date ต้องไม่ย้อนหลัง", "บล็อก", "≥ วันนี้",
     "Effective date เป็นวันย้อนหลัง — บัญชีทำจ่ายตามวันนั้นไม่ได้แล้ว", "ใช้งาน", "", False),
    ("S5", "ตอนบันทึก", "PO ที่มีใบเคลมค้าง ขอจ่ายไม่ได้", "บล็อก", "ไม่มีเคลมเปิดอยู่",
     "PO นี้มีใบเคลมค้าง — ขอจ่ายไม่ได้จนกว่าจะได้ใบลดหนี้", "ใช้งาน", "", False),
    ("S6", "ตอนบันทึก", "เอกสารบังคับต้องครบก่อนขอจ่าย", "บล็อก", "ครบทุกใบที่ is_required",
     "ยังขาดเอกสาร: ใบแจ้งหนี้ — หัวหน้าจะตีกลับ", "ใช้งาน", "", False),
    ("S7", "ตอนบันทึก", "ติ๊กประทับข้อความแล้ว ข้อความต้องไม่ว่าง", "บล็อก", "ต้องมีข้อความ",
     "ติ๊กประทับข้อความแล้วแต่ข้อความว่าง", "ใช้งาน", "", False),
]
r = 5
for code, when, rule, sev, crit, msg, status, note, hot in RULES:
    cell(ws, r, 1, code, bg=C_LOCK, al="center", bold=True)
    cell(ws, r, 2, when, al="center")
    cell(ws, r, 3, rule)
    cell(ws, r, 4, sev, bg=C_EDIT, al="center")
    cell(ws, r, 5, crit, bg=C_WARN_BG if hot else C_EDIT, fg=C_WARN_FG if hot else C_TITLE, al="center", bold=hot)
    cell(ws, r, 6, msg)
    cell(ws, r, 7, status, bg=C_EDIT, al="center")
    cell(ws, r, 8, note, bg=C_EDIT)
    ws.row_dimensions[r].height = 28
    r += 1
RULE_LAST = r - 1
dv(ws, "SEV", "D5:D%d" % RULE_LAST)
dv(ws, "RULEST", "G5:G%d" % RULE_LAST)
ws.freeze_panes = "C5"
ws.auto_filter.ref = "A4:H%d" % RULE_LAST

r += 1
for lab, formula in [("กฎทั้งหมด", '=COUNTA(A5:A%d)' % RULE_LAST),
                     ("ระดับบล็อก", '=COUNTIF(D5:D%d,"บล็อก")' % RULE_LAST),
                     ("ระดับเตือน", '=COUNTIF(D5:D%d,"เตือน")' % RULE_LAST),
                     ("รอทีมยืนยันเกณฑ์", '=COUNTIF(G5:G%d,"รอทีมยืนยันเกณฑ์")' % RULE_LAST)]:
    cell(ws, r, 2, lab, bold=True)
    cell(ws, r, 4, formula, bg=C_LOCK, bold=True, al="center")
    r += 1

# ============================================================
# 7. สิทธิ์แต่ละแผนก
# ============================================================
ws = wb.create_sheet("สิทธิ์แต่ละแผนก")
title(ws, "สิทธิ์แต่ละแผนก",
      "C=สร้าง · R=ดู · U=แก้ · A=อนุมัติ · — =ไม่มีสิทธิ์ · แก้ได้ทั้งตาราง ถ้าเห็นว่าไม่ตรงกับงานจริง", 11)
widths(ws, [30, 9, 11, 9, 11, 9, 9, 9, 11, 9, 30])
H = ["หน้าที่", "SR", "SR Head", "QC", "QC Head", "LS", "WH", "AC", "AC Head", "GM", "หมายเหตุของทีม"]
for i, h in enumerate(H, 1):
    cell(ws, 4, i, h)
hdr(ws, 4, 11)

PERM = [
    ("สร้าง/แก้ PO", "C R U", "R U", "R", "R", "R", "R", "R", "R", "R", ""),
    ("ยอดเงิน / ราคา", "R U", "R U", "—", "—", "—", "—", "R U", "R U", "R", "QC/LS/WH ไม่เห็นช่องราคาเลย ไม่ใช่แค่แก้ไม่ได้"),
    ("ใบขอราคา / ใบเสนอราคา", "C R U", "R U", "—", "—", "—", "—", "R", "R", "R", "เพิ่มใหม่ล่าสุด"),
    ("อัพโหลด PI / PO", "C U", "C U", "—", "—", "—", "—", "R", "R", "R", ""),
    ("อนุมัติ PO", "—", "A", "—", "—", "—", "—", "—", "—", "A", ""),
    ("ผลตรวจ QC", "R", "R", "C U", "C U A", "R", "R", "R", "R", "R", ""),
    ("เปิดรหัสสินค้า", "C (ขอ)", "A", "—", "—", "U", "—", "—", "—", "—", "LS เป็นคนบันทึกรหัสจาก B1 กลับ"),
    ("จองรถ", "R", "R", "R", "R", "C U", "R", "—", "—", "R", ""),
    ("GRPO", "R", "R", "R", "R", "R", "C U", "R", "R", "R", ""),
    ("ใบแจ้งหนี้", "R", "R", "—", "—", "—", "—", "C U", "R", "R", "รอยืนยันว่า SR หรือ AC เป็นคนอัพ"),
    ("ตั้งเรื่องขอจ่าย", "—", "—", "—", "—", "—", "—", "C", "R", "R", ""),
    ("อนุมัติจ่าย", "—", "—", "—", "—", "—", "—", "—", "A", "A", "เกินวงเงินต้องถึงผู้บริหาร"),
    ("บันทึกการจ่าย + สลิป", "R", "R", "—", "—", "—", "—", "C U", "R U", "R", "ต้องคนละคนกับผู้อนุมัติ"),
    ("กลับรายการจ่าย", "—", "—", "—", "—", "—", "—", "—", "A", "R", "หัวหน้าบัญชีเท่านั้น"),
    ("เคลม", "R U", "A", "C U", "A", "—", "R", "R", "R", "R", ""),
    ("ปลดล็อกเคลม", "—", "—", "—", "A", "—", "—", "—", "—", "—", "QC ที่ตรวจ ปลดล็อกของตัวเองไม่ได้"),
    ("ตรวจ 3 ทาง / ปิด PO", "R", "R", "—", "—", "—", "—", "C U", "A", "R", ""),
    ("ดูรายงานทั้งบริษัท", "—", "R", "—", "R", "—", "—", "R", "R", "R", ""),
    ("แจ้งปัญหา (bug)", "C", "C", "C", "C", "C", "C", "C", "C", "C", "ทุกคนแจ้งได้จากทุกหน้า"),
    ("ตั้งค่าระบบ / ผู้ใช้", "—", "—", "—", "—", "—", "—", "—", "—", "—", "IT/แอดมินเท่านั้น และไม่อยู่ในสายอนุมัติ"),
]
r = 5
for row in PERM:
    cell(ws, r, 1, row[0], bold=True)
    for i, v in enumerate(row[1:10], start=2):
        bg = C_LOCK if v == "—" else C_EDIT
        cell(ws, r, i, v, bg=bg, al="center")
    cell(ws, r, 11, row[10], bg=C_EDIT)
    ws.row_dimensions[r].height = 24
    r += 1
ws.freeze_panes = "B5"

r += 1
c = cell(ws, r, 1, "กฎแยกหน้าที่ที่บังคับด้วยโปรแกรม ไม่ใช่ด้วยความไว้ใจ: "
                    "ผู้สร้างอนุมัติงานตัวเองไม่ได้ · ผู้ตั้งเรื่องขอจ่าย ≠ ผู้อนุมัติ ≠ ผู้บันทึกจ่าย · "
                    "QC ที่ตรวจ ปลดล็อกเคลมของตัวเองไม่ได้ · แอดมินระบบไม่อยู่ในสายอนุมัติ",
         bg=C_GOOD_BG, fg=C_GOOD_FG)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
ws.row_dimensions[r].height = 34

# ============================================================
# 8. การแจ้งเตือน
# ============================================================
ws = wb.create_sheet("การแจ้งเตือน")
title(ws, "การแจ้งเตือน",
      "ภายในใช้ Lark อย่างเดียว (ตัดสินใจแล้ว) · ชั้น 1 ในเว็บทุกเรื่อง · ชั้น 2 Lark เฉพาะที่ต้องลงมือ · ชั้น 3 อีเมลสรุป", 6)
widths(ws, [40, 30, 22, 22, 16, 32])
H = ["เหตุการณ์", "ใครได้รับ", "ช่องทาง", "เมื่อไหร่", "เปิดใช้", "หมายเหตุของทีม"]
for i, h in enumerate(H, 1):
    cell(ws, 4, i, h)
hdr(ws, 4, 6)

NOTIF = [
    ("งานเข้าคิวของคุณ", "เจ้าของขั้นถัดไป", "1 + 2", "ทันที", "TRUE", "ปิดไม่ได้ — กันคนอ้างว่าไม่รู้"),
    ("เอกสารที่คุณต้องอัพยังไม่มา", "เจ้าของเอกสาร", "1 + 2", "ทุกเช้า 08:30", "TRUE", ""),
    ("ค้างเกิน SLA 100%", "ผู้ถืองาน", "1 + 2", "วันละครั้ง", "TRUE", ""),
    ("ค้างเกิน SLA 150%", "ผู้ถืองาน + หัวหน้าแผนก", "1 + 2 + 3", "วันละครั้ง", "TRUE", ""),
    ("ค้างเกิน SLA 200%", "+ ผู้บริหาร (ในสรุปรายวัน)", "3 อีเมล", "08:30", "TRUE", ""),
    ("รออนุมัติ", "ผู้อนุมัติตามวงเงิน", "1 + 2", "ทันที", "TRUE", ""),
    ("ตรวจรับไม่ผ่าน / เปิดเคลม", "SR + AC + หัวหน้า QC", "1 + 2", "ทันที", "TRUE", "แจ้งพร้อมกัน 3 แผนก"),
    ("ได้รับใบลดหนี้ → ปลดล็อกจ่าย", "AC", "1 + 2", "ทันที", "TRUE", ""),
    ("งวดจ่ายครบกำหนดใน 3 วัน", "AC", "1 + 3", "08:30", "TRUE", ""),
    ("งวดเลยกำหนดแล้วยังไม่จ่าย", "AC + หัวหน้าบัญชี", "1 + 2 + 3", "ทุกวัน 08:30", "TRUE", ""),
    ("ตั้งเรื่องขอจ่าย", "ผู้อนุมัติตามวงเงิน", "1 + 2", "ทันที", "TRUE", ""),
    ("อนุมัติ / ตีกลับคำขอจ่าย", "ผู้ตั้งเรื่อง", "1 + 2", "ทันที", "TRUE", ""),
    ("บันทึกจ่ายแล้ว", "SR ผู้ดูแล PO", "1 + 2", "ทันที", "TRUE", ""),
    ("บันทึกจ่ายแล้ว — แจ้งผู้ขาย", "ผู้ขาย (อีเมล + สลิป)", "3 อีเมล", "ทันที", "TRUE",
     "ตัดขั้นที่ AC ส่งสลิปให้ SR แล้ว SR ส่งต่อ"),
    ("ธนาคารตีกลับ", "AC + หัวหน้าบัญชี", "1 + 2", "ทันที", "TRUE", ""),
    ("จ่ายครบทุกงวด", "SR + AC", "1 + 2", "ทันที", "TRUE", ""),
    ("ใบเสร็จตัวจริงยังไม่มาเกิน 15 วัน", "AC", "1 + 2", "ทุก 7 วัน", "TRUE", ""),
    ("การตรวจสอบไม่ผ่าน (ยอดไม่ตรง)", "ผู้กด + แผนกที่ต้องแก้", "1 + 2", "ทันที", "TRUE", ""),
    ("PO ปิดแล้ว", "ทุกคนที่เคยแตะ PO นี้", "1 ในเว็บ", "ทันที", "TRUE", ""),
    ("สรุปงานค้างของฉัน", "ทุกคน", "3 อีเมล", "08:30 จ–ศ", "TRUE", "ไม่มีงานค้าง = ไม่ส่ง"),
    ("แจ้งบั๊กระดับ «ทำงานต่อไม่ได้»", "กลุ่ม Lark ของ IT", "2 Lark", "ทันที", "TRUE", ""),
    ("Lark ส่งไม่สำเร็จครบ 5 ครั้ง", "ผู้รับเดิม (ผ่านอีเมลแทน)", "3 อีเมล", "อัตโนมัติ", "TRUE",
     "จำเป็นเพราะไม่มีช่องทางที่สองแล้ว"),
    ("Lark ล้มเหลว > 10 ข้อความใน 15 นาที", "แอดมินระบบ + แบนเนอร์ในเว็บ", "1 + 3", "ทันที", "TRUE", ""),
    ("ดึงข้อมูลจาก B1 ล้มเหลว", "แอดมินระบบ", "1 + 2", "ทันที", "TRUE", ""),
]
r = 5
for ev, who, ch, when, on, note in NOTIF:
    cell(ws, r, 1, ev)
    cell(ws, r, 2, who)
    cell(ws, r, 3, ch, bg=C_EDIT, al="center")
    cell(ws, r, 4, when, bg=C_EDIT, al="center")
    cell(ws, r, 5, on, bg=C_EDIT, al="center")
    cell(ws, r, 6, note, bg=C_EDIT)
    ws.row_dimensions[r].height = 24
    r += 1
dv(ws, "TIER", "C5:C%d" % (r - 1))
ws.freeze_panes = "B5"

r += 1
cell(ws, r, 1, "กติกากันสแปม (สำคัญกว่าตัวระบบแจ้งเตือนเอง)", bold=True, size=11); r += 1
for g in [
    "ไม่แจ้งการกระทำของตัวเอง — ตัดข้อความไปได้ประมาณ 40%",
    "เรื่องเดิมไม่เด้งซ้ำใน 24 ชม. (dedupe_key = เหตุการณ์ + เอกสาร + ผู้รับ + วัน)",
    "เกิน 3 เรื่องใน 10 นาที รวมเป็นข้อความเดียว",
    "เงียบ 20:00–07:00 และวันหยุด ยกเว้นเรื่องด่วน",
    "ผู้ใช้ปิดรายประเภทได้เอง ยกเว้น «งานเข้าคิวของคุณ» ที่ปิดไม่ได้",
]:
    cell(ws, r, 1, "• " + g)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1

# ============================================================
# 9. สถานะ
# ============================================================
ws = wb.create_sheet("สถานะ")
title(ws, "สถานะทั้งหมดของแต่ละเอกสาร", "สถานะไหนเปลี่ยนไปไหนได้บ้าง และใครเป็นคนทำให้เปลี่ยน", 6)
widths(ws, [22, 22, 30, 34, 22, 30])
H = ["เอกสาร", "สถานะ", "หมายความว่า", "เปลี่ยนไปสถานะไหนได้", "ใครทำให้เปลี่ยน", "หมายเหตุของทีม"]
for i, h in enumerate(H, 1):
    cell(ws, 4, i, h)
hdr(ws, 4, 6)

STATES = [
    ("Purchase_Orders", "ACTIVE", "กำลังดำเนินการ", "COMPLETED / CANCELLED", "ระบบ / SR", ""),
    ("Purchase_Orders", "COMPLETED", "ปิดบัญชีแล้ว", "(สิ้นสุด)", "AC", "ปิดได้เมื่อผ่านกฎ V13–V18 ครบ"),
    ("Purchase_Orders", "CANCELLED", "ยกเลิก", "(สิ้นสุด)", "SR + หัวหน้า", ""),
    ("PO_Payments", "PENDING", "รอตั้งเรื่องขอจ่าย", "REQUESTED / BLOCKED / CANCELLED", "ระบบ", ""),
    ("PO_Payments", "BLOCKED", "ล็อกเพราะมีเคลมค้าง", "PENDING", "ระบบ (เมื่อได้ใบลดหนี้)", ""),
    ("PO_Payments", "REQUESTED", "ตั้งเรื่องแล้ว รออนุมัติ", "APPROVED / PENDING", "AC ตั้ง · ผู้อนุมัติตัดสิน", ""),
    ("PO_Payments", "APPROVED", "อนุมัติแล้ว รอทำจ่าย", "PAID / PENDING", "หัวหน้าบัญชี / ผู้บริหาร", ""),
    ("PO_Payments", "PAID", "จ่ายแล้ว มีสลิป", "FAILED / VOID", "AC (คนละคนกับผู้อนุมัติ)", ""),
    ("PO_Payments", "FAILED", "ธนาคารตีกลับ", "APPROVED", "AC", ""),
    ("PO_Payments", "VOID", "กลับรายการเพราะจ่ายผิด", "(สิ้นสุด — เปิดงวดใหม่แทน)", "หัวหน้าบัญชี", "ไม่มีการลบ ของเดิมอยู่เป็นหลักฐาน"),
    ("PO_Payments", "CANCELLED", "ยกเลิกงวด", "(สิ้นสุด)", "หัวหน้าบัญชี", ""),
    ("Claims", "OPEN", "เปิดเรื่องแล้ว", "SENT_TO_SUPPLIER", "QC / Sales Coordinator", ""),
    ("Claims", "SENT_TO_SUPPLIER", "ส่งเคลมให้ผู้ขายแล้ว", "CN_RECEIVED", "SR", ""),
    ("Claims", "CN_RECEIVED", "ได้รับใบลดหนี้", "CLOSED", "AC", "ตัวปลดล็อกงวดจ่าย"),
    ("Claims", "CLOSED", "ปิดเรื่อง", "(สิ้นสุด)", "หัวหน้า QC", ""),
    ("ItemCode_Requests", "DRAFT", "ร่าง", "WAIT_APPROVE", "SR", ""),
    ("ItemCode_Requests", "WAIT_APPROVE", "รอหัวหน้าอนุมัติ", "WAIT_LS / REJECTED", "หัวหน้า SR", ""),
    ("ItemCode_Requests", "WAIT_LS", "รอ LS เปิดรหัสใน B1", "DONE", "LS", ""),
    ("ItemCode_Requests", "DONE", "เปิดรหัสแล้ว", "(สิ้นสุด)", "LS", "PO ที่รออยู่เดินต่อเองอัตโนมัติ"),
    ("ItemCode_Requests", "REJECTED", "ไม่อนุมัติ", "(สิ้นสุด)", "หัวหน้า SR", ""),
    ("Price_Requests", "NEW", "ยังไม่มีใบเสนอราคา", "QUOTING / CANCELLED", "SR", "เพิ่มใหม่ล่าสุด"),
    ("Price_Requests", "QUOTING", "กำลังเทียบราคา", "CONVERTED / CANCELLED", "SR", "เพิ่มใหม่ล่าสุด"),
    ("Price_Requests", "CONVERTED", "เปิด PO แล้ว", "(สิ้นสุด)", "SR", "เพิ่มใหม่ล่าสุด · เพิ่มใบเสนอราคาใหม่ไม่ได้แล้ว"),
    ("Price_Requests", "CANCELLED", "ยกเลิก", "(สิ้นสุด)", "SR", "เพิ่มใหม่ล่าสุด"),
    ("Payment_Requests", "PENDING", "รอหัวหน้าอนุมัติ", "APPROVED / REJECTED",
     "หัวหน้าบัญชี หรือผู้บริหารตามวงเงิน", "เพิ่มใหม่ล่าสุด"),
    ("Payment_Requests", "APPROVED", "อนุมัติแล้ว รอบัญชีทำจ่าย", "(สิ้นสุด)",
     "หัวหน้าบัญชี / ผู้บริหาร", "เพิ่มใหม่ล่าสุด · งวดจ่ายเปลี่ยนเป็น APPROVED ตาม"),
    ("Payment_Requests", "REJECTED", "ตีกลับพร้อมเหตุผล", "(สิ้นสุด — ขอใหม่เป็นคำขอใบใหม่)",
     "หัวหน้าบัญชี / ผู้บริหาร", "เพิ่มใหม่ล่าสุด · ของเดิมไม่ถูกลบ"),
    ("Issues", "NEW", "รอ IT รับเรื่อง", "ACK", "IT (SLA 1 วันทำการ)", ""),
    ("Issues", "ACK", "รับเรื่องแล้ว", "FIXING / WONTFIX", "IT", ""),
    ("Issues", "FIXING", "กำลังแก้", "FIXED", "IT", ""),
    ("Issues", "FIXED", "แก้แล้ว รอผู้แจ้งยืนยัน", "CLOSED / FIXING", "IT แก้ · ผู้แจ้งยืนยัน", "ผู้แจ้งเป็นคนกดปิด ไม่ใช่ IT"),
    ("Issues", "CLOSED", "ปิดเรื่อง", "(สิ้นสุด)", "ผู้แจ้ง", "ไม่มีปุ่มลบ"),
]
r = 5
prev = None
for tbl, st, mean, nxt, who, note in STATES:
    isnew = tbl in ("Price_Requests", "Payment_Requests")
    bg = C_NEW if isnew else None
    cell(ws, r, 1, tbl if tbl != prev else "", bold=(tbl != prev), bg=bg)
    cell(ws, r, 2, st, bg=C_LOCK, bold=True)
    cell(ws, r, 3, mean, bg=bg)
    cell(ws, r, 4, nxt, bg=C_EDIT)
    cell(ws, r, 5, who, bg=C_EDIT)
    cell(ws, r, 6, note, bg=C_EDIT)
    prev = tbl
    ws.row_dimensions[r].height = 24
    r += 1
ws.freeze_panes = "C5"
ws.auto_filter.ref = "A4:F%d" % (r - 1)

# ============================================================
# 10. ค่าตั้งต้น
# ============================================================
ws = wb.create_sheet("ค่าตั้งต้น")
title(ws, "ค่าตั้งต้นทั้งหมด (Config)",
      "ตัวเลขที่ปรับได้ทั้งระบบอยู่ในแท็บเดียว · แถวสีแดงคือค่าที่ผมตั้งเองและยังไม่มีใครยืนยัน", 6)
widths(ws, [30, 26, 20, 46, 20, 30])
H = ["รหัสค่า", "หมวด", "ค่า (แก้ได้)", "ใช้ทำอะไร", "ใครยืนยัน", "หมายเหตุของทีม"]
for i, h in enumerate(H, 1):
    cell(ws, 4, i, h)
hdr(ws, 4, 6)

CONFIG = [
    ("SLA_DEFAULT_HOURS", "ขั้นตอนงาน", "24", "SLA ตั้งต้นของขั้นที่ไม่ได้กำหนดเอง", "หัวหน้าทุกแผนก", "", True),
    ("SLA_ESCALATE_1_PCT", "ขั้นตอนงาน", "150", "เกิน SLA กี่ % จึงแจ้งหัวหน้า", "หัวหน้าทุกแผนก", "", True),
    ("SLA_ESCALATE_2_PCT", "ขั้นตอนงาน", "200", "เกิน SLA กี่ % จึงเข้าสรุปของผู้บริหาร", "ผู้บริหาร", "", True),
    ("INVOICE_TOLERANCE_WARN_PCT", "ตรวจสอบ", "2", "ยอดใบแจ้งหนี้ต่างจาก PO กี่ % จึงเตือน", "AC", "กฎ V9", True),
    ("INVOICE_TOLERANCE_BLOCK_PCT", "ตรวจสอบ", "5", "ต่างกี่ % จึงบล็อก", "AC", "กฎ V10", True),
    ("PAYMENT_VARIANCE_TOLERANCE_THB", "จ่ายเงิน", "1", "ยอดจ่ายต่างจากที่คำนวณกี่บาทจึงต้องใส่เหตุผล", "AC", "กฎ P7", True),
    ("PAYMENT_BACKDATE_DAYS", "จ่ายเงิน", "7", "จ่ายย้อนหลังได้กี่วันโดยไม่ต้องขออนุมัติเพิ่ม", "AC", "กฎ P9", True),
    ("APPROVAL_LIMIT_AC_HEAD_THB", "จ่ายเงิน", "50000", "วงเงินที่หัวหน้าบัญชีอนุมัติได้เอง", "AC + ผู้บริหาร", "เกินนี้ต้องถึงผู้บริหาร", True),
    ("SWIFT_ALWAYS_GM", "จ่ายเงิน", "TRUE", "จ่ายต่างประเทศต้องถึงผู้บริหารทุกวงเงิน", "AC + ผู้บริหาร", "", True),
    ("WHT_RATE_GOODS", "ภาษี", "0", "อัตราหัก ณ ที่จ่าย — ซื้อสินค้า (%)", "AC", "", True),
    ("WHT_RATE_FREIGHT", "ภาษี", "1", "อัตราหัก ณ ที่จ่าย — ค่าขนส่ง (%)", "AC", "", True),
    ("WHT_RATE_SERVICE", "ภาษี", "3", "อัตราหัก ณ ที่จ่าย — ค่าบริการ (%)", "AC", "", True),
    ("WHT_RATE_RENT", "ภาษี", "5", "อัตราหัก ณ ที่จ่าย — ค่าเช่า (%)", "AC", "", True),
    ("WHT_RATE_ADS", "ภาษี", "2", "อัตราหัก ณ ที่จ่าย — ค่าโฆษณา (%)", "AC", "", True),
    ("BANK_FEE_DEFAULT_THB", "จ่ายเงิน", "35", "ค่าธรรมเนียมโอนตั้งต้น", "AC", "", True),
    ("BANK_FEE_BORNE_BY", "จ่ายเงิน", "OUR", "ใครรับภาระค่าธรรมเนียมโดยปกติ", "AC", "OUR หรือ SUPPLIER", True),
    ("FILE_MAX_SIZE_MB", "ไฟล์", "20", "ขนาดไฟล์สูงสุดที่รับ", "IT", "กฎ V5", False),
    ("IMAGE_RESIZE_LONG_EDGE_PX", "ไฟล์", "1600", "ย่อรูปด้านยาวเหลือกี่พิกเซลก่อนอัพ", "IT", "ลดพื้นที่ ~94%", False),
    ("IMAGE_JPEG_QUALITY", "ไฟล์", "0.75", "คุณภาพ JPEG หลังบีบอัด", "IT", "", False),
    ("SCAN_DPI_RECOMMENDED", "ไฟล์", "200", "ตั้งค่าเครื่องสแกน (dpi, เฉดเทา)", "IT + ธุรการ", "ลดจาก 300 dpi สี", False),
    ("DOC_VERSIONS_KEEP", "ไฟล์", "3", "เก็บกี่เวอร์ชันล่าสุด (ฉบับที่อนุมัติเก็บเสมอ)", "IT", "", False),
    ("DOC_VERSION_PURGE_DAYS", "ไฟล์", "90", "ลบเวอร์ชันเกินโควตาหลังกี่วัน", "IT", "", False),
    ("ARCHIVE_AFTER_MONTHS", "ไฟล์", "12", "PO ปิดเกินกี่เดือนจึงย้ายเข้าคลัง", "AC", "", True),
    ("DOC_RETENTION_YEARS", "ไฟล์", "5", "เก็บเอกสารกี่ปีตามกฎหมายบัญชี", "AC + ที่ปรึกษากฎหมาย", "ต้องยืนยันว่าใช้กับทุกประเภทเอกสาร", True),
    ("QUIET_HOURS_START", "แจ้งเตือน", "20:00", "เริ่มเวลาเงียบ", "หัวหน้าทุกแผนก", "", True),
    ("QUIET_HOURS_END", "แจ้งเตือน", "07:00", "จบเวลาเงียบ", "หัวหน้าทุกแผนก", "", True),
    ("DAILY_DIGEST_TIME", "แจ้งเตือน", "08:30", "เวลาส่งสรุปรายวัน", "หัวหน้าทุกแผนก", "", True),
    ("NOTIF_DEDUPE_HOURS", "แจ้งเตือน", "24", "เรื่องเดิมไม่เด้งซ้ำภายในกี่ชั่วโมง", "IT", "", False),
    ("NOTIF_BATCH_THRESHOLD", "แจ้งเตือน", "3", "เกินกี่เรื่องใน 10 นาทีจึงรวมเป็นข้อความเดียว", "IT", "", False),
    ("NOTIF_RETRY_MAX", "แจ้งเตือน", "5", "ลองส่งซ้ำสูงสุดกี่ครั้งก่อนตกไปที่อีเมล", "IT", "1 → 5 → 25 นาที", False),
    ("SYNC_INTERVAL_MINUTES", "เชื่อมต่อ B1", "30", "ดึงข้อมูลธุรกรรมจาก B1 ทุกกี่นาที", "IT + ผู้ดูแล B1", "", False),
    ("SYNC_MASTER_TIME", "เชื่อมต่อ B1", "02:00", "ดึงข้อมูล master (ผู้ขาย/สินค้า) ตอนกี่โมง", "IT + ผู้ดูแล B1", "", False),
    ("SYNC_OVERLAP_DAYS", "เชื่อมต่อ B1", "1", "ย้อนดึงซ้ำกี่วันกันตกหล่น", "IT", "UpdateDate ของ B1 เก็บระดับวัน", False),
    ("SYNC_STALE_WARN_MINUTES", "เชื่อมต่อ B1", "90", "ข้อมูลเก่ากว่ากี่นาทีจึงขึ้นแบนเนอร์เตือน", "IT", "", False),
    ("ISSUE_ACK_SLA_HOURS", "แจ้งบั๊ก", "8", "IT ต้องรับเรื่องภายในกี่ชั่วโมงทำการ", "IT", "", True),

    ("DRIVE_ROOT_NAME", "ไฟล์และไดรฟ์", "MGS-Documents", "ชื่อ Shared drive กลางที่เดียวของทั้งบริษัท",
     "IT", "ย้ายไดรฟ์ = ย้ายไฟล์ทั้งระบบ ต้องวางแผนก่อน", False),
    ("CASH_LIMIT", "ซื้อเงินสด", "50,000", "เพดานยอดซื้อเงินสดต่อรายการ เกินกว่านี้ต้องเปิด PO (กฎ C4)",
     "ผู้บริหาร + AC", "ตั้งไว้เท่าวงเงินอนุมัติจ่าย — ยังไม่มีใครยืนยัน", True),
    ("CASH_QC_BYPASS_ENABLED", "ซื้อเงินสด", "เปิด", "อนุญาตให้ติ๊กข้าม QC ได้หรือไม่",
     "หัวหน้า QC", "ปิด = ห้ามข้าม QC ทุกกรณี ต้องตรวจทุกล็อต", True),
    ("MERGE_ON_COMPLETE", "รวมไฟล์", "เปิด", "รวมเอกสารเป็นไฟล์เดียวอัตโนมัติเมื่อปิดรายการ",
     "AC", "", False),
    ("MERGE_KEEP_DAYS", "รวมไฟล์", "90", "เก็บต้นฉบับต่ออีกกี่วันหลังรวมเล่ม แล้วจึงลบ (0 = ไม่ลบเลย)",
     "AC + ผู้สอบบัญชี", "ค่านี้คือสิ่งที่ทำให้พื้นที่ไม่บวมจากการรวมไฟล์ — ยังไม่มีใครยืนยัน", True),
    ("STAMP_DEFAULT_POSITION", "ขอทำจ่าย", "BR", "ตำแหน่งเริ่มต้นของข้อความที่ประทับ (TL/TR/BR/CT)",
     "AC", "", False),
    ("STAMP_DEFAULT_SIZE_PT", "ขอทำจ่าย", "13", "ขนาดตัวอักษรเริ่มต้นของข้อความที่ประทับ",
     "AC", "", False),
]
r = 5
for key, cat, val, use, who, note, hot in CONFIG:
    cell(ws, r, 1, key, bg=C_LOCK)
    cell(ws, r, 2, cat, al="center")
    cell(ws, r, 3, val, bg=C_WARN_BG if hot else C_EDIT, fg=C_WARN_FG if hot else C_TITLE, al="center", bold=True)
    cell(ws, r, 4, use)
    cell(ws, r, 5, who, bg=C_EDIT, al="center")
    cell(ws, r, 6, note, bg=C_EDIT)
    ws.row_dimensions[r].height = 24
    r += 1
CFG_LAST = r - 1
ws.freeze_panes = "B5"
ws.auto_filter.ref = "A4:F%d" % CFG_LAST

r += 1
cell(ws, r, 1, "ค่าที่ยังต้องให้ทีมยืนยัน", bold=True)
cell(ws, r, 3, str(sum(1 for c in CONFIG if c[6])), bg=C_LOCK, bold=True, al="center")
r += 1
cell(ws, r, 1, "ค่าทั้งหมด", bold=True)
cell(ws, r, 3, "=COUNTA(A5:A%d)" % CFG_LAST, bg=C_LOCK, bold=True, al="center")

# ============================================================
# 11. เชื่อมต่อ B1
# ============================================================
ws = wb.create_sheet("เชื่อมต่อ B1")
title(ws, "เชื่อมต่อ SAP Business One",
      "อ่านอย่างเดียว ไม่เขียนกลับ · ตัวส่งข้อมูลรันในออฟฟิศแล้วผลักออกมา ไม่ต้องเปิดพอร์ตเข้า", 7)
widths(ws, [24, 24, 28, 30, 16, 16, 30])
H = ["ข้อมูลในระบบเรา", "ตารางใน B1", "ฟิลด์ที่ใช้", "ใช้ทำอะไร", "ความถี่", "คีย์จับคู่", "หมายเหตุของทีม"]
for i, h in enumerate(H, 1):
    cell(ws, 4, i, h)
hdr(ws, 4, 7)

B1MAP = [
    ("PO (หัวเอกสาร)", "OPOR", "DocEntry, DocNum, CardCode, CardName, DocDate, DocDueDate, DocTotal, DocCur, GroupNum, DocStatus, CANCELED, UpdateDate",
     "เลข PO · ผู้ขาย · มูลค่า · เงื่อนไขชำระ", "ทุก 30 นาที", "DocEntry", ""),
    ("PO (บรรทัด)", "POR1", "LineNum, ItemCode, Dscription, Quantity, unitMsr, Price, LineTotal, OpenQty, WhsCode",
     "สินค้า จำนวน หน่วย · OpenQty บอกว่ายังค้างรับเท่าไหร่", "ทุก 30 นาที", "DocEntry + LineNum", ""),
    ("ผู้ขาย", "OCRD", "CardCode, CardName, LicTradNum, E_Mail, Phone1, Currency (CardType='S')",
     "ชื่อผู้ขายสะกดตรงกันทุกใบ + อีเมลส่งสลิป", "ทุกคืน", "CardCode", ""),
    ("สินค้า", "OITM", "ItemCode, ItemName, InvntryUom, ItmsGrpCod",
     "ยืนยันว่ารหัสสินค้าเปิดใน B1 แล้วจริง", "ทุกคืน", "ItemCode", ""),
    ("ใบรับสินค้า (GRPO)", "OPDN / PDN1", "DocNum, DocDate · ItemCode, Quantity, BaseEntry, BaseLine, BaseType",
     "BaseType = 22 คือโยงกลับไปใบสั่งซื้อ", "ทุก 30 นาที", "DocEntry + LineNum", "BaseType 22=PO, 20=GRPO, 18=A/P Invoice"),
    ("ใบแจ้งหนี้ผู้ขาย", "OPCH / PCH1", "DocNum, NumAtCard, DocDate, DocTotal, VatSum · BaseEntry, BaseType",
     "NumAtCard = เลขใบแจ้งหนี้ของผู้ขาย ใช้ตรวจ 3 ทาง + กันรับซ้ำ", "ทุก 30 นาที", "DocEntry", ""),
    ("การจ่ายเงินออก", "OVPM / VPM2", "DocNum, DocDate, TrsfrSum, TrsfrRef, TrsfrDate, TrsfrAcct · ใบแจ้งหนี้ที่จ่าย",
     "เทียบกับที่บันทึกในระบบเรา — เจอทันทีถ้าไม่ตรง", "ทุก 30 นาที", "DocEntry", "ของแถมที่มีค่ามาก วันนี้ไม่มีใครเห็น"),
    ("เงื่อนไขชำระ", "OCTG", "GroupNum, PymntGroup, ExtraDays, ExtraMonth, NumOfInst, StartFrom",
     "แปลงเป็นงวดจ่ายอัตโนมัติ — งวดในระบบไม่มีทางไม่ตรงกับสัญญา", "ทุกคืน", "GroupNum", "ขอรายการที่ใช้จริงจากผู้ดูแล B1"),
]
r = 5
for a, b, c, d, e, f, g in B1MAP:
    cell(ws, r, 1, a, bold=True)
    cell(ws, r, 2, b, bg=C_LOCK, al="center")
    cell(ws, r, 3, c, size=9)
    cell(ws, r, 4, d)
    cell(ws, r, 5, e, bg=C_EDIT, al="center")
    cell(ws, r, 6, f, bg=C_LOCK, al="center")
    cell(ws, r, 7, g, bg=C_EDIT)
    ws.row_dimensions[r].height = 46
    r += 1

r += 1
cell(ws, r, 1, "ไม่ดึงข้อมูลเหล่านี้ (ตั้งใจไม่ขอ เพื่อให้อนุมัติสิทธิ์ผ่านง่าย)", bold=True, size=11); r += 1
for x in ["บัญชีแยกประเภท (OJDT / JDT1)", "ต้นทุนสินค้า", "ข้อมูลลูกค้า", "ข้อมูลพนักงาน / เงินเดือน"]:
    cell(ws, r, 1, "• " + x)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 1

r += 1
cell(ws, r, 1, "กฎการซิงก์", bold=True, size=11); r += 1
for x in [
    "Upsert ด้วย DocEntry ของ B1 ไม่ใช่ insert — ดึงซ้ำกี่ครั้งก็ไม่เกิดข้อมูลซ้ำ",
    "ดึงเฉพาะที่ UpdateDate เปลี่ยน + ย้อนซ้ำ 1 วันกันตกหล่น (UpdateDate ของ B1 เก็บระดับวัน)",
    "ไฟล์ไม่ครบ = ข้ามรอบ ไม่เขียนทับของเดิม — ข้อมูลเก่าที่ถูกต้องดีกว่าข้อมูลใหม่ที่พัง",
    "ผู้ใช้ฐานข้อมูลมีสิทธิ์ SELECT เท่านั้น จำกัดเฉพาะ 7 ตารางข้างบน",
    "บันทึกทุกรอบใน Sync_Log · ตรวจนับรายวันว่าจำนวน PO เปิดอยู่ตรงกับ B1",
    "B1 ล่ม = ระบบยังใช้ได้ ขึ้นป้ายบอกว่าข้อมูลเก่าแค่ไหน ไม่บล็อกการทำงาน",
]:
    cell(ws, r, 1, "• " + x)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1

r += 1
c = cell(ws, r, 1, "ยังไม่ทราบ: B1 รุ่นย่อย (9.x / 10.0) และฐานข้อมูลเป็น MS SQL หรือ HANA — "
                    "ตัดสินว่าจะใช้ Service Layer (10.0 ขึ้นไป) หรืออ่าน SQL แบบ read-only (ได้ทุกรุ่น)",
         bg=C_WARN_BG, fg=C_WARN_FG, bold=True)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
ws.row_dimensions[r].height = 32

# ============================================================
# 12. ประเด็นค้าง
# ============================================================
ws = wb.create_sheet("ประเด็นค้าง")
title(ws, "ประเด็นที่ต้องตัดสินใจก่อนขึ้นระบบ",
      "รวบรวมจากทุกเอกสาร · ช่อง «คำตอบของทีม» และ «สถานะ» กรอกได้เลย — นี่คือแท็บที่ควรเปิดก่อนแท็บอื่น", 8)
widths(ws, [5, 16, 46, 30, 18, 34, 16, 12])
H = ["#", "หมวด", "ประเด็น", "ค่าที่ผมตั้งไว้ก่อน", "ใครตัดสิน", "คำตอบของทีม", "สถานะ", "ความสำคัญ"]
for i, h in enumerate(H, 1):
    cell(ws, 4, i, h)
hdr(ws, 4, 8)

ISSUES_LIST = [
    ("เชื่อมต่อ B1", "B1 รุ่นย่อยอะไร (9.x / 10.0) และฐานข้อมูล MS SQL หรือ HANA",
     "ยังไม่ทราบ", "ผู้ดูแล B1", "", "รอตัดสินใจ", "สูง"),
    ("เชื่อมต่อ B1", "ใครดูแล B1 และขอสิทธิ์อ่าน 7 ตารางต้องผ่านใคร",
     "ยังไม่ทราบ", "IT + ผู้บริหาร", "", "รอตัดสินใจ", "สูง"),
    ("เชื่อมต่อ B1", "มี job ส่งออกข้อมูลจาก B1 อยู่แล้วหรือไม่ (B1if / Crystal / SQL Agent)",
     "ยังไม่ทราบ", "ผู้ดูแล B1", "", "รอตัดสินใจ", "กลาง"),
    ("เชื่อมต่อ B1", "รายการเงื่อนไขชำระ (OCTG) ที่บริษัทใช้จริงมีกี่แบบ",
     "ยังไม่ทราบ", "AC + ผู้ดูแล B1", "", "รอตัดสินใจ", "สูง"),
    ("เชื่อมต่อ B1", "มีเครื่องในออฟฟิศที่เปิดตลอดและรัน scheduled task ได้หรือไม่",
     "ยังไม่ทราบ", "IT", "", "รอตัดสินใจ", "สูง"),
    ("เชื่อมต่อ B1", "ตารางงวดชำระต่อเอกสารใน B1 ชื่ออะไรในรุ่นที่ใช้",
     "ยังไม่ยืนยัน", "ผู้ดูแล B1", "", "รอตัดสินใจ", "กลาง"),
    ("ซื้อเงินสด", "เพดานยอดซื้อเงินสดต่อรายการควรเป็นเท่าไหร่",
     "50,000 บาท", "ผู้บริหาร + AC", "", "รอตัดสินใจ", "สูง"),
    ("ซื้อเงินสด", "ใครมีสิทธิ์อนุมัติการข้าม QC — หัวหน้า QC เท่านั้น หรือจัดซื้อระบุเองได้",
     "จัดซื้อระบุเอง แล้วหัวหน้า QC เห็นย้อนหลังทุกใบ", "หัวหน้า QC", "", "รอตัดสินใจ", "สูง"),
    ("ซื้อเงินสด", "เงินสดที่พนักงานสำรองจ่าย ต้องมีสายอนุมัติเบิกคืนแยกหรือไม่",
     "ยังไม่ทำ — บันทึกไว้เฉย ๆ", "AC", "", "รอตัดสินใจ", "กลาง"),
    ("รวมไฟล์", "ต้องเก็บไฟล์ต้นฉบับตามกฎหมายกี่ปี หรือเก็บไฟล์รวมพอ",
     "เก็บต้นฉบับ 90 วันหลังรวมเล่มแล้วลบ", "AC + ผู้สอบบัญชี", "", "รอตัดสินใจ", "สูง"),
    ("รวมไฟล์", "ไฟล์รวมต้องใส่รหัสผ่านหรือลายน้ำ «สำเนา» หรือไม่",
     "ไม่ใส่", "AC", "", "รอตัดสินใจ", "ต่ำ"),
    ("ขอทำจ่าย", "ข้อความที่ประทับลงเอกสารต้องมีลายเซ็นดิจิทัลหรือ QR ตรวจสอบหรือไม่",
     "ข้อความอย่างเดียว ไม่มี QR", "AC + ผู้สอบบัญชี", "", "รอตัดสินใจ", "กลาง"),
    ("จ่ายเงิน", "วงเงินที่หัวหน้าบัญชีอนุมัติได้เอง",
     "50,000 บาท", "AC + ผู้บริหาร", "", "รอตัดสินใจ", "สูง"),
    ("จ่ายเงิน", "อัตราภาษีหัก ณ ที่จ่ายที่บริษัทใช้จริง และประเภทรายจ่ายที่ต้องมี",
     "0/1/3/5/2%", "AC", "", "รอตัดสินใจ", "สูง"),
    ("จ่ายเงิน", "ค่าธรรมเนียมโอน ปกติบริษัทรับภาระหรือหักจากผู้ขาย",
     "บริษัทรับภาระ", "AC", "", "รอตัดสินใจ", "กลาง"),
    ("จ่ายเงิน", "จ่ายย้อนหลังได้กี่วันโดยไม่ต้องขออนุมัติเพิ่ม",
     "7 วัน", "AC", "", "รอตัดสินใจ", "กลาง"),
    ("จ่ายเงิน", "รายการบัญชีธนาคารที่ใช้จ่ายออก (ทำเป็นตัวเลือกแทนการพิมพ์)",
     "ยังไม่ทราบ", "AC", "", "รอตัดสินใจ", "กลาง"),
    ("จ่ายเงิน", "ปัจจุบันใครส่งสลิปให้ผู้ขาย และจะให้ระบบส่งแทนได้เลยไหม",
     "ให้ระบบส่ง", "AC + SR", "", "รอตัดสินใจ", "กลาง"),
    ("ตรวจสอบ", "เกณฑ์ส่วนต่างใบแจ้งหนี้ — เตือนที่ 2% บล็อกที่ 5%",
     "2% / 5%", "AC", "", "รอตัดสินใจ", "สูง"),
    ("ขั้นตอนงาน", "SLA แต่ละขั้นตอนควรเป็นเท่าไหร่จริง (ดูแท็บ «ขั้นตอนงาน»)",
     "24–336 ชม.", "หัวหน้าทุกแผนก", "", "รอตัดสินใจ", "สูง"),
    ("ขั้นตอนงาน", "flow 17 ขั้นตรงกับงานจริงไหม มีขั้นไหนหายหรือเกิน",
     "17 ขั้น", "หัวหน้าทุกแผนก", "", "รอตัดสินใจ", "สูง"),
    ("ขั้นตอนงาน", "ใครควรอัพโหลดใบแจ้งหนี้ — SR (ผู้รับจากผู้ขาย) หรือ AC (ผู้ใช้เอกสาร)",
     "AC", "SR + AC", "", "รอตัดสินใจ", "กลาง"),
    ("ขั้นตอนงาน", "ให้ผู้ขายส่งใบเสร็จตัวจริงถึง AC โดยตรงได้ไหม (ตัด SR ออกจริง)",
     "ส่งตรงถึง AC", "SR + AC", "", "รอตัดสินใจ", "กลาง"),
    ("เอกสาร", "Pattern label ใช้กับสินค้ากลุ่มไหน ต้อง track ในระบบไหม",
     "ไม่บังคับ", "SR + QC", "", "รอตัดสินใจ", "ต่ำ"),
    ("เอกสาร", "ตัด PO ที่ยังไม่เซ็นออกได้จริงไหม ไม่มีขั้นตอนไหนต้องใช้",
     "ตัดออก", "SR + AC", "", "รอตัดสินใจ", "กลาง"),
    ("เอกสาร", "ฟิลด์จริงของ «ส่วนที่ 3» ในฟอร์ม QC-F006 (หน้า 2–3 เป็นภาพสแกน)",
     "อนุมานเอง", "QC", "", "รอตัดสินใจ", "สูง"),
    ("เอกสาร", "ใบเคลมกับผู้ขาย ใช้ฟอร์มเดียวกับใบร้องเรียนลูกค้าหรือมีฟอร์มแยก",
     "ฟอร์มเดียวกัน", "QC + SR", "", "รอตัดสินใจ", "กลาง"),
    ("เอกสาร", "INSPECTION REPORT มีเลขที่เอกสารควบคุมไหม",
     "ยังไม่ทราบ", "QC", "", "รอตัดสินใจ", "ต่ำ"),
    ("เอกสาร", "ใบขอเปิดรหัสสินค้ามีฟอร์มควบคุมเดิมอยู่แล้วหรือเขียนอิสระ",
     "ยังไม่ทราบ", "SR + LS", "", "รอตัดสินใจ", "ต่ำ"),
    ("เอกสาร", "นโยบายเก็บเอกสารกี่ปี และใช้กับทุกประเภทเอกสารหรือไม่",
     "5 ปี", "AC + ที่ปรึกษากฎหมาย", "", "รอตัดสินใจ", "กลาง"),
    ("แจ้งเตือน", "lark_user_id ของพนักงานทุกคน",
     "ยังไม่มี", "IT + HR", "", "รอตัดสินใจ", "สูง"),
    ("แจ้งเตือน", "กลุ่ม Lark เดิม 3 กลุ่ม จะให้เหลือกลุ่มไหนรับแจ้งเตือนจากระบบ",
     "ยังไม่ทราบ", "หัวหน้าทุกแผนก", "", "รอตัดสินใจ", "กลาง"),
    ("แจ้งเตือน", "เวลาเงียบ 20:00–07:00 เหมาะสมไหม",
     "20:00–07:00", "หัวหน้าทุกแผนก", "", "รอตัดสินใจ", "ต่ำ"),
    ("ผู้ใช้/สิทธิ์", "ใครคือแอดมินระบบ (ต้องไม่อยู่ในสายอนุมัติ)",
     "ยังไม่ทราบ", "ผู้บริหาร", "", "รอตัดสินใจ", "สูง"),
    ("ผู้ใช้/สิทธิ์", "ใครรับเรื่องบั๊ก และ SLA รับเรื่องกี่ชั่วโมง",
     "8 ชม.ทำการ", "IT", "", "รอตัดสินใจ", "กลาง"),
    ("ผู้ใช้/สิทธิ์", "ตารางสิทธิ์ในแท็บ «สิทธิ์แต่ละแผนก» ตรงกับงานจริงไหม",
     "ตามที่ร่างไว้", "หัวหน้าทุกแผนก", "", "รอตัดสินใจ", "สูง"),
    ("ขอบเขต", "จะย้ายการเซ็นออกจาก Dochub จริงไหม หรือคงไว้สำหรับเอกสารภายนอก",
     "คง Dochub เฉพาะ PO ที่ผู้ขายเซ็น", "ผู้บริหาร + AC", "", "รอตัดสินใจ", "สูง"),
    ("ขอบเขต", "เริ่มที่งานซื้อในประเทศ (อาหาร) ก่อนแล้วค่อยขยาย ใช่ไหม",
     "ใช่", "ผู้บริหาร", "", "รอตัดสินใจ", "กลาง"),
    ("ขอบเขต", "ใครเป็นเจ้าของระบบหลังใช้งานจริง (ดูแลผู้ใช้ สิทธิ์ SLA)",
     "ยังไม่ทราบ", "ผู้บริหาร", "", "รอตัดสินใจ", "สูง"),
    ("ทดลองใช้", "ผลทดลองใช้ 15 นาทีจากแต่ละแผนก (ใช้แบบทดสอบในเอกสาร 10)",
     "ยังไม่ได้ทำ", "หัวหน้าทุกแผนก", "", "รอตัดสินใจ", "สูง"),
    ("โครงสร้าง", "ตาราง Price_Requests / Quotations ที่เพิ่งเพิ่ม ตรงกับวิธีขอราคาจริงไหม",
     "ตามที่ร่างไว้", "SR", "", "รอตัดสินใจ", "สูง"),
]
r = 5
for i, (cat, issue, default, who, ans, status, prio) in enumerate(ISSUES_LIST, start=1):
    cell(ws, r, 1, i, al="center")
    cell(ws, r, 2, cat, al="center")
    cell(ws, r, 3, issue)
    cell(ws, r, 4, default, bg=C_WARN_BG, fg=C_WARN_FG, al="center")
    cell(ws, r, 5, who, al="center")
    cell(ws, r, 6, ans, bg=C_EDIT)
    cell(ws, r, 7, status, bg=C_EDIT, al="center")
    cell(ws, r, 8, prio, bg=C_EDIT, al="center")
    ws.row_dimensions[r].height = 28
    r += 1
ISS_LAST = r - 1
dv(ws, "ISSUEST", "G5:G%d" % ISS_LAST)
dv(ws, "PRIO", "H5:H%d" % ISS_LAST)
ws.freeze_panes = "C5"
ws.auto_filter.ref = "A4:H%d" % ISS_LAST

r += 1
for lab, formula in [
    ("รวมประเด็นทั้งหมด", "=COUNTA(C5:C%d)" % ISS_LAST),
    ("รอตัดสินใจ", '=COUNTIF(G5:G%d,"รอตัดสินใจ")' % ISS_LAST),
    ("ตัดสินใจแล้ว", '=COUNTIF(G5:G%d,"ตัดสินใจแล้ว")' % ISS_LAST),
    ("ความสำคัญสูง ที่ยังรอตัดสินใจ", '=COUNTIFS(H5:H%d,"สูง",G5:G%d,"รอตัดสินใจ")' % (ISS_LAST, ISS_LAST)),
]:
    cell(ws, r, 2, lab, bold=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    cell(ws, r, 4, formula, bg=C_LOCK, bold=True, al="center")
    r += 1

r += 1
c = cell(ws, r, 2, "ทุกค่าในคอลัมน์ «ค่าที่ผมตั้งไว้ก่อน» เป็นค่าที่ผมเลือกเองเพื่อให้ระบบเดินได้ "
                    "ไม่ได้มาจากนโยบายของบริษัท — ถ้าไม่แก้ ระบบจะใช้ค่าเหล่านี้จริงตอนขึ้นใช้งาน",
         bg=C_WARN_BG, fg=C_WARN_FG, bold=True)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
ws.row_dimensions[r].height = 32

# ---------- finish ----------
# ตรวจว่าจำนวนแถวพจนานุกรมตรงกับที่ใช้ในสูตรจริง
assert DICT_LAST_ROW == 4 + len(D), "DICT_LAST_ROW ไม่ตรงกับจำนวนแถวจริง: %d vs %d" % (DICT_LAST_ROW, 4 + len(D))
assert DICT_LAST == DICT_LAST_ROW, "แถวสุดท้ายของพจนานุกรมไม่ตรง: %d vs %d" % (DICT_LAST, DICT_LAST_ROW)

# เขียนสูตรสรุปหน้าแรกใหม่ด้วยช่วงที่คำนวณจากข้อมูลจริง — กันช่วงกินแถวหมายเหตุหรือขาดแถว
TBL_LAST   = 4 + len(TABLES)
STAGE_LAST = 4 + len(STAGES)
DOC_LAST   = 4 + len(DOCS)
FINAL_STATS = [
    ("จำนวนตารางทั้งหมด",        '=COUNTA(สารบัญตาราง!$B$5:$B$%d)' % TBL_LAST),
    ("จำนวนคอลัมน์ทั้งหมด",      '=COUNTA(พจนานุกรมข้อมูล!$B$5:$B$%d)' % DICT_LAST),
    ("ขั้นตอนงาน",               '=COUNTA(ขั้นตอนงาน!$B$5:$B$%d)' % STAGE_LAST),
    ("เอกสารบังคับต่อ PO",       '=COUNTIF(เอกสารที่ต้องมี!$E$5:$E$%d,"บังคับ")' % DOC_LAST),
    ("กฎตรวจสอบ",               '=COUNTA(กฎตรวจสอบ!$A$5:$A$%d)' % RULE_LAST),
    ("ค่าตั้งต้นที่ปรับได้",      '=COUNTA(ค่าตั้งต้น!$A$5:$A$%d)' % CFG_LAST),
    ("ประเด็นที่ยังรอตัดสินใจ",   '=COUNTIF(ประเด็นค้าง!$G$5:$G$%d,"รอตัดสินใจ")' % ISS_LAST),
    ("ประเด็นที่ตัดสินใจแล้ว",    '=COUNTIF(ประเด็นค้าง!$G$5:$G$%d,"ตัดสินใจแล้ว")' % ISS_LAST),
]
assert [l for l, _ in FINAL_STATS] == SUMMARY_LABELS, "ลำดับตัวเลขสรุปไม่ตรงกับที่เขียนไว้ตอนแรก"
for i, (lab, formula) in enumerate(FINAL_STATS):
    rr = SUMMARY_ROW + i
    assert START_SHEET.cell(row=rr, column=1).value == lab, "แถวสรุปเลื่อน: %s" % lab
    x = START_SHEET.cell(row=rr, column=2)
    x.value = formula

# openpyxl ไม่เก็บค่าที่คำนวณไว้ — บังคับให้คำนวณตอนเปิดไฟล์
wb.calculation.fullCalcOnLoad = True

lst.sheet_state = "hidden"
wb._sheets.insert(0, wb._sheets.pop(wb.sheetnames.index("เริ่มที่นี่")))
for s in wb.worksheets:
    s.sheet_properties.tabColor = "1F3D33"
wb["ประเด็นค้าง"].sheet_properties.tabColor = "A02B30"
wb["เริ่มที่นี่"].sheet_properties.tabColor = "96731C"

wb.save(OUT)
print("saved", OUT)
print("sheets:", wb.sheetnames)
print("dict rows:", len(D))
