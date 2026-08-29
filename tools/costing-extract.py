#!/usr/bin/env python3
"""สกัดข้อมูลจากไฟล์ Costing ที่ออกจาก SAP เพื่อขึ้น Dashboard ตั้งเบิกทำจ่าย

ไฟล์ Costing เป็น **รายงานที่ออกจาก SAP** ไม่ใช่ไฟล์ที่คนพิมพ์เอง
และบริษัทมี PO ทั้งฝั่ง **Mech** และ **Food** จึงมีไฟล์คนละใบ กฎเดียวกัน

รันด้วย:
  python3 tools/costing-extract.py "Costing Mech.xlsx"                      # เดาสายจากชื่อไฟล์
  python3 tools/costing-extract.py "Costing Mech.xlsx" "Costing Food.xlsx"  # รวมหลายไฟล์
  python3 tools/costing-extract.py a.xlsx -m MECH -o out.json               # ระบุสายเอง

กติกาตามที่ทีมกำหนด
  1. อ่านชีตแรกสุดเท่านั้น
  2. PO Payment Term ว่าง -> "UNKNOWN"  (กันระบบพัง ไม่ใช่ทิ้งแถว)
  3. ตัดแถวที่ Payment Term มี LC / L/C ออก (ไม่สนตัวพิมพ์)
  4. ตรวจ 4 ฟิลด์  PO Number · Supplier · Price · Due Date  -> Payment_Status
  5. ส่งออก 7 คอลัมน์ + ธงความเสี่ยงยอดศูนย์

หลักที่ยึด: ไฟล์เพี้ยนได้ แต่สคริปต์ต้องไม่ตาย — คอลัมน์หายก็บอกว่าหายอะไร
ไม่ใช่ traceback ยาวเป็นหน้า
"""
import argparse, json, os, re, sys
from collections import Counter, OrderedDict

try:
    import openpyxl
except ImportError:
    sys.exit("ต้องติดตั้ง openpyxl ก่อน:  pip install openpyxl")

# คอลัมน์ที่ต้องใช้ — ชื่อทางเลือกเผื่อไฟล์รอบหน้าเปลี่ยนหัวคอลัมน์เล็กน้อย
COLS = OrderedDict([
    ("PO Number",       ["po number", "po no", "po_no", "เลขที่ po"]),
    ("Supplier",        ["supplier", "vendor", "ผู้ขาย"]),
    ("PO Payment Term", ["po payment term", "payment term", "term", "เงื่อนไขชำระ"]),
    ("Price",           ["price", "unit price", "ราคา"]),
    ("Currency",        ["currency", "cur", "สกุลเงิน"]),
    ("Due Date",        ["due date", "duedate", "วันครบกำหนด"]),
])
KEY_FIELDS = ["PO Number", "Supplier", "Price", "Due Date"]   # 4 ฟิลด์ที่ตัดสินความครบ
LC_RE = re.compile(r"l\s*/?\s*c", re.IGNORECASE)              # จับ LC · L/C · l / c
UNKNOWN = "UNKNOWN"

# สายสินค้า — บริษัทมี PO ทั้งฝั่ง Mech และ Food ไฟล์คนละใบ กฎเดียวกัน
# OTHER มาจากของจริง: ไฟล์ Mech มีแถว PO-O3… (ค่าเดินทาง ค่าบริการ) ปนอยู่ด้วย
MODULES = {"MECH": "เครื่องจักร & โซลาร์", "FOOD": "อาหาร", "OTHER": "ค่าใช้จ่ายอื่น"}
# เลข PO แต่ละชุดขึ้นต้นต่างกัน (จากไฟล์จริงและใบตรวจ QC: PO-M2… / PO-F1… / PO-O3…)
PO_PREFIX = {"M": "MECH", "F": "FOOD", "O": "OTHER"}


def guess_module(path, rows, idx, explicit=""):
    """หาสายสินค้าตามลำดับ: ที่สั่งมา -> เลข PO ในไฟล์ -> ชื่อไฟล์
    เดาไม่ได้ก็คืน UNSET ไม่ใช่เดามั่ว — เอาไปแสดงให้คนเลือกดีกว่าลงผิดสาย"""
    if explicit.upper() in MODULES:
        return explicit.upper()
    # ดูจากเลข PO จริงในไฟล์ ตัวไหนมากสุดชนะ (แม่นกว่าชื่อไฟล์ที่คนตั้งเอง)
    votes = Counter()
    if "PO Number" in idx:
        for r in rows[1:]:
            v = str(r[idx["PO Number"]] or "").strip().upper()
            m = re.match(r"^PO-([A-Z])", v)
            if m and m.group(1) not in PO_PREFIX:
                continue
            if m:
                votes[PO_PREFIX[m.group(1)]] += 1
    if votes:
        return votes.most_common(1)[0][0]
    name = norm(path)
    for k in MODULES:
        if k.lower() in name:
            return k
    return "UNSET"


def norm(s):
    return re.sub(r"\s+", " ", str(s or "")).strip().lower()


def is_blank(v):
    """ว่างจริง ๆ เท่านั้น — เลข 0 ไม่ใช่ค่าว่าง (ยอดศูนย์เป็นคนละเรื่อง ดู zero_price)"""
    return v is None or str(v).strip() == "" or norm(v) in ("nan", "none", "null", "-")


def map_columns(header):
    """จับคู่หัวคอลัมน์จริงกับที่ต้องใช้ · เจอชื่อซ้ำให้ใช้คอลัมน์แรก (ไฟล์นี้มี Import duty ซ้ำ)"""
    seen, idx, missing = {}, {}, []
    for i, h in enumerate(header):
        k = norm(h)
        if k and k not in seen:
            seen[k] = i
    for want, aliases in COLS.items():
        for a in [norm(want)] + aliases:
            if a in seen:
                idx[want] = seen[a]
                break
        else:
            missing.append(want)
    return idx, missing


def parse_due(v):
    """คืน ISO ไว้เรียงลำดับ — ไฟล์นี้เขียนแบบ dd.mm.yyyy · รองรับ / และ - ด้วย
    แปลงไม่ได้ก็คืนค่าว่าง ไม่ใช่ทำให้ทั้งแถวตก"""
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    m = re.match(r"^\s*(\d{1,2})[./-](\d{1,2})[./-](\d{4})\s*$", str(v or ""))
    if m:
        d, mo, y = m.groups()
        return "%s-%02d-%02d" % (y, int(mo), int(d))
    m = re.match(r"^\s*(\d{4})-(\d{1,2})-(\d{1,2})", str(v or ""))
    if m:
        y, mo, d = m.groups()
        return "%s-%02d-%02d" % (y, int(mo), int(d))
    return ""


def to_num(v):
    try:
        return float(str(v).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def extract(path, module=""):
    if not os.path.exists(path):
        return {"error": "ไม่พบไฟล์ %s — ตรวจชื่อไฟล์และที่อยู่อีกครั้ง" % path}
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        # ไฟล์เสียหรือไม่ใช่ .xlsx ก็ต้องบอกเป็นภาษาคน ไม่ใช่ traceback
        return {"error": "เปิดไฟล์ %s ไม่ได้ (%s) — ต้องเป็นไฟล์ .xlsx ที่ไม่เสียหาย"
                         % (path, type(e).__name__)}
    ws = wb[wb.sheetnames[0]]                       # กติกาข้อ 1 — ชีตแรกสุด
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"error": "ชีตแรกไม่มีข้อมูลเลย", "sheet": ws.title}

    idx, missing = map_columns(rows[0])
    if missing:
        return {"error": "ไม่พบคอลัมน์ที่ต้องใช้: " + ", ".join(missing),
                "sheet": ws.title, "headers": [str(h) for h in rows[0] if h]}

    mod = guess_module(path, rows, idx, module)

    get = lambda r, k: r[idx[k]] if idx[k] < len(r) else None
    out, terms = [], Counter()
    # ตั้งค่าเริ่มต้นเป็น 0 ทุกตัว — Counter ปกติจะไม่มีคีย์ที่นับได้ 0
    # ทำให้ผลที่ส่งออกขาดคีย์ไปเฉย ๆ และเทียบกับฝั่ง Apps Script ไม่ได้
    stats = Counter({"read": 0, "excluded_lc": 0, "complete": 0, "incomplete": 0,
                     "unknown_term": 0, "zero_price": 0})

    for n, r in enumerate(rows[1:], start=2):
        if all(is_blank(c) for c in r):             # แถวว่างล้วน = ท้ายตาราง ข้ามไป
            continue
        stats["read"] += 1

        # กติกาข้อ 2 — ว่างแปลว่า UNKNOWN ไม่ใช่แปลว่าทิ้ง
        term_raw = get(r, "PO Payment Term")
        term = UNKNOWN if is_blank(term_raw) else str(term_raw).strip()
        if term == UNKNOWN:
            stats["unknown_term"] += 1
        terms[term] += 1

        # กติกาข้อ 3 — LC จ่ายผ่านธนาคาร ไม่เข้ากระบวนการตั้งเบิกนี้
        if LC_RE.search(term):
            stats["excluded_lc"] += 1
            continue

        # กติกาข้อ 4 — ครบ 4 ฟิลด์หรือไม่ และถ้าไม่ครบ ขาดอะไรบ้าง
        miss = [f for f in KEY_FIELDS if is_blank(get(r, f))]
        status = ("🟢 พร้อมทำจ่าย (Complete)" if not miss
                  else "🔴 ข้อมูลไม่ครบ (" + ", ".join(miss) + ")")
        stats["complete" if not miss else "incomplete"] += 1

        price = to_num(get(r, "Price"))
        zero = price is not None and price == 0        # มุมมองความเสี่ยง — กันตั้งเบิกยอดศูนย์
        if zero:
            stats["zero_price"] += 1

        out.append({
            "row":             n,
            "Module":          mod,
            "PO Number":       "" if is_blank(get(r, "PO Number")) else str(get(r, "PO Number")).strip(),
            "Supplier":        "" if is_blank(get(r, "Supplier")) else str(get(r, "Supplier")).strip(),
            "PO Payment Term": term,
            "Price":           price,
            "Currency":        "" if is_blank(get(r, "Currency")) else str(get(r, "Currency")).strip(),
            "Due Date":        "" if is_blank(get(r, "Due Date")) else str(get(r, "Due Date")).strip(),
            "Due_ISO":         parse_due(get(r, "Due Date")),
            "Payment_Status":  status,
            "Missing":         miss,
            "Zero_Price":      zero,
        })

    return {"sheet": ws.title, "module": mod, "file": path, "rows": out,
            "stats": dict(stats), "terms": terms.most_common()}


def dump_grid(path, out):
    """ตารางดิบ ไม่ผ่านกฎใด ๆ — ให้ฝั่ง Apps Script รับไปประมวลผลเองแล้วเทียบผลกัน"""
    import datetime
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    cell = lambda c: "" if c is None else (
        c.isoformat() if isinstance(c, (datetime.date, datetime.datetime)) else c)
    grid = [[cell(c) for c in row] for row in ws.iter_rows(values_only=True)]
    with open(out, "w", encoding="utf-8") as f:
        json.dump(grid, f, ensure_ascii=False)
    print("เขียนตารางดิบ %d แถว ที่ %s" % (len(grid), out))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", nargs="+", help="ไฟล์ Costing หนึ่งไฟล์ขึ้นไป (Mech / Food)")
    ap.add_argument("-m", "--module", default="",
                    help="ระบุสายสินค้าเอง MECH หรือ FOOD · ไม่ระบุ = เดาจากเลข PO ในไฟล์")
    ap.add_argument("-o", "--out", default="")
    ap.add_argument("--grid", default="",
                    help="ส่งออกตารางดิบของไฟล์แรกเป็น JSON — ใช้เทียบผลกับฝั่ง Apps Script")
    a = ap.parse_args()

    if a.grid:
        if os.path.exists(a.xlsx[0]):
            dump_grid(a.xlsx[0], a.grid)
        else:
            print("ข้ามการส่งออกตารางดิบ — ไม่พบไฟล์ " + a.xlsx[0])

    all_rows, all_stats, files, bad = [], Counter(), [], 0
    for path in a.xlsx:
        res = extract(path, a.module)
        if "error" in res:
            print("ข้ามไฟล์ %s — %s" % (path, res["error"]))
            if res.get("headers"):
                print("  หัวคอลัมน์ที่เจอ: " + " | ".join(res["headers"][:30]))
            bad += 1
            continue
        s = res["stats"]
        files.append({"file": path, "sheet": res["sheet"], "module": res["module"],
                      "rows": len(res["rows"]), "stats": s})
        all_rows += res["rows"]
        all_stats.update(s)
        print("%-6s %-40s ชีต %-10s อ่าน %3d | ตัด LC %3d | เหลือ %3d | 🟢 %3d 🔴 %d"
              % (res["module"], path.split("/")[-1][:40], res["sheet"], s.get("read", 0),
                 s.get("excluded_lc", 0), len(res["rows"]),
                 s.get("complete", 0), s.get("incomplete", 0)))
        if res["module"] == "UNSET":
            print("  เตือน: เดาสายสินค้าไม่ได้จากเลข PO หรือชื่อไฟล์ — ระบุด้วย -m MECH|FOOD")

    if bad and not all_rows:
        sys.exit(1)

    t = all_stats
    print("-" * 72)
    print("รวมทุกไฟล์ %d ไฟล์ | อ่านมา %d แถว | ตัด LC ออก %d | เหลือเข้ากระบวนการ %d"
          % (len(files), t.get("read", 0), t.get("excluded_lc", 0), len(all_rows)))
    print("  🟢 พร้อมทำจ่าย %d   🔴 ข้อมูลไม่ครบ %d"
          % (t.get("complete", 0), t.get("incomplete", 0)))
    print("  เงื่อนไขจ่ายว่าง (นับเป็น UNKNOWN) %d | ราคาเป็นศูนย์ %d"
          % (t.get("unknown_term", 0), t.get("zero_price", 0)))

    if a.out:
        with open(a.out, "w", encoding="utf-8") as f:
            json.dump({"files": files, "rows": all_rows, "stats": dict(t)},
                      f, ensure_ascii=False, indent=1)
        print("เขียนผลลัพธ์ที่ " + a.out)


if __name__ == "__main__":
    main()
